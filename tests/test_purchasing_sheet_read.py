# tests/test_purchasing_sheet_read.py
"""Reading a generated purchasing workbook back in.

The headline test is a round trip: build a sheet with build_purchasing_sheet,
read it with read_purchasing_sheet, and check the rows survive. That is the
whole reason the reader lives beside the writer — it fails the day a column
moves.
"""
import os
import sys

import openpyxl
import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import bom_purchasing as bp  # noqa: E402

BUY = {"Number": "SF-000067", "Title": "SF-000067", "Row Order": "1",
       "Item Qty": 2, "Units": "Each", "Description (Item,CO)": "pull handle",
       "Source": "Buy", "Vendor": "Acme", "Cost Per": 1.5,
       "Lead Time (Business Days)": 5}
MAKE = {"Number": "CD-001200", "Title": "CD-001200", "Row Order": "2",
        "Item Qty": 1, "Units": "Each", "Description (Item,CO)": "adapter plate",
        "Source": "Make", "Vendor": "Machine Shop", "Cost Per": 40.0,
        "Lead Time (Business Days)": 15}


def _df(rows):
    df, err = bp.coerce_bom_dataframe(pd.DataFrame(rows))
    assert err is None, err
    return df


def _sheet(tmp_path, rows, assembly="CD-001608"):
    """Write a purchasing workbook and return its path."""
    out = tmp_path / "CD-001608 Purchasing Sheet.xlsx"
    bp.build_purchasing_sheet(_df(rows), str(out), assembly)
    return str(out)


def test_round_trip_preserves_the_rows(tmp_path):
    path = _sheet(tmp_path, [BUY, MAKE])
    df, assembly, error = bp.read_purchasing_sheet(path)

    assert error is None
    assert assembly == "CD-001608"
    assert len(df) == 2

    by_title = {r["Title"]: r for _i, r in df.iterrows()}
    assert set(by_title) == {"SF-000067", "CD-001200"}

    make = by_title["CD-001200"]
    assert make["Source"] == "Make"
    assert make["Vendor"] == "Machine Shop"
    assert make["Description (Item,CO)"] == "adapter plate"
    assert float(make["Item Qty"]) == 1.0
    assert float(make["Cost Per"]) == 40.0
    assert float(make["Lead Time (Business Days)"]) == 15.0


def test_the_canonical_column_names_come_back(tmp_path):
    """The sheet heads them "Name" and "Description"; the reader restores the
    internal names the rest of the codebase keys on."""
    path = _sheet(tmp_path, [BUY])
    df, _assembly, error = bp.read_purchasing_sheet(path)

    assert error is None
    assert "Title" in df.columns
    assert "Description (Item,CO)" in df.columns
    assert "Name" not in df.columns


def test_a_workbook_with_no_purchasing_tab_is_refused(tmp_path):
    """A raw Inventor export looks exactly like this. The message has to name
    the tab, so the remedy is obvious without reading the code."""
    path = tmp_path / "CD-001608 BOM.xlsx"
    pd.DataFrame([{"Item": "1", "Filename": "CD-001612.ipt"}]).to_excel(
        path, index=False)

    df, assembly, error = bp.read_purchasing_sheet(str(path))

    assert error is not None
    assert "Purchasing" in error
    assert df.empty
    assert assembly == ""


def test_a_missing_file_is_refused_not_raised(tmp_path):
    df, _assembly, error = bp.read_purchasing_sheet(str(tmp_path / "nope.xlsx"))
    assert error is not None
    assert "not found" in error.lower()
    assert df.empty


def test_the_header_row_is_found_when_it_is_not_at_hdr_row(tmp_path):
    """A sheet generated before a future layout change still reads."""
    path = tmp_path / "shifted.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = bp.PURCHASING_SHEET_NAME
    ws["A1"] = "CD-009999"
    headers = [bp.HEADER_LABELS.get(c, c) for c in bp.SHEET_COLUMNS]
    ws.append([])                      # push the header down one extra row
    for offset, name in enumerate(headers, 1):
        ws.cell(row=bp.HDR_ROW + 1, column=offset, value=name)
    row = {"Title": "CD-001200", "Source": "Make", "Vendor": "Acme"}
    for offset, col in enumerate(bp.SHEET_COLUMNS, 1):
        ws.cell(row=bp.HDR_ROW + 2, column=offset, value=row.get(col))
    wb.save(path)

    df, assembly, error = bp.read_purchasing_sheet(str(path))

    assert error is None
    assert assembly == "CD-009999"
    assert list(df["Title"]) == ["CD-001200"]


def test_a_duplicated_column_is_refused_rather_than_silently_dropped(tmp_path):
    """These workbooks are hand-edited — filling in suppliers is the whole
    workflow — so a copy-pasted column is a plausible accident. Keeping only
    the last one would lose a supplier column with nothing said."""
    path = tmp_path / "duplicated.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = bp.PURCHASING_SHEET_NAME
    ws["A1"] = "CD-001608"
    headers = [bp.HEADER_LABELS.get(c, c) for c in bp.SHEET_COLUMNS]
    headers.append("Vendor")           # the accidental duplicate
    for offset, name in enumerate(headers, 1):
        ws.cell(row=bp.HDR_ROW, column=offset, value=name)
    wb.save(path)

    df, _assembly, error = bp.read_purchasing_sheet(str(path))

    assert error is not None
    assert "Vendor" in error
    assert df.empty


def test_the_unmatched_note_never_becomes_a_row(tmp_path):
    """Against a real generated workbook: an unpriced Buy row makes the
    writer append an "Unmatched (n)" note below the table, always preceded by
    a fully blank row. This proves the reader stops cleanly at that blank row
    before it ever reaches the note — not that the UNMATCHED_NOTE_PREFIX
    guard itself fires, since the blank-row break always wins first here. See
    test_the_unmatched_note_guard_holds_even_without_a_blank_row for that."""
    unpriced = dict(BUY, Vendor=None, **{"Cost Per": None})
    path = _sheet(tmp_path, [unpriced, MAKE])

    df, _assembly, error = bp.read_purchasing_sheet(path)

    assert error is None
    assert len(df) == 2
    assert not any(str(t).startswith(bp.UNMATCHED_NOTE_PREFIX)
                   for t in df["Title"])


def test_the_unmatched_note_guard_holds_even_without_a_blank_row(tmp_path):
    """The real writer always inserts a blank row before the note, so the
    blank-row break above the UNMATCHED_NOTE_PREFIX check always fires first
    against real output — the prefix check itself is never exercised there.
    Hand-build a workbook with the note placed immediately after the data,
    with no blank row between, so this test actually depends on the guard."""
    path = tmp_path / "no_blank_row.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = bp.PURCHASING_SHEET_NAME
    ws["A1"] = "CD-001608"
    headers = [bp.HEADER_LABELS.get(c, c) for c in bp.SHEET_COLUMNS]
    for offset, name in enumerate(headers, 1):
        ws.cell(row=bp.HDR_ROW, column=offset, value=name)
    row = {"Title": "CD-001200", "Source": "Make", "Vendor": "Acme"}
    for offset, col in enumerate(bp.SHEET_COLUMNS, 1):
        ws.cell(row=bp.HDR_ROW + 1, column=offset, value=row.get(col))

    note_row = bp.HDR_ROW + 2  # immediately after the data — no blank row
    n_cols = len(bp.SHEET_COLUMNS)
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=n_cols)
    ws.cell(row=note_row, column=1,
            value=f"{bp.UNMATCHED_NOTE_PREFIX}1) — no price in reference: SF-000067")
    wb.save(path)

    df, _assembly, error = bp.read_purchasing_sheet(str(path))

    assert error is None
    assert len(df) == 1
    assert not any(str(t).startswith(bp.UNMATCHED_NOTE_PREFIX)
                   for t in df["Title"])


def test_sub_total_is_not_readable_and_that_is_expected(tmp_path):
    """Sub Total is an Excel formula. openpyxl with data_only=True returns
    None unless Excel has opened and re-saved the file, so callers recompute
    the line total rather than read this column."""
    path = _sheet(tmp_path, [MAKE])
    df, _assembly, error = bp.read_purchasing_sheet(path)

    assert error is None
    assert df.loc[0, "Sub Total"] is None
