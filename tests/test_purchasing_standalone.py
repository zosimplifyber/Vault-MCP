import ast
import os
import sys

import openpyxl
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import bom_purchasing as bp  # noqa: E402


def _ref_workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "purchased parts"
    ws.append(["Number", "Vendor", "Cost Per"])
    ws.append(["SF-1", "Acme", 3.5])
    wb.save(path)


def test_reference_path_override_is_used(tmp_path, monkeypatch):
    # Auto-discovery finds nothing, so only the override can supply vendor data.
    monkeypatch.setattr(bp, "find_purchased_items_file", lambda: None)
    ref = tmp_path / "ref.xlsx"
    _ref_workbook(ref)
    bom = tmp_path / "bom.txt"
    bom.write_text(              # the reference is keyed on the file name
        "Item\tPart Number\tBOM Structure\tQTY\tDescription\tFilename\n"
        "1\tSF-999\tPurchased\t2\tpart\tSF-1.ipt\n",
        encoding="utf-8",
    )
    result = bp.generate_from_file(str(bom), "ASM", str(tmp_path), reference_path=str(ref))
    assert not result.get("error"), result
    ws = openpyxl.load_workbook(result["output_path"])["Purchasing"]
    header = [c.value for c in ws[3]]
    # Number is not a sheet column any more — key the row by its Description.
    desc, ven = header.index("Description") + 1, header.index("Vendor") + 1
    vendors = {ws.cell(r, desc).value: ws.cell(r, ven).value for r in range(4, 5)}
    assert vendors.get("part") == "Acme"   # from the override, not OneDrive


def test_default_source_is_mslist_only_no_excel_autofind(tmp_path, monkeypatch):
    # Default source is Microsoft-List-only: without a sign-in, generation still
    # succeeds (costs blank) and must NOT fall back to the Excel file.
    import purchasing_reference as pref
    monkeypatch.setattr(pref, "resolve_reference_config", lambda override=None: {
        "source": "mslist",
        "mslist": {"tenant_id": "T", "client_id": "C"},
        "column_map": {},
    })
    monkeypatch.setattr(pref, "load_mslist_dataframe",
                        lambda *a, **k: (_ for _ in ()).throw(RuntimeError("not signed in")))
    called = {"find": False}
    monkeypatch.setattr(bp, "find_purchased_items_file",
                        lambda: called.__setitem__("find", True) or None)
    bom = tmp_path / "bom.txt"
    bom.write_text("Item\tPart Number\tQTY\tDescription\n1\tSF-1\t2\tp\n", encoding="utf-8")
    result = bp.generate_from_file(str(bom), "ASM", str(tmp_path))
    assert not result.get("error"), result
    assert called["find"] is False   # List-only default: no Excel auto-discovery


def test_standalone_does_not_import_app_or_vault():
    src = open(os.path.join(ROOT, "purchasing_standalone.py"), encoding="utf-8").read()
    mods = set()
    for n in ast.walk(ast.parse(src)):
        if isinstance(n, ast.Import):
            mods |= {a.name.split(".")[0] for a in n.names}
        elif isinstance(n, ast.ImportFrom) and n.module:
            mods.add(n.module.split(".")[0])
    assert not (mods & {"app", "mcp_server", "vault_rest_api", "gui"}), mods
    assert "bom_purchasing" in mods


def test_standalone_gui_constructs(monkeypatch):
    tk = pytest.importorskip("tkinter")
    monkeypatch.setattr(bp, "find_purchased_items_file", lambda: None)
    import purchasing_standalone as ps
    try:
        app = ps.App()
    except tk.TclError:
        pytest.skip("no display available")
    app.withdraw()
    app.update_idletasks()
    assert "Simplifyber" in app.title()
    app.destroy()
