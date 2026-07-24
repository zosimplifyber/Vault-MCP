# tests/test_bom_purchasing_inventor.py
import os
import sys

import pandas as pd
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import bom_purchasing as bp  # noqa: E402


def test_coerce_maps_inventor_headers_and_translates_source():
    df = pd.DataFrame({
        "Item": ["1", "2", "2.1"],
        "Part Number": ["SF-001580", "SF-001803", "SF-001885"],
        "BOM Structure": ["Normal", "Normal", "Purchased"],
        "Unit QTY": ["Each", "Each", "Each"],
        "QTY": [2, 1, 8],
        "Description": ["adapter plate", "bladder tool", "hex screw"],
        "REV": ["", "1", "1"],
        "Material": ["Aluminum", "", "Stainless Steel"],
        "Material Finish": ["", "", "Black Oxide"],
    })
    out, err = bp.coerce_bom_dataframe(df)
    assert err is None
    assert list(out["Number"]) == ["SF-001580", "SF-001803", "SF-001885"]
    assert list(out["Item Qty"]) == [2, 1, 8]
    assert list(out["Row Order"]) == ["1", "2", "2.1"]
    assert list(out["Source"]) == ["Make", "Make", "Buy"]      # translated
    assert list(out["Units"]) == ["Each", "Each", "Each"]
    assert out["Material Finish"].tolist() == ["", "", "Black Oxide"]
    for col in ("Category Name", "State", "Position Number", "Title (Item,CO)"):
        assert col in out.columns


def test_coerce_accepts_canonical_vault_headers_unchanged():
    df = pd.DataFrame({c: ["x"] for c in bp.BOM_COLUMNS})
    out, err = bp.coerce_bom_dataframe(df)
    assert err is None
    assert set(bp.BOM_COLUMNS).issubset(out.columns)


def test_coerce_errors_when_no_part_number():
    df = pd.DataFrame({"QTY": [1, 2], "Description": ["a", "b"]})
    out, err = bp.coerce_bom_dataframe(df)
    assert err is not None and "part number" in err.lower()


def test_coerce_errors_when_number_mixes_none_and_blank():
    df = pd.DataFrame({
        "Part Number": [None, "", None],
        "QTY": [1, 2, 3],
        "Description": ["a", "b", "c"],
    })
    out, err = bp.coerce_bom_dataframe(df)
    assert err is not None and "part number" in err.lower()


def test_coerce_does_not_crash_on_both_number_and_part_number():
    # Inventor branch triggers (missing most canonical cols) but Number already
    # exists alongside its synonym Part Number -> must not create a duplicate col.
    df = pd.DataFrame({
        "Number": ["SF-1", "SF-2"],
        "Part Number": ["X-1", "X-2"],
        "QTY": [1, 2],
        "Description": ["a", "b"],
    })
    out, err = bp.coerce_bom_dataframe(df)
    assert err is None
    assert list(out["Number"]) == ["SF-1", "SF-2"]      # canonical wins
    assert list(out["Item Qty"]) == [1, 2]


def test_coerce_errors_when_all_quantities_blank():
    df = pd.DataFrame({
        "Part Number": ["A", "B"],
        "QTY": ["", ""],
        "Description": ["a", "b"],
    })
    out, err = bp.coerce_bom_dataframe(df)
    assert err is not None and "quantit" in err.lower()


def test_read_bom_file_reads_tab_delimited_txt(tmp_path):
    p = tmp_path / "bom.txt"
    p.write_text(
        "Item\tPart Number\tQTY\tDescription\n"
        "1\tSF-001580\t2\tadapter plate\n"
        "2.1\tSF-001885\t8\thex screw\n",
        encoding="utf-8",
    )
    df = bp.read_bom_file(str(p))
    assert list(df.columns[:4]) == ["Item", "Part Number", "QTY", "Description"]
    assert len(df) == 2
    assert str(df.iloc[1]["Part Number"]) == "SF-001885"


def test_generate_from_file_inventor_export_populates(tmp_path, monkeypatch):
    p = tmp_path / "bom.txt"
    p.write_text(
        "Item\tPart Number\tBOM Structure\tUnit QTY\tQTY\tDescription\tREV\tMaterial\tMaterial Finish\n"
        "1\tSF-001580\tNormal\tEach\t2\tadapter plate\t\tAluminum\t\n"
        "2\tSF-001803\tNormal\tEach\t1\tbladder tool\t1\t\t\n"
        "2.1\tSF-001885\tPurchased\tEach\t8\thex screw\t1\tSteel\tBlack Oxide\n",
        encoding="utf-8",
    )

    # Reference "enrichment" that would overwrite Material — prove export wins.
    def fake_enrich(df, reference_path=""):
        df = df.copy()
        df["Material"] = "REF-MATERIAL"
        df["Vendor"] = "Acme"
        df["Cost Per"] = 1.5
        return df, 1, 1, [], []
    monkeypatch.setattr(bp, "_enrich_with_reference", fake_enrich)

    out_dir = tmp_path / "out"
    result = bp.generate_from_file(str(p), "CD-001608", str(out_dir))
    assert not result.get("error"), result
    assert os.path.isfile(result["output_path"])

    import openpyxl
    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb["Purchasing"]
    header = [c.value for c in ws[3]]
    mat_col = header.index("Material") + 1
    num_col = header.index("Number") + 1
    mats = {ws.cell(r, num_col).value: ws.cell(r, mat_col).value
            for r in range(4, 4 + 3)}
    assert mats["SF-001580"] == "Aluminum"       # from export
    assert mats["SF-001803"] == "REF-MATERIAL"   # blank in export → ref


def _hier_df():
    # 2 is an assembly (has child 2.1); 1 is a leaf part.
    return pd.DataFrame({
        "Number": ["SF-1", "SF-2", "SF-21"],
        "Row Order": ["1", "2", "2.1"],
        "Position Number": [None, None, None],
        "Item Qty": [2, 1, 8],
        "Units": ["Each", "Each", "Each"],
        "Category Name": [None, None, None],   # Inventor exports have none
        "Revision": [None, "1", "1"],
        "State": [None, None, None],
        "Title (Item,CO)": [None, None, None],
        "Description (Item,CO)": ["leaf", "assy", "screw"],
        "Source": ["Buy", "Make", "Buy"],
        "Material": ["Al", None, "Steel"],
        "Material Finish": [None, None, "Black Oxide"],
        "Vendor": ["Acme", None, "Bolts Inc"],
        "Cost Per": [1.0, None, 0.25],
        "Vendor Number": [None, None, None],
        "HS/HTS Code": [None, None, None],
        "Shipping": [None, None, None],
        "Tax/Tariff": [None, None, None],
        "Lead Time (Business Days)": [None, None, None],
    })


def test_material_finish_is_a_column_and_populates(tmp_path):
    import openpyxl
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(_hier_df(), str(out), "ASM")
    ws = openpyxl.load_workbook(str(out))["Purchasing"]
    header = [c.value for c in ws[3]]
    assert "Material Finish" in header
    mf_col = header.index("Material Finish") + 1
    assert ws.cell(6, mf_col).value == "Black Oxide"  # row 4+2 = the screw


def test_assembly_detected_by_children_without_category(tmp_path):
    import openpyxl
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(_hier_df(), str(out), "ASM")
    ws = openpyxl.load_workbook(str(out))["Purchasing"]
    header = [c.value for c in ws[3]]
    cost_col = header.index("Cost Per") + 1
    # Row 5 is "2" (assembly) → Cost Per is a SUM formula over its child rows.
    assert str(ws.cell(5, cost_col).value).startswith("=SUM(")


def test_assembly_costs_sheet_lists_assemblies_and_grand_total(tmp_path):
    import openpyxl
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(_hier_df(), str(out), "ASM")
    wb = openpyxl.load_workbook(str(out))
    assert "Assembly Costs" in wb.sheetnames
    ws = wb["Assembly Costs"]
    cells = [c.value for col in ws.iter_cols() for c in col]
    text = "\n".join(str(v) for v in cells if v is not None)
    assert "SF-2" in text                     # the one assembly is listed
    assert "GRAND TOTAL" in text
    # grand total is a formula referencing the Purchasing sheet
    assert any(isinstance(v, str) and v.startswith("=") and "Purchasing!" in v
               for v in cells)


def _purchasing_col_letter(ws_purchasing, name):
    from openpyxl.utils import get_column_letter
    header = [c.value for c in ws_purchasing[3]]
    return get_column_letter(header.index(name) + 1)


def test_assembly_costs_formulas_reference_exact_cells(tmp_path):
    import openpyxl
    # Two top-level assemblies (A, B), each with one nested child.
    df = pd.DataFrame({
        "Number":                 ["A", "A1", "B", "B1"],
        "Row Order":              ["1", "1.1", "2", "2.1"],
        "Position Number":        [None, None, None, None],
        "Item Qty":               [1, 3, 2, 4],
        "Units":                  ["Each", "Each", "Each", "Each"],
        "Category Name":          [None, None, None, None],
        "Revision":               [None, None, None, None],
        "State":                  [None, None, None, None],
        "Title (Item,CO)":        [None, None, None, None],
        "Description (Item,CO)":  ["asmA", "childA", "asmB", "childB"],
        "Source":                 ["Make", "Buy", "Make", "Buy"],
        "Material":               [None, None, None, None],
        "Material Finish":        [None, None, None, None],
        "Vendor":                 [None, "V", None, "V"],
        "Cost Per":               [None, 1.0, None, 2.0],
        "Vendor Number":          [None, None, None, None],
        "HS/HTS Code":            [None, None, None, None],
        "Shipping":               [None, None, None, None],
        "Tax/Tariff":             [None, None, None, None],
        "Lead Time (Business Days)": [None, None, None, None],
    })
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(df, str(out), "ASM")
    wb = openpyxl.load_workbook(str(out))
    pur = wb["Purchasing"]
    cost_col = _purchasing_col_letter(pur, "Cost Per")
    sub_col = _purchasing_col_letter(pur, "Sub Total")
    ac = wb["Assembly Costs"]
    # Purchasing data starts at row 4: A->row4, B->row6 (positional index * 1 + 4).
    # Assembly Costs data starts at row 3 (its own HDR=2). Two assemblies: A, B.
    assert ac.cell(3, 4).value == f"=Purchasing!{cost_col}4"   # A cost-to-make-one
    assert ac.cell(4, 4).value == f"=Purchasing!{cost_col}6"   # B cost-to-make-one
    # Grand total row = 5; sums exactly the two top-level Sub Total cells.
    assert ac.cell(5, 4).value == f"=Purchasing!{sub_col}4+Purchasing!{sub_col}6"


def test_unmatched_parts_are_highlighted_and_listed(tmp_path):
    import openpyxl
    df = _hier_df()
    # SF-21 is a Buy part with no Vendor → unmatched.
    df.loc[df["Number"] == "SF-21", "Vendor"] = None
    df.loc[df["Number"] == "SF-21", "Cost Per"] = None
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(df, str(out), "ASM")
    ws = openpyxl.load_workbook(str(out))["Purchasing"]
    header = [c.value for c in ws[3]]
    num_col = header.index("Number") + 1
    # the screw row (4+2=6) Number cell is filled with the unmatched color
    fill = ws.cell(6, num_col).fill
    assert fill.fgColor.rgb and fill.fgColor.rgb.endswith(bp.UNMATCHED_FILL)
    # an "Unmatched" note listing SF-21 appears somewhere below the table
    text = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
    assert "Unmatched" in text and "SF-21" in text


def test_assembly_buy_row_is_not_flagged_unmatched(tmp_path):
    import openpyxl
    # An assembly (has a child) dirtily tagged Source "Buy" with no Vendor must
    # NOT be highlighted or listed — its cost comes from the child roll-up.
    df = _hier_df()
    df.loc[df["Number"] == "SF-2", "Source"] = "Buy"
    df.loc[df["Number"] == "SF-2", "Vendor"] = None
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(df, str(out), "ASM")
    ws = openpyxl.load_workbook(str(out))["Purchasing"]
    header = [c.value for c in ws[3]]
    num_col = header.index("Number") + 1
    # SF-1 (Vendor Acme) and SF-21 (Vendor Bolts Inc) are matched; SF-2 is an
    # assembly → no row is unmatched, so no note is written.
    text = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
    assert "Unmatched" not in text
    # and SF-2's row (row 5) is not amber-highlighted
    assert not (ws.cell(5, num_col).fill.fgColor.rgb or "").endswith(bp.UNMATCHED_FILL)


def test_cd001608_end_to_end(tmp_path, monkeypatch):
    import openpyxl
    fixture = os.path.join(ROOT, "tests", "fixtures", "CD-001608-inventor-bom.txt")
    assert os.path.isfile(fixture)

    # Hermetic: no reference file → everything unmatched, costs blank.
    monkeypatch.setattr(bp, "find_purchased_items_file", lambda: None)

    result = bp.generate_from_file(fixture, "CD-001608", str(tmp_path))
    assert not result.get("error"), result

    wb = openpyxl.load_workbook(result["output_path"])
    assert {"Purchasing", "By Vendor", "Assembly Costs"}.issubset(wb.sheetnames)

    n_rows = len(bp.read_bom_file(fixture))     # 46 BOM lines
    ws = wb["Purchasing"]
    header = [c.value for c in ws[3]]
    num_col = header.index("Number") + 1
    qty_col = header.index("Item Qty") + 1
    numbers = [ws.cell(r, num_col).value for r in range(4, 4 + n_rows)]
    qtys = [ws.cell(r, qty_col).value for r in range(4, 4 + n_rows)]

    assert "SF-001580" in numbers                # real part number populated
    assert 120 in qtys                           # a real quantity from QTY
    # a library part with no SF number still appears (will be unmatched)
    assert any(isinstance(n, str) and n.startswith("ISO 4762") for n in numbers)

    acost = "\n".join(
        str(c.value) for col in wb["Assembly Costs"].iter_cols()
        for c in col if c.value
    )
    assert "CD-001621" in acost and "GRAND TOTAL" in acost


def test_coerce_normalizes_float_row_order_from_txt(tmp_path):
    # A .txt whose Item column is all single-level numerics parses as float64;
    # ensure Row Order comes out as clean strings so hierarchy parsing works.
    p = tmp_path / "bom.txt"
    p.write_text(
        "Item\tPart Number\tQTY\tDescription\n"
        "1\tSF-1\t2\tleaf\n"
        "14\tSF-2\t1\tassy\n"
        "14.1\tSF-3\t8\tchild\n",
        encoding="utf-8",
    )
    df = bp.read_bom_file(str(p))
    out, err = bp.coerce_bom_dataframe(df)
    assert err is None
    assert list(out["Row Order"]) == ["1", "14", "14.1"]   # not "14.0"
    children = bp.build_children_map(out)
    assert children[1] == [2]   # "14" is the parent of "14.1"


def test_row_order_tenth_child_not_collapsed(tmp_path):
    lines = ["Item\tPart Number\tQTY\tDescription", "2\tSF-2\t1\tassy"]
    for k in range(1, 11):
        lines.append(f"2.{k}\tSF-2{k}\t1\tchild{k}")
    (tmp_path / "bom.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    df = bp.read_bom_file(str(tmp_path / "bom.txt"))
    out, err = bp.coerce_bom_dataframe(df)
    assert err is None
    assert list(out["Row Order"]) == ["2"] + [f"2.{k}" for k in range(1, 11)]
    children = bp.build_children_map(out)
    assert children[0] == list(range(1, 11))   # all 10 distinct children under "2"


def test_unmatched_note_dedupes_repeated_part_numbers(tmp_path):
    import openpyxl
    df = _hier_df()
    dup = df[df["Number"] == "SF-21"].copy()
    dup["Row Order"] = "3"
    df2 = pd.concat([df, dup], ignore_index=True)
    df2.loc[df2["Number"] == "SF-21", "Vendor"] = None
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(df2, str(out), "ASM")
    ws = openpyxl.load_workbook(str(out))["Purchasing"]
    text = "\n".join(str(c.value) for col in ws.iter_cols() for c in col if c.value)
    assert "Unmatched (1)" in text   # de-duped: SF-21 counted once, not twice


def test_assembly_with_unpriced_descendant_is_marked(tmp_path):
    import openpyxl
    df = _hier_df()
    df.loc[df["Number"] == "SF-21", "Cost Per"] = None   # child leaf unpriced
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(df, str(out), "ASM")
    ac = openpyxl.load_workbook(str(out))["Assembly Costs"]
    text = "\n".join(str(c.value) for col in ac.iter_cols() for c in col if c.value)
    assert "*" in text
    assert "Includes unpriced parts" in text


def test_assembly_costs_no_marker_when_all_priced(tmp_path):
    import openpyxl
    df = _hier_df()   # SF-1 ($1) and SF-21 ($0.25) both priced; SF-2 is the assembly
    out = tmp_path / "s.xlsx"
    bp.build_purchasing_sheet(df, str(out), "ASM")
    ac = openpyxl.load_workbook(str(out))["Assembly Costs"]
    text = "\n".join(str(c.value) for col in ac.iter_cols() for c in col if c.value)
    assert "Includes unpriced parts" not in text
