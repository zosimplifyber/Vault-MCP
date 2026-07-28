# tests/test_purchasing_sheet_layout.py
"""Workbook layout: filters on the Purchasing tab, what lands on By Vendor."""
import os
import sys

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import bom_purchasing as bp  # noqa: E402

HDR_ROW = 3          # Purchasing tab: title, date, then headers


def _df(rows: list[dict]) -> pd.DataFrame:
    df, err = bp.coerce_bom_dataframe(pd.DataFrame(rows))
    assert err is None, err
    return df


def _book(tmp_path, df: pd.DataFrame):
    out = tmp_path / "sheet.xlsx"
    bp.build_purchasing_sheet(df, str(out), "TEST-001")
    return openpyxl.load_workbook(out)


BUY = {"Number": "SF-000067", "Row Order": "1", "Item Qty": 2, "Units": "Each",
       "Description (Item,CO)": "pull handle", "Source": "Buy",
       "Vendor": "Acme", "Cost Per": 1.5}
MAKE_OUTSOURCED = {"Number": "CD-001200", "Row Order": "2", "Item Qty": 1,
                   "Units": "Each", "Description (Item,CO)": "adapter plate",
                   "Source": "Make", "Vendor": "Machine Shop", "Cost Per": 40.0}
MAKE_INHOUSE = {"Number": "CD-001201", "Row Order": "3", "Item Qty": 1,
                "Units": "Each", "Description (Item,CO)": "bracket",
                "Source": "Make", "Vendor": None, "Cost Per": None}


class TestAutoFilter:
    def test_purchasing_tab_filters_the_full_header_row(self, tmp_path):
        df = _df([BUY, MAKE_OUTSOURCED])
        ws = _book(tmp_path, df)["Purchasing"]
        last_col = get_column_letter(len(bp.ALL_COLUMNS))
        assert ws.auto_filter.ref == f"A{HDR_ROW}:{last_col}{HDR_ROW + len(df)}"

    def test_filter_range_stops_above_the_unmatched_note(self, tmp_path):
        # An unpriced Buy row adds an "Unmatched (n)" note under the table; the
        # filter must not swallow it.
        df = _df([dict(BUY, Vendor=None, **{"Cost Per": None}), MAKE_OUTSOURCED])
        wb = _book(tmp_path, df)
        ws = wb["Purchasing"]
        note_row = HDR_ROW + len(df) + 2
        assert str(ws.cell(row=note_row, column=1).value).startswith("Unmatched")
        assert int(ws.auto_filter.ref.split(":")[1].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ")) < note_row


class TestTitleColumn:
    def test_the_sheet_shows_number_and_title_side_by_side(self, tmp_path):
        df = _df([dict(BUY, Title="CD-001578")])
        ws = _book(tmp_path, df)["Purchasing"]
        header = [c.value for c in ws[HDR_ROW]]
        assert header.index("Title") == header.index("Number") + 1
        assert ws.cell(row=4, column=header.index("Number") + 1).value == "SF-000067"
        assert ws.cell(row=4, column=header.index("Title") + 1).value == "CD-001578"


class TestVendorTabSources:
    def _numbers(self, wb):
        ws = wb["By Vendor"]
        col = bp.VENDOR_COLUMNS.index("Number") + 1
        return [ws.cell(row=r, column=col).value
                for r in range(3, ws.max_row + 1)]

    def test_includes_make_parts_that_have_a_vendor(self, tmp_path):
        wb = _book(tmp_path, _df([BUY, MAKE_OUTSOURCED]))
        assert "CD-001200" in self._numbers(wb)

    def test_excludes_make_parts_with_no_vendor(self, tmp_path):
        wb = _book(tmp_path, _df([BUY, MAKE_INHOUSE]))
        assert "CD-001201" not in self._numbers(wb)

    def test_still_includes_bought_parts(self, tmp_path):
        wb = _book(tmp_path, _df([BUY, MAKE_OUTSOURCED, MAKE_INHOUSE]))
        assert "SF-000067" in self._numbers(wb)
