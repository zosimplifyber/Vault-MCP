# tests/test_vault_state.py
"""Vault lifecycle state for BOM part numbers.

Vault's file search is keyword-based and matches properties as well as names —
searching "CD-001582" really does return an unrelated "SF-001915" alongside
"CD-001582.iam". Picking a state off the wrong file would be worse than showing
none, so a hit only counts when the file's name IS the part number.
"""
import json
import os
import sys

import pandas as pd

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import bom_purchasing as bp  # noqa: E402
import vault_state as vs  # noqa: E402


def _f(name, state):
    return {"name": name, "state": state, "lifecycleState": {"name": state}}


class TestPickStateFile:
    def test_ignores_a_keyword_only_hit(self):
        # "SF-001915" comes back for the query "CD-001582" — different part.
        assert vs.pick_state_file("CD-001582", [_f("SF-001915", "Released")]) is None

    def test_prefers_the_cad_model_over_other_files_with_the_same_number(self):
        files = [_f("CD-001582 BOM.xlsx", "Work in Progress"),
                 _f("CD-001582.idw", "Released"),
                 _f("CD-001582.iam", "Work in Progress")]
        assert vs.pick_state_file("CD-001582", files)["name"] == "CD-001582.iam"

    def test_matches_a_file_stored_without_an_extension(self):
        # Real vault data: purchased-part files are named "SF-001658" flat.
        assert vs.pick_state_file("SF-001658", [_f("SF-001658", "Released")])["name"] == "SF-001658"

    def test_matches_case_insensitively(self):
        hit = vs.pick_state_file("SF-001658", [_f("sf-001658.ipt", "Released")])
        assert hit is not None

    def test_ignores_a_longer_name_that_merely_starts_with_the_number(self):
        assert vs.pick_state_file("CD-001582", [_f("CD-001582 BOM.xlsx", "Released")]) is None


class TestEntityFiltering:
    def test_an_item_never_lends_its_state_to_a_file(self):
        # search_files also returns ItemVersions: SF-001922 the *item* is Work in
        # Progress while its CAD file CD-001578.ipt is Released.
        rows = [{"name": "SF-001922", "state": "Work in Progress",
                 "entityType": "ItemVersion"}]
        assert vs.pick_state_file("SF-001922", rows) is None

    def test_a_folder_does_not_mask_the_real_file(self):
        rows = [{"name": "ISO 4762 - M6 x 16 - Stainless Steel", "state": None,
                 "subfolderCount": 0},
                _f("ISO 4762 - M6 x 16 Stainless Steel.ipt", "Work in Progress")]
        hit = vs.pick_state_file("ISO 4762 - M6 x 16 - Stainless Steel", rows)
        assert hit["name"] == "ISO 4762 - M6 x 16 Stainless Steel.ipt"


class TestFileNameLookup:
    ROWS = [_f("CD-001578.ipt", "Released"),
            _f("CD-001578.SLDPRT", "Released"),
            _f("CD-001578_perf.stl", "Work in Progress")]

    def test_an_exact_file_name_picks_that_file(self):
        assert vs.pick_state_file("CD-001578.ipt", self.ROWS)["name"] == "CD-001578.ipt"

    def test_the_number_alone_still_prefers_the_model(self):
        assert vs.pick_state_file("CD-001578", self.ROWS)["name"] == "CD-001578.ipt"


class TestLooseAndPrefixMatching:
    def test_interior_punctuation_is_ignored(self):
        # BOM says "M6 x 50 - Stainless Steel"; the file has no dash.
        rows = [_f("ISO 4762 - M6 x 50 Stainless Steel.ipt", "Work in Progress")]
        hit = vs.pick_state_file("ISO 4762 - M6 x 50 - Stainless Steel", rows)
        assert hit["name"] == "ISO 4762 - M6 x 50 Stainless Steel.ipt"

    def test_a_typod_duplicate_is_not_matched(self):
        rows = [_f("ISO 4762 - M6 x 50ISO Stainless Steel.ipt", "Released")]
        assert vs.pick_state_file("ISO 4762 - M6 x 50 - Stainless Steel", rows) is None

    def test_a_trailing_detail_in_the_file_name_still_matches(self):
        rows = [_f("DIN 934 - M5 x 0.8.ipt", "Work in Progress")]
        assert vs.pick_state_file("DIN 934 - M5", rows)["name"] == "DIN 934 - M5 x 0.8.ipt"

    def test_a_shorter_size_does_not_grab_a_longer_one(self):
        # "M6 x 1" must never pick up "M6 x 10".
        rows = [_f("ISO 4762 - M6 x 10 Stainless Steel.ipt", "Released")]
        assert vs.pick_state_file("ISO 4762 - M6 x 1", rows) is None

    def test_candidates_that_disagree_about_the_state_stay_blank(self):
        rows = [_f("ISO 4762 - M6 x 20 A.ipt", "Released"),
                _f("ISO 4762 - M6 x 20 B.ipt", "Work in Progress")]
        assert vs.pick_state_file("ISO 4762 - M6 x 20", rows) is None


class TestStateOf:
    def test_reads_the_lifecycle_state_name(self):
        assert vs.state_of(_f("SF-1.ipt", "Released")) == "Released"

    def test_falls_back_to_the_flat_state_field(self):
        assert vs.state_of({"name": "SF-1.ipt", "state": "Released"}) == "Released"


class TestLookupWithoutVault:
    def test_missing_config_returns_a_warning_not_an_error(self, tmp_path):
        cfg = tmp_path / "config.json"
        cfg.write_text(json.dumps({"purchasing_reference": {}}), encoding="utf-8")
        states, warnings = vs.lookup_file_states(["SF-000067"], config_path=str(cfg))
        assert states == {}
        assert warnings and "Vault" in warnings[0]

    def test_no_numbers_does_not_even_look_for_config(self):
        assert vs.lookup_file_states([]) == ({}, [])

    def test_a_failing_fetch_degrades_to_a_warning(self, tmp_path, monkeypatch):
        cfg = tmp_path / "config.json"
        cfg.write_text(json.dumps({"vault": {"servername": "s", "username": "u",
                                             "password": "p", "database": "d"}}),
                       encoding="utf-8")

        def boom(*a, **k):
            raise RuntimeError("Vault sign-in failed")
        monkeypatch.setattr(vs, "_fetch_states", boom)

        states, warnings = vs.lookup_file_states(["SF-000067"], config_path=str(cfg))
        assert states == {}
        assert warnings and "sign-in failed" in warnings[0]


class TestSheetWiring:
    def test_blank_state_is_filled_from_vault(self, monkeypatch):
        monkeypatch.setattr(bp.vault_state, "lookup_file_states",
                            lambda nums, **k: ({"SF-000067": "Released"}, []))
        df = pd.DataFrame({"Number": ["SF-000067"], "State": [None]})
        warnings = bp._fill_state_from_vault(df)
        assert df.loc[0, "State"] == "Released"
        assert warnings == []

    def test_the_aliases_are_handed_to_the_lookup(self, monkeypatch):
        seen = {}

        def spy(numbers, **kwargs):
            seen.update(kwargs)
            return {}, []
        monkeypatch.setattr(bp.vault_state, "lookup_file_states", spy)
        df = pd.DataFrame({"Number": ["SF-001922"], "State": [None],
                           "File Name": ["CD-001578.ipt"]})
        bp._fill_state_from_vault(df)
        assert seen["aliases"] == {"SF-001922": ["CD-001578.ipt"]}

    def test_a_state_the_bom_already_carries_is_not_overwritten(self, monkeypatch):
        monkeypatch.setattr(bp.vault_state, "lookup_file_states",
                            lambda nums, **k: ({"SF-000067": "Released"}, []))
        df = pd.DataFrame({"Number": ["SF-000067"], "State": ["In Review"]})
        bp._fill_state_from_vault(df)
        assert df.loc[0, "State"] == "In Review"

    def test_parts_vault_does_not_know_stay_blank(self, monkeypatch):
        monkeypatch.setattr(bp.vault_state, "lookup_file_states",
                            lambda nums, **k: ({}, ["offline"]))
        df = pd.DataFrame({"Number": ["ISO 4762 - M6 x 10"], "State": [None]})
        warnings = bp._fill_state_from_vault(df)
        assert pd.isna(df.loc[0, "State"])
        assert warnings == ["offline"]

    def test_state_is_a_column_on_the_sheet_again(self):
        assert "State" in bp.BOM_COLUMNS
        assert "State" in bp.COLUMN_WIDTHS


class TestLookupKeys:
    def test_the_file_name_column_is_searched_first(self):
        df = pd.DataFrame({"Number": ["SF-001922"], "File Name": ["CD-001578.ipt"],
                           "Title": ["CD-001578"]})
        assert bp._state_lookup_aliases(df)["SF-001922"] == ["CD-001578.ipt", "CD-001578"]

    def test_the_canonical_name_column_is_used_as_the_file_name(self):
        # coerce_bom_dataframe renames the export's Filename column to Name.
        df = pd.DataFrame({"Number": ["SF-001922"], "Name": ["CD-001578.ipt"]})
        assert bp._state_lookup_aliases(df)["SF-001922"] == ["CD-001578.ipt"]

    def test_the_title_is_used_when_there_is_no_file_name(self):
        # The BOM's Part Number is the item number; Title carries the CD- number.
        df = pd.DataFrame({"Number": ["SF-001920"], "Title": ["CD-001574"]})
        assert bp._state_lookup_aliases(df)["SF-001920"] == ["CD-001574"]

    def test_rows_without_either_have_no_alias(self):
        df = pd.DataFrame({"Number": ["ISO 7089 - 5"], "Title": [float("nan")]})
        assert bp._state_lookup_aliases(df) == {}

    def test_alternate_file_name_headers_are_recognised(self):
        df = pd.DataFrame({"Number": ["SF-001922"], "Filename": ["CD-001578.ipt"]})
        assert bp._state_lookup_aliases(df)["SF-001922"] == ["CD-001578.ipt"]


class TestGeneratorWiring:
    def _bom(self, tmp_path):
        p = tmp_path / "bom.csv"
        p.write_text("Item,Part Number,QTY,BOM Structure,Description\n"
                     "1,SF-000067,2,Purchased,pull handle\n", encoding="utf-8")
        return p

    def test_generated_sheet_carries_the_vault_state(self, tmp_path, monkeypatch):
        import openpyxl
        monkeypatch.setattr(bp, "_enrich_with_reference",
                            lambda df, reference_path="": (df, 0, 0, [], []))
        monkeypatch.setattr(bp.vault_state, "lookup_file_states",
                            lambda nums, **k: ({"SF-000067": "Released"}, []))
        res = bp.generate_from_file(str(self._bom(tmp_path)), "TEST-1",
                                    output_dir=str(tmp_path))
        ws = openpyxl.load_workbook(res["output_path"])["Purchasing"]
        header = [c.value for c in ws[3]]
        assert ws.cell(row=4, column=header.index("State") + 1).value == "Released"

    def test_vault_warnings_reach_the_caller(self, tmp_path, monkeypatch):
        monkeypatch.setattr(bp, "_enrich_with_reference",
                            lambda df, reference_path="": (df, 0, 0, [], []))
        monkeypatch.setattr(bp.vault_state, "lookup_file_states",
                            lambda nums, **k: ({}, ["No Vault connection configured"]))
        res = bp.generate_from_file(str(self._bom(tmp_path)), "TEST-1",
                                    output_dir=str(tmp_path))
        assert any("No Vault connection" in w for w in res["warnings"])
