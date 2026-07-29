# tests/test_check_file_properties.py
"""Unit tests for the file-based property checker.

Runs against responses recorded from the live vault (``tests/fixtures/
vault_file_cd001659*.json``), so nothing here touches the network.
"""
import asyncio
import json
import os
import re
import sys

import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
for path in (ROOT, os.path.join(ROOT, "scripts")):
    if path not in sys.path:
        sys.path.insert(0, path)

import check_file_properties as cfp  # noqa: E402

FIXTURES = os.path.join(ROOT, "tests", "fixtures")


def _fixture(name):
    with open(os.path.join(FIXTURES, name), encoding="utf-8") as fh:
        return json.load(fh)


@pytest.fixture()
def search_payload():
    """A recorded /file-versions?q=CD-001659.iam&option[propDefIds]=all response."""
    return _fixture("vault_file_cd001659.json")


@pytest.fixture()
def uses_payload():
    """A recorded /file-versions/124814/uses response (CAD BOM children)."""
    return _fixture("vault_file_cd001659_uses.json")


@pytest.fixture()
def rules():
    return cfp.load_json(cfp.DEFAULT_RULES_PATH)


@pytest.fixture()
def properties(search_payload):
    record = search_payload["results"][0]
    defs = cfp.extract_definition_index(search_payload)
    return cfp.flatten_file_properties(record, defs)


@pytest.fixture()
def failing_properties(properties):
    """CD-001659.iam with two gated columns emptied, so it reports FAIL.

    The real file passes everything now that Title, CAD Category and
    Description (File) are ungated, so tests that need a failure build one.
    """
    return dict(properties, **{"Engineer": "", "Project": ""})


# --------------------------------------------------------------------------- extraction

def test_definitions_come_from_the_response_itself(search_payload):
    """Vault embeds the property definitions, so no extra round-trip is needed."""
    defs = cfp.extract_definition_index(search_payload)
    assert defs, "the response should carry included.propertyDefinition"
    assert defs["77"]["displayName"] == "State"


def test_extract_definition_index_tolerates_a_missing_included_block():
    assert cfp.extract_definition_index({}) == {}
    assert cfp.extract_definition_index({"included": {}}) == {}
    assert cfp.extract_definition_index(None) == {}


def test_properties_resolve_to_their_display_names(properties):
    """The names have to match the keys in file_property_rules.json."""
    assert properties["Category Name"] == "Assembly - Engineering"
    assert properties["Description (File)"] == "bmw kft90 hot a adapter plate assembly"
    assert properties["Title"] == "CD-001659"
    assert properties["Source"] == "Make"
    assert properties["Engineer"] == "Alan Y."
    assert properties["Engr Approved By"] == "Zak O."
    assert properties["Project"] == "BMW"
    assert properties["Revision"] == "4"
    assert properties["State"] == "Released"
    assert properties["File Name"] == "CD-001659.iam"
    assert properties["CAD Category"] == ""


def test_historical_twins_do_not_shadow_the_live_value(properties):
    """Vault returns State and 'State (Historical)' as separate definitions."""
    assert properties["State"] == "Released"
    assert "State (Historical)" in properties


def test_a_pinned_version_recovers_its_state_from_lifecycle():
    """A CAD BOM child is whatever version its parent pins — often not latest.

    Vault blanks the live State on those and reports the real one via
    lifecycleState. Without the fallback every pinned child would report
    'State missing (required)' while the same file checked alone reports
    Released.
    """
    defs = {
        "77": {"displayName": "State", "systemName": "State"},
        "78": {"displayName": "State (Historical)", "systemName": "State(Ver)"},
    }
    record = {
        "name": "CD-001171.ipt", "state": "", "revision": "4",
        "lifecycleState": {"name": "Released", "displayName": "Released"},
        "properties": [
            {"propertyDefinitionId": "77", "value": ""},
            {"propertyDefinitionId": "78", "value": "Released"},
        ],
    }
    flat = cfp.flatten_file_properties(record, defs)
    assert flat["State"] == "Released"
    assert flat["State (Historical)"] == "Released"


def test_the_historical_twin_is_the_fallback_when_lifecycle_is_absent():
    defs = {
        "77": {"displayName": "State", "systemName": "State"},
        "78": {"displayName": "State (Historical)", "systemName": "State(Ver)"},
    }
    record = {
        "name": "CD-001171.ipt", "state": "",
        "properties": [
            {"propertyDefinitionId": "77", "value": ""},
            {"propertyDefinitionId": "78", "value": "Work in Progress"},
        ],
    }
    assert cfp.flatten_file_properties(record, defs)["State"] == "Work in Progress"


def test_a_live_state_is_never_overwritten_by_the_historical_one():
    """The latest version's own State wins — the fallback is only for blanks."""
    defs = {
        "77": {"displayName": "State", "systemName": "State"},
        "78": {"displayName": "State (Historical)", "systemName": "State(Ver)"},
    }
    record = {
        "name": "CD-001171.ipt", "state": "Work in Progress",
        "lifecycleState": {"name": "Released"},
        "properties": [
            {"propertyDefinitionId": "77", "value": "Work in Progress"},
            {"propertyDefinitionId": "78", "value": "Released"},
        ],
    }
    assert cfp.flatten_file_properties(record, defs)["State"] == "Work in Progress"


def test_a_genuinely_stateless_record_stays_blank():
    """No state anywhere means blank — the rule should still flag it."""
    defs = {"77": {"displayName": "State", "systemName": "State"}}
    record = {"name": "X.ipt", "state": "",
              "properties": [{"propertyDefinitionId": "77", "value": ""}]}
    assert cfp.flatten_file_properties(record, defs)["State"] == ""


def test_unnamed_properties_are_skipped_not_invented():
    """A value whose definition didn't come back must not become a fake key."""
    record = {"name": "X.ipt", "properties": [
        {"propertyDefinitionId": "999", "value": "orphan"},
    ]}
    flat = cfp.flatten_file_properties(record, {})
    assert "orphan" not in flat.values()
    assert flat == {"File Name": "X.ipt"}


def test_a_populated_value_wins_over_an_empty_one():
    defs = {"1": {"displayName": "Vendor"}, "2": {"displayName": "Vendor"}}
    record = {"properties": [
        {"propertyDefinitionId": "1", "value": "MiSUMi"},
        {"propertyDefinitionId": "2", "value": ""},
    ]}
    assert cfp.flatten_file_properties(record, defs)["Vendor"] == "MiSUMi"

    record_reversed = {"properties": [
        {"propertyDefinitionId": "2", "value": ""},
        {"propertyDefinitionId": "1", "value": "MiSUMi"},
    ]}
    assert cfp.flatten_file_properties(record_reversed, defs)["Vendor"] == "MiSUMi"


# --------------------------------------------------------------------------- name matching

def test_exact_name_match_wins_over_the_first_result():
    records = [{"name": "CD-001659-BRACKET.ipt"}, {"name": "CD-001659.iam"}]
    record, note = cfp.select_file_record(records, "CD-001659.iam")
    assert record["name"] == "CD-001659.iam"
    assert note is None


def test_name_matching_ignores_case():
    records = [{"name": "CD-001659.iam"}]
    record, note = cfp.select_file_record(records, "cd-001659.IAM")
    assert record["name"] == "CD-001659.iam"
    assert note is None


def test_an_inexact_match_is_flagged_rather_than_checked_silently():
    records = [{"name": "CD-001659-BRACKET.ipt"}, {"name": "CD-001659-PLATE.ipt"}]
    record, note = cfp.select_file_record(records, "CD-001659")
    assert record["name"] == "CD-001659-BRACKET.ipt"
    assert note and "CD-001659-BRACKET.ipt" in note


def test_no_results_raises():
    with pytest.raises(RuntimeError, match="No files found"):
        cfp.select_file_record([], "NOPE.ipt")


# --------------------------------------------------------------------------- rules

def test_cd001659_passes_every_gated_property(properties, rules):
    """The end-to-end expectation for the worked example.

    Title, CAD Category and Description (File) are all populated oddly — the
    description carries a customer name and digits — but none of the three is
    gated any more, and every gated property is filled in.
    """
    result = cfp.evaluate_against_rules(
        properties, properties["Category Name"], rules)
    assert result["category_resolved"] == "Assembly - Engineering"

    failed = [(r["property"], r["failures"])
              for r in result["report"]["results"] if not r["passed"]]
    assert failed == []
    assert result["report"]["passed"] == result["report"]["total"]


def test_a_blank_gated_property_still_fails_cd001659(properties, rules):
    """Guard against the rule set having been hollowed out entirely."""
    props = dict(properties, **{"Engineer": ""})
    result = cfp.evaluate_against_rules(props, "Assembly - Engineering", rules)
    failed = {r["property"] for r in result["report"]["results"] if not r["passed"]}
    assert failed == {"Engineer"}


def test_the_compliant_properties_pass(properties, rules):
    result = cfp.evaluate_against_rules(
        properties, properties["Category Name"], rules)
    passed = {r["property"] for r in result["report"]["results"] if r["passed"]}
    assert {"Source", "Engineer", "Engr Approved By", "Project",
            "Revision", "State", "Category Name", "File Name"} <= passed


def test_an_iso_standard_purchased_part_is_exempt_from_a_vendor_number(rules):
    """A generic ISO screw has no single supplier SKU — but it still has a supplier."""
    props = {
        "File Name": "ISO 4762 M6x20.ipt", "Title": "ISO 4762 M6x20",
        "Description (File)": "socket head cap screw",
        "Revision": "1", "State": "Released", "Source": "Buy",
        "Vendor": "McMASTER-CARR", "Engineer": "Zak O.",
        "Engr Approved By": "Zak O.", "Designer": "Zak O.", "Project": "General",
        "Category Name": "Part - Purchased", "CAD Category": "Part - Purchased",
    }
    result = cfp.evaluate_against_rules(props, "Part - Purchased", rules)
    assert result["report"]["failed"] == 0, [
        (r["property"], r["failures"])
        for r in result["report"]["results"] if not r["passed"]
    ]


def test_a_non_standard_purchased_part_needs_a_vendor_number(rules):
    props = {
        "File Name": "CD-002000.ipt", "Title": "linear rail block",
        "Description (File)": "linear rail block",
        "Revision": "1", "State": "Released", "Source": "Buy",
        "Vendor": "MiSUMi", "Engineer": "Zak O.", "Engr Approved By": "Zak O.",
        "Designer": "Zak O.", "Project": "General",
        "Category Name": "Part - Purchased", "CAD Category": "Part - Purchased",
    }
    result = cfp.evaluate_against_rules(props, "Part - Purchased", rules)
    number = next(r for r in result["report"]["results"]
                  if r["property"] == "Vendor Number")
    assert not number["passed"]


def test_an_unknown_category_skips_rather_than_passes(properties, rules):
    result = cfp.evaluate_against_rules(properties, "Documents", rules)
    assert result["category_resolved"] is None
    assert result["report"] is None


def test_category_aliases_resolve(rules):
    assert cfp.resolve_category("Assembly-Engineering", rules) == "Assembly - Engineering"
    assert cfp.resolve_category("Purchased Part", rules) == "Part - Purchased"


# --------------------------------------------------------------------------- rules file

def test_every_rule_set_is_well_formed(rules):
    """A bad regex in the JSON should fail here, not mid-report."""
    assert rules["categories"], "there should be at least one rule set"
    for category, spec in rules["categories"].items():
        for prop, rule in (spec.get("properties") or {}).items():
            where = f"{category}.{prop}"
            assert isinstance(rule, dict), where
            for key in ("pattern",):
                if rule.get(key):
                    re.compile(rule[key])
            for pattern in rule.get("forbidden_patterns") or []:
                re.compile(pattern)
            clause = rule.get("required_unless")
            if clause:
                assert clause.get("property"), where
                re.compile(clause["matches_pattern"])
                # The property the exemption reads must be one this rule set
                # actually checks, or the exemption can never fire.
                assert clause["property"] in spec["properties"], where


CONTENT_CENTER = "Part - Content Center"
PURCHASED = "Part - Purchased"

# Bought parts — catalogue hardware and Inventor library files. Nobody in-house
# designs, engineers, approves, or bills these to a project, so they are exempt
# from the sign-off columns.
BOUGHT = (PURCHASED, CONTENT_CENTER)

# In-house work, where engineering ownership is expected and gated.
IN_HOUSE = ("Assembly - Engineering", "Part - Engineering", "Drawing - Engineering")

# Which categories are exempt from each gated column. Every column below is
# required in every category EXCEPT the ones listed. Dropping one from a
# category that isn't exempt silently weakens the gate, so it fails here.
GATED_COLUMNS = {
    "State": (),
    "Source": (CONTENT_CENTER,),
    "Revision": (PURCHASED,),
    "Engineer": BOUGHT,
    "Engr Approved By": BOUGHT,
    "Project": BOUGHT,
    # Assemblies too: the design credit lives on the child parts.
    "Designer": ("Assembly - Engineering",) + BOUGHT,
}

# Declared so their values show up in the report, but never gated anywhere.
NEVER_REQUIRED = ("Title", "CAD Category", "Description (File)")

PART_CATEGORIES = ("Part - Engineering", "Part - Purchased", CONTENT_CENTER)


@pytest.mark.parametrize("prop,exempt", sorted(GATED_COLUMNS.items()))
def test_the_gated_columns_are_required_where_they_should_be(rules, prop, exempt):
    for category, spec in rules["categories"].items():
        rule = (spec.get("properties") or {}).get(prop)
        assert rule is not None, f"{category} has no rule for {prop}"
        expected = category not in exempt
        assert bool(rule.get("required")) is expected, (
            f"{category}.{prop} required should be {expected}"
        )


@pytest.mark.parametrize("prop,exempt", sorted(GATED_COLUMNS.items()))
def test_a_missing_gated_column_is_reported(rules, prop, exempt):
    """Blank it out: gated categories must flag it, exempt ones must not."""
    for category in rules["categories"]:
        props = {p: "placeholder" for p in rules["categories"][category]["properties"]}
        props[prop] = ""
        result = cfp.evaluate_against_rules(props, category, rules)
        failed = {r["property"] for r in result["report"]["results"] if not r["passed"]}
        if category in exempt:
            assert prop not in failed, f"{category} should not gate {prop}"
        else:
            assert prop in failed, f"{category} did not flag a blank {prop}"


def test_content_center_parts_skip_the_sign_off_fields(rules):
    """Library fasteners carry no in-house engineering ownership."""
    props = {p: "" for p in rules["categories"][CONTENT_CENTER]["properties"]}
    props.update({
        "File Name": "ISO 4762 M6x20.ipt", "Revision": "1", "State": "Released",
        "Material": "Steel", "Vendor": "McMASTER-CARR",
        "Vendor Number": "91290A326", "Category Name": CONTENT_CENTER,
    })
    result = cfp.evaluate_against_rules(props, CONTENT_CENTER, rules)
    assert result["report"]["failed"] == 0, [
        (r["property"], r["failures"])
        for r in result["report"]["results"] if not r["passed"]
    ]


@pytest.mark.parametrize("source", ["Buy", "N/A", "Make", ""])
def test_content_center_accepts_any_source(rules, source):
    """Library files inherit whatever Source the template carries.

    In this vault that's a mix of 'N/A' (7), blank (5) and 'Buy' (2), so an
    allowed_values list would fail most of them even with required:false.
    """
    props = {p: "placeholder" for p in rules["categories"][CONTENT_CENTER]["properties"]}
    props["Source"] = source
    result = cfp.evaluate_against_rules(props, CONTENT_CENTER, rules)
    checked = next(r for r in result["report"]["results"]
                   if r["property"] == "Source")
    assert checked["passed"], f"Content Center rejected Source={source!r}"


@pytest.mark.parametrize("category", ("Assembly - Engineering", "Part - Engineering"))
def test_other_categories_still_constrain_source(rules, category):
    """Ungating Content Center must not have loosened Source everywhere."""
    props = {p: "placeholder" for p in rules["categories"][category]["properties"]}
    props["Source"] = "N/A"
    result = cfp.evaluate_against_rules(props, category, rules)
    checked = next(r for r in result["report"]["results"]
                   if r["property"] == "Source")
    assert not checked["passed"], f"{category} should still require Source=Make"


@pytest.mark.parametrize("category", BOUGHT)
def test_bought_parts_accept_not_reviewed(rules, category):
    """You don't engineering-review a catalogue part, so the string is fine."""
    rule = rules["categories"][category]["properties"]["Engr Approved By"]
    assert "NOT REVIEWED" not in (rule.get("forbidden_values") or []), category

    props = {p: "placeholder" for p in rules["categories"][category]["properties"]}
    props["Engr Approved By"] = "NOT REVIEWED"
    result = cfp.evaluate_against_rules(props, category, rules)
    approved = next(r for r in result["report"]["results"]
                    if r["property"] == "Engr Approved By")
    assert approved["passed"], f"{category} should tolerate NOT REVIEWED"


def test_an_assembly_with_no_designer_still_passes(properties, rules):
    props = dict(properties, **{"Designer": ""})
    result = cfp.evaluate_against_rules(props, "Assembly - Engineering", rules)
    designer = next(r for r in result["report"]["results"]
                    if r["property"] == "Designer")
    assert designer["passed"]


@pytest.mark.parametrize("prop", NEVER_REQUIRED)
def test_title_and_cad_category_are_reported_but_never_gated(rules, prop):
    """Still visible in the report, but they can never fail a file."""
    for category, spec in rules["categories"].items():
        rule = (spec.get("properties") or {}).get(prop)
        if rule is None:
            continue                    # not every category declares them
        assert rule.get("required") is not True, f"{category}.{prop} is required"

        # Blank, and anything else, has to pass.
        for value in ("", "CD-001659", "anything at all"):
            props = {p: "placeholder" for p in spec["properties"]}
            props[prop] = value
            result = cfp.evaluate_against_rules(props, category, rules)
            checked = next(r for r in result["report"]["results"]
                           if r["property"] == prop)
            assert checked["passed"], f"{category}.{prop}={value!r} failed"


@pytest.mark.parametrize("prop", NEVER_REQUIRED)
def test_the_ungated_properties_still_appear_in_the_report(properties, rules, prop):
    result = cfp.evaluate_against_rules(
        properties, properties["Category Name"], rules)
    reported = {r["property"] for r in result["report"]["results"]}
    assert prop in reported


@pytest.mark.parametrize("category", IN_HOUSE)
def test_engr_approved_by_rejects_not_reviewed_on_in_house_work(rules, category):
    """Where a review is expected, 'NOT REVIEWED' means it hasn't happened."""
    spec = rules["categories"][category]
    rule = spec["properties"]["Engr Approved By"]
    assert "NOT REVIEWED" in (rule.get("forbidden_values") or []), category
    assert "NOT REVIEWED" not in (rule.get("allowed_values") or []), category

    props = {p: "placeholder" for p in spec["properties"]}
    props["Engr Approved By"] = "NOT REVIEWED"
    result = cfp.evaluate_against_rules(props, category, rules)
    approved = next(r for r in result["report"]["results"]
                    if r["property"] == "Engr Approved By")
    assert not approved["passed"], f"{category} accepted NOT REVIEWED"


def test_the_two_category_groups_cover_every_rule_set(rules):
    """If a category is added, it has to be classified in one group or the other."""
    assert set(IN_HOUSE) | set(BOUGHT) == set(rules["categories"])


@pytest.mark.parametrize("category", PART_CATEGORIES)
def test_every_part_requires_a_vendor(rules, category):
    rule = rules["categories"][category]["properties"]["Vendor"]
    assert rule.get("required") is True, f"{category} does not require Vendor"
    assert "required_unless" not in rule, (
        f"{category} lets Vendor off the hook — it should be required outright"
    )


@pytest.mark.parametrize("category", PART_CATEGORIES)
def test_a_blank_vendor_is_reported_on_every_part(rules, category):
    props = {p: "placeholder"
             for p in rules["categories"][category]["properties"]}
    props["Vendor"] = ""
    result = cfp.evaluate_against_rules(props, category, rules)
    vendor = next(r for r in result["report"]["results"] if r["property"] == "Vendor")
    assert not vendor["passed"], f"{category} accepted a blank Vendor"


def test_an_iso_standard_part_still_needs_a_vendor(rules):
    """The standard-part exemption applies to Vendor Number, never to Vendor."""
    props = {
        "File Name": "ISO 4762 M6x20.ipt", "Title": "ISO 4762 M6x20",
        "Description (File)": "socket head cap screw",
        "Revision": "1", "State": "Released", "Source": "Buy",
        "Vendor": "", "Engineer": "Zak O.", "Engr Approved By": "Zak O.",
        "Designer": "Zak O.", "Project": "General",
        "Category Name": "Part - Purchased", "CAD Category": "Part - Purchased",
    }
    result = cfp.evaluate_against_rules(props, "Part - Purchased", rules)
    by_name = {r["property"]: r for r in result["report"]["results"]}
    assert not by_name["Vendor"]["passed"], "Vendor has no standard-part exemption"
    assert by_name["Vendor Number"]["passed"], "Vendor Number keeps its exemption"


def test_file_rules_do_not_use_item_only_property_names(rules):
    """Guards against copy-paste drift back toward item_property_rules.json."""
    item_only = {"Title (Item,CO)", "Description (Item,CO)", "Number", "Units"}
    for category, spec in rules["categories"].items():
        overlap = item_only & set(spec.get("properties") or {})
        assert not overlap, f"{category} uses item-side property names: {overlap}"


# --------------------------------------------------------------------------- children

def test_cad_bom_children_come_back_enriched(uses_payload, monkeypatch):
    """/uses returns children with their properties, so no extra fetch per child."""
    defs = cfp.extract_definition_index(uses_payload)
    children = []
    for row in uses_payload["results"]:
        child = row["childFile"]
        children.append(cfp.flatten_file_properties(child, defs))

    assert len(children) == 3
    names = sorted(c["File Name"] for c in children)
    assert names == ["CD-001162.ipt", "CD-001171.ipt", "CD-001624.ipt"]
    for props in children:
        assert props["Category Name"] == "Part - Engineering"
        assert props["Description (File)"]


class _UsesOnlyAPI:
    """Serves a /uses payload; has no latest-version lookup."""

    def __init__(self, payload):
        self.payload = payload

    async def get_file_uses(self, **_kwargs):
        return {"error": False, "status_code": 200, "data": self.payload}


def test_children_are_deduplicated_by_file_version(uses_payload):
    """A part used twice in an assembly is checked once."""
    import asyncio

    doubled = dict(uses_payload)
    doubled["results"] = list(uses_payload["results"]) + [uses_payload["results"][0]]

    children = asyncio.run(cfp.fetch_cad_children(
        _UsesOnlyAPI(doubled), "1", "124814", use_latest=False))
    assert len(children) == 3
    assert len({c["file_version_id"] for c in children}) == 3


def test_the_same_file_pinned_at_two_versions_collapses_to_one_row(uses_payload):
    """Dedupe keys on the File master, not the version, so two pinned
    versions of one file don't become two rows."""
    import asyncio
    import copy

    payload = copy.deepcopy(uses_payload)
    other_version = copy.deepcopy(payload["results"][0])
    other_version["childFile"]["id"] = "999999"        # different version...
    # ...same File master, which is what identifies the file
    payload["results"].append(other_version)

    children = asyncio.run(cfp.fetch_cad_children(
        _UsesOnlyAPI(payload), "1", "124814", use_latest=False))
    assert len(children) == 3


def test_children_are_checked_at_their_latest_version(uses_payload):
    """The pinned version is what /uses returns; we grade the newest one.

    Otherwise a file you just fixed keeps reporting its old failures because
    the parent still references the pre-fix version.
    """
    import asyncio

    latest = {
        "results": [{
            "name": "CD-001624.ipt", "id": "999001", "revision": "9",
            "state": "Released", "category": "Part - Engineering",
            "file": {"id": "111"},
            "properties": [{"propertyDefinitionId": "1", "value": "Make"}],
        }],
        "included": {"propertyDefinition": {
            "1": {"displayName": "Source", "systemName": "Source"}}},
    }

    class FakeAPI(_UsesOnlyAPI):
        async def search_file_versions(self, **kwargs):
            assert kwargs.get("latest_only") is True
            return {"error": False, "status_code": 200, "data": latest}

    children = asyncio.run(cfp.fetch_cad_children(
        FakeAPI(uses_payload), "1", "124814"))

    upgraded = next(c for c in children if c["file_name"] == "CD-001624.ipt")
    assert upgraded["properties"]["Source"] == "Make"
    assert upgraded["properties"]["Revision"] == "9"
    assert upgraded["file_version_id"] == "999001"
    # The version the parent pins is still recorded for context.
    assert upgraded["pinned_revision"] == "3"


def test_a_child_whose_latest_version_cannot_be_read_is_an_error(uses_payload):
    """Never grade stale data as though it were current — say so instead."""
    import asyncio

    class FakeAPI(_UsesOnlyAPI):
        async def search_file_versions(self, **_kwargs):
            return {"error": True, "status_code": 503,
                    "data": {"message": "service unavailable"}}

    children = asyncio.run(cfp.fetch_cad_children(
        FakeAPI(uses_payload), "1", "124814"))
    assert all(c["error"] for c in children)
    assert "latest version" in children[0]["error"]


def test_a_child_that_no_longer_resolves_is_an_error(uses_payload):
    import asyncio

    class FakeAPI(_UsesOnlyAPI):
        async def search_file_versions(self, **_kwargs):
            return {"error": False, "status_code": 200, "data": {"results": []}}

    children = asyncio.run(cfp.fetch_cad_children(
        FakeAPI(uses_payload), "1", "124814"))
    assert all(c["error"] for c in children)


def test_the_latest_version_lookups_are_concurrency_capped():
    """A big assembly must not open one socket per child.

    vault_rest_api builds a fresh httpx client per request, so nothing else
    throttles this — an unbounded gather over a 200-part BOM hits Vault with
    200 simultaneous connections.
    """
    import asyncio

    children = [{
        "childFile": {
            "name": f"CD-{i:06d}.ipt", "id": str(500000 + i),
            "file": {"id": str(i)}, "revision": "1", "state": "Released",
            "category": "Part - Engineering",
        },
        "fileAssocType": "Dependency",
    } for i in range(200)]

    class FakeAPI(_UsesOnlyAPI):
        def __init__(self):
            super().__init__({
                "results": children,
                "included": {"propertyDefinition": {
                    "1": {"displayName": "Source", "systemName": "Source"}}},
            })
            self.in_flight = 0
            self.peak = 0

        async def search_file_versions(self, **kwargs):
            self.in_flight += 1
            self.peak = max(self.peak, self.in_flight)
            # Yield so every coroutine gather has already scheduled gets a
            # chance to enter before the first one returns.
            await asyncio.sleep(0)
            self.in_flight -= 1
            return {"error": False, "status_code": 200, "data": {"results": [
                {"name": kwargs["query"], "id": "999", "file": {"id": "1"}},
            ]}}

    api = FakeAPI()
    asyncio.run(cfp.fetch_cad_children(api, "1", "124814"))

    assert api.peak <= cfp.MAX_CONCURRENCY, (
        f"{api.peak} concurrent lookups; expected at most {cfp.MAX_CONCURRENCY}"
    )


def test_a_failed_bom_walk_raises_with_the_vault_message():
    import asyncio

    class FakeAPI:
        async def get_file_uses(self, **_kwargs):
            return {"error": True, "status_code": 403,
                    "data": {"message": "forbidden"}}

    with pytest.raises(RuntimeError, match="CAD BOM walk failed"):
        asyncio.run(cfp.fetch_cad_children(FakeAPI(), "1", "124814"))


# --------------------------------------------------------------------------- exit codes

def _result(report=None, category="Assembly - Engineering", children=None):
    return {
        "report": report, "category_resolved": category,
        "children": children or [],
    }


def test_exit_code_zero_when_everything_passes():
    assert cfp.result_exit_code(_result({"failed": 0, "passed": 5, "total": 5})) == 0


def test_exit_code_one_when_the_file_fails():
    assert cfp.result_exit_code(_result({"failed": 2, "passed": 3, "total": 5})) == 1


def test_exit_code_one_when_only_a_child_fails():
    child = {"error": None, "category_resolved": "Part - Engineering",
             "report": {"failed": 1}}
    result = _result({"failed": 0, "passed": 5, "total": 5}, children=[child])
    assert cfp.result_exit_code(result) == 1


def test_exit_code_two_when_no_rule_set_matched():
    assert cfp.result_exit_code(_result(None, category=None)) == 2


def test_a_skipped_child_is_not_a_failure():
    child = {"error": None, "category_resolved": None, "report": None}
    result = _result({"failed": 0, "passed": 5, "total": 5}, children=[child])
    assert cfp.child_status(child) == "SKIP"
    assert cfp.result_exit_code(result) == 0


# --------------------------------------------------------------------------- reporting

def test_markdown_report_names_every_failure(failing_properties, rules):
    evaluated = cfp.evaluate_against_rules(
        failing_properties, failing_properties["Category Name"], rules)
    md = cfp.format_markdown_report({
        "file_name": "CD-001659.iam",
        "info": {"properties": failing_properties, "note": None},
        "children": [], "children_error": None, "recursive": False,
        **evaluated,
    })
    assert "# File Property Compliance — `CD-001659.iam`" in md
    assert "**FAIL**" in md
    for prop in ("Engineer", "Project"):
        assert f"`{prop}`" in md


def test_markdown_escapes_pipes_so_tables_do_not_break():
    """Alternation regexes like (CD|SF|MFG|DT) can appear verbatim in failures."""
    md = cfp.format_markdown_report({
        "file_name": "X.ipt",
        "info": {"properties": {}, "note": None},
        "category_raw": "Part - Engineering",
        "category_resolved": "Part - Engineering",
        "report": {
            "failed": 1, "passed": 0, "total": 1,
            "results": [{
                "property": "Title", "passed": False,
                "value": "a|b",
                "failures": [r"contains forbidden pattern /(CD|SF|MFG|DT)/"],
            }],
        },
        "children": [], "children_error": None, "recursive": False,
    })
    row = next(line for line in md.splitlines() if line.startswith("| `Title`"))
    # Three columns means four pipes; any extra is an unescaped one leaking in.
    assert row.count("|") - row.count("\\|") == 4, row


def test_gui_renders_a_report_without_blowing_up(properties, rules):
    """The GUI's render path is only exercised at runtime — pin it here."""
    tk = pytest.importorskip("tkinter")
    try:
        root = tk.Tk()
    except tk.TclError:
        pytest.skip("no display available")
    root.withdraw()
    try:
        from gui.file_property_check import run_gui
        run_gui(parent=root)
        root.update_idletasks()

        window = next(w for w in root.winfo_children()
                      if isinstance(w, tk.Toplevel))
        text = _find_text_widget(window)
        assert text is not None, "the report pane should exist"

        evaluated = cfp.evaluate_against_rules(
            properties, properties["Category Name"], rules)
        window.render_for_test({
            "file_name": "CD-001659.iam",
            "info": {"properties": properties, "note": None},
            "children": [], "children_error": None, "recursive": False,
            **evaluated,
        })
        root.update_idletasks()

        rendered = text.get("1.0", "end")
        assert "CD-001659.iam" in rendered
        assert "Assembly - Engineering" in rendered
        for prop in ("Title", "Description (File)", "CAD Category"):
            assert prop in rendered
        assert "16/16 properties passed" in rendered
    finally:
        root.destroy()


def _find_text_widget(widget):
    import tkinter as tk
    for child in widget.winfo_children():
        if isinstance(child, tk.Text):
            return child
        found = _find_text_widget(child)
        if found is not None:
            return found
    return None


# --------------------------------------------------------------------------- excel

@pytest.fixture()
def checked_result(failing_properties, rules):
    """A result covering all three statuses: a failing top file, a failing
    child, and a child whose category has no rule set."""
    child_props = dict(failing_properties, **{
        "File Name": "CD-001624.ipt",
        "Category Name": "Part - Engineering",
        "Vendor": "",
        "Material": "Aluminum 6061",
    })
    return {
        "file_name": "CD-001659.iam",
        "info": {"properties": failing_properties, "note": None},
        "recursive": True,
        "children_error": None,
        "children": [
            {
                "file_name": "CD-001624.ipt", "file_version_id": "1",
                "assoc_type": "Dependency", "properties": child_props, "error": None,
                **cfp.evaluate_against_rules(child_props, "Part - Engineering", rules),
            },
            {
                "file_name": "notes.xlsx", "file_version_id": "2",
                "assoc_type": "Dependency", "properties": {}, "error": None,
                "category_raw": "Documents", "category_resolved": None,
                "report": None,
            },
        ],
        **cfp.evaluate_against_rules(
            failing_properties, failing_properties["Category Name"], rules),
    }


def test_export_writes_a_two_sheet_workbook(checked_result, tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    out = tmp_path / "report.xlsx"
    written = cfp.export_to_excel(checked_result, out)

    assert written == str(out)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Summary", "Detail"]
    for ws in wb:
        assert ws.freeze_panes == "A4", "headers should stay put while scrolling"
        assert ws.auto_filter.ref, "every sheet should be filterable"


def test_summary_has_one_row_per_file(checked_result, tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    out = tmp_path / "report.xlsx"
    cfp.export_to_excel(checked_result, out)
    ws = openpyxl.load_workbook(out)["Summary"]

    header = [c.value for c in ws[3]]
    assert header == ["File", "Category", "Description", "Status", "Failures"], (
        "Passed/Total were dropped; Description says what each part is"
    )

    rows = {r[0]: r for r in ws.iter_rows(min_row=4, values_only=True)}
    assert set(rows) == {"CD-001659.iam", "CD-001624.ipt", "notes.xlsx"}
    assert rows["CD-001659.iam"][3] == "FAIL"
    assert "Engineer" in rows["CD-001659.iam"][4]
    assert "Vendor" in rows["CD-001624.ipt"][4]


def test_summary_carries_each_files_description(checked_result, tmp_path):
    """What the part IS, next to whether it passed."""
    openpyxl = pytest.importorskip("openpyxl")
    out = tmp_path / "report.xlsx"
    cfp.export_to_excel(checked_result, out)
    rows = {r[0]: r for r in
            openpyxl.load_workbook(out)["Summary"].iter_rows(min_row=4,
                                                             values_only=True)}
    expected = checked_result["info"]["properties"]["Description (File)"]
    assert expected, "the fixture should have a description to carry"
    assert rows["CD-001659.iam"][2] == expected
    assert rows["CD-001624.ipt"][2] == expected      # child inherits it here


def test_a_file_with_no_description_gets_a_blank_cell(checked_result, tmp_path):
    """No description is an empty cell, never the string 'None'."""
    openpyxl = pytest.importorskip("openpyxl")
    checked_result["info"]["properties"] = dict(
        checked_result["info"]["properties"], **{"Description (File)": None})
    out = tmp_path / "report.xlsx"
    cfp.export_to_excel(checked_result, out)
    rows = {r[0]: r for r in
            openpyxl.load_workbook(out)["Summary"].iter_rows(min_row=4,
                                                             values_only=True)}
    assert rows["CD-001659.iam"][2] in ("", None)
    # notes.xlsx has no properties at all — same treatment.
    assert rows["notes.xlsx"][2] in ("", None)


def test_a_file_with_no_rule_set_exports_as_skip_not_pass(checked_result, tmp_path):
    """The whole point of SKIP — it must never read as compliant in the sheet."""
    openpyxl = pytest.importorskip("openpyxl")
    out = tmp_path / "report.xlsx"
    cfp.export_to_excel(checked_result, out)
    wb = openpyxl.load_workbook(out)

    summary = {r[0]: r for r in wb["Summary"].iter_rows(min_row=4, values_only=True)}
    assert summary["notes.xlsx"][3] == "SKIP"
    assert "No rule set" in summary["notes.xlsx"][4]

    detail = [r for r in wb["Detail"].iter_rows(min_row=4, values_only=True)
              if r[0] == "notes.xlsx"]
    assert len(detail) == 1
    assert detail[0][3] == "SKIP"
    assert "No rule set" in detail[0][5]


def test_detail_has_one_row_per_property_checked(checked_result, tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    out = tmp_path / "report.xlsx"
    cfp.export_to_excel(checked_result, out)
    ws = openpyxl.load_workbook(out)["Detail"]

    rows = [r for r in ws.iter_rows(min_row=4, values_only=True)]
    top = [r for r in rows if r[0] == "CD-001659.iam"]
    assert len(top) == checked_result["report"]["total"]

    failing = [r for r in top if r[3] == "FAIL"]
    assert sorted(r[2] for r in failing) == ["Engineer", "Project"]
    assert all(r[5] for r in failing), "every failing row must say why"


def test_export_of_a_single_file_check_has_no_child_rows(properties, rules, tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    result = {
        "file_name": "CD-001659.iam",
        "info": {"properties": properties, "note": None},
        "children": [], "children_error": None, "recursive": False,
        **cfp.evaluate_against_rules(properties, properties["Category Name"], rules),
    }
    out = tmp_path / "single.xlsx"
    cfp.export_to_excel(result, out)
    ws = openpyxl.load_workbook(out)["Summary"]
    assert len(list(ws.iter_rows(min_row=4, values_only=True))) == 1


def test_default_export_path_is_named_for_the_file(tmp_path):
    path = cfp.default_export_path("CD-001659.iam", directory=tmp_path)
    assert path.parent == tmp_path
    assert path.suffix == ".xlsx"
    assert path.name.startswith("property-check_CD-001659_")


def test_exports_default_to_the_downloads_folder():
    """Matches where the MFG package builder and purchasing sheet land."""
    from pathlib import Path as _Path
    path = cfp.default_export_path("CD-001659.iam")
    assert path.parent == _Path.home() / "Downloads"


def test_export_creates_downloads_if_it_is_missing(checked_result, tmp_path,
                                                   monkeypatch):
    """A machine with no Downloads folder still gets its report."""
    pytest.importorskip("openpyxl")
    from pathlib import Path as _Path
    fake_home = tmp_path / "home"
    fake_home.mkdir()
    monkeypatch.setattr(_Path, "home", classmethod(lambda cls: fake_home))

    target = cfp.default_export_path("CD-001659.iam")
    assert not target.parent.exists()
    cfp.export_to_excel(checked_result, target)
    assert target.exists()


def test_export_to_a_locked_file_raises_a_useful_message(checked_result, tmp_path,
                                                         monkeypatch):
    pytest.importorskip("openpyxl")
    import openpyxl.workbook.workbook as wbmod

    def boom(self, filename):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr(wbmod.Workbook, "save", boom)
    with pytest.raises(RuntimeError, match="open in Excel"):
        cfp.export_to_excel(checked_result, tmp_path / "locked.xlsx")


def test_export_creates_the_target_directory(checked_result, tmp_path):
    pytest.importorskip("openpyxl")
    out = tmp_path / "nested" / "deeper" / "report.xlsx"
    cfp.export_to_excel(checked_result, out)
    assert out.exists()


def test_gui_export_button_unlocks_only_after_a_successful_check(checked_result):
    tk = pytest.importorskip("tkinter")
    pytest.importorskip("openpyxl")
    try:
        root = tk.Tk()
    except tk.TclError:
        pytest.skip("no display available")
    root.withdraw()
    try:
        from gui.file_property_check import run_gui
        run_gui(parent=root)
        root.update_idletasks()
        window = next(w for w in root.winfo_children()
                      if isinstance(w, tk.Toplevel))
        button = window.export_button_for_test

        assert str(button["state"]) == "disabled", "nothing to export yet"

        window.finish_for_test(checked_result, None)
        root.update_idletasks()
        assert str(button["state"]) == "normal"

        window.finish_for_test(None, "Vault sign-in failed")
        root.update_idletasks()
        assert str(button["state"]) == "disabled", (
            "a failed check must not leave a stale result exportable"
        )
    finally:
        root.destroy()


def test_markdown_report_says_pass_when_clean(properties, rules):
    clean = dict(properties, **{"Description (File)": "adapter plate assembly"})
    evaluated = cfp.evaluate_against_rules(clean, "Assembly - Engineering", rules)
    md = cfp.format_markdown_report({
        "file_name": "CD-001659.iam",
        "info": {"properties": clean, "note": None},
        "children": [], "children_error": None, "recursive": False,
        **evaluated,
    })
    assert "**PASS**" in md


# --------------------------------------------------------------------------- session reuse

def test_check_file_name_reuses_a_supplied_session(monkeypatch, tmp_path):
    """With api and vault_id supplied, no sign-in happens — proven two ways:
    VaultRestAPI is replaced with a constructor that raises if called, and
    config_path points at a file that does not exist, so falling through to
    the sign-in branch fails loudly (an assertion) rather than quietly
    reaching the real Vault server.
    """
    class BoomAPI:
        async def create_session(self, **_kw):
            raise AssertionError("must not sign in when a session is supplied")

    def boom_vault_rest_api(**_kw):
        raise AssertionError(
            "must not construct a new VaultRestAPI when a session is supplied")

    monkeypatch.setattr(cfp, "VaultRestAPI", boom_vault_rest_api)

    async def fake_fetch_file(api, vault_id, name):
        assert vault_id == "V1"
        return {"record": {}, "file_version_id": "9", "file_id": "8",
                "properties": {"Category Name": "Part"}, "note": ""}

    monkeypatch.setattr(cfp, "fetch_file", fake_fetch_file)

    result = asyncio.run(cfp.check_file_name(
        "CD-001659.ipt", api=BoomAPI(), vault_id="V1", recursive=False,
        config_path=tmp_path / "does-not-exist.json"))

    assert result["file_name"] == "CD-001659.ipt"
    assert result["info"]["file_version_id"] == "9"


def test_check_file_name_reuse_forwards_the_session_into_the_recursive_walk(
        monkeypatch, tmp_path):
    """Task 5's engine calls this with recursive=True and a reused session —
    the CAD BOM walk must reuse the same api/vault_id too, not just the
    top-level file lookup.
    """
    class BoomAPI:
        async def create_session(self, **_kw):
            raise AssertionError("must not sign in when a session is supplied")

    def boom_vault_rest_api(**_kw):
        raise AssertionError(
            "must not construct a new VaultRestAPI when a session is supplied")

    monkeypatch.setattr(cfp, "VaultRestAPI", boom_vault_rest_api)

    reuse_api = BoomAPI()

    async def fake_fetch_file(api, vault_id, name):
        assert api is reuse_api
        assert vault_id == "V1"
        return {"record": {}, "file_version_id": "9", "file_id": "8",
                "properties": {"Category Name": "Assembly - Engineering"},
                "note": ""}

    seen = {}

    async def fake_fetch_cad_children(api, vault_id, file_version_id, *,
                                       limit=500, **_kw):
        seen["api"] = api
        seen["vault_id"] = vault_id
        seen["file_version_id"] = file_version_id
        return []

    monkeypatch.setattr(cfp, "fetch_file", fake_fetch_file)
    monkeypatch.setattr(cfp, "fetch_cad_children", fake_fetch_cad_children)

    result = asyncio.run(cfp.check_file_name(
        "CD-001659.iam", api=reuse_api, vault_id="V1", recursive=True,
        config_path=tmp_path / "does-not-exist.json"))

    assert seen["api"] is reuse_api
    assert seen["vault_id"] == "V1"
    assert seen["file_version_id"] == "9"
    assert result["children"] == []
    assert result["children_error"] is None


def _write_config(tmp_path):
    cfg = tmp_path / "config.json"
    cfg.write_text(json.dumps({"vault": {
        "servername": "s", "username": "u", "password": "p", "database": "d",
    }}), encoding="utf-8")
    return cfg


class _SigningAPI:
    """Records whether create_session was called and hands back a vault id."""

    def __init__(self):
        self.signed_in = False

    async def create_session(self, **_kw):
        self.signed_in = True
        return {"error": False, "data": {"vaultInformation": {"id": "V-SIGNED-IN"}}}


def test_check_file_name_signs_in_when_no_session_supplied(monkeypatch, tmp_path):
    """The default path (no api/vault_id) must still sign in from config.json."""
    fake_api = _SigningAPI()
    monkeypatch.setattr(cfp, "VaultRestAPI", lambda **_kw: fake_api)

    seen_vault_id = {}

    async def fake_fetch_file(api, vault_id, name):
        seen_vault_id["value"] = vault_id
        return {"record": {}, "file_version_id": "1", "file_id": "1",
                "properties": {"Category Name": "Part"}, "note": ""}

    monkeypatch.setattr(cfp, "fetch_file", fake_fetch_file)

    asyncio.run(cfp.check_file_name(
        "CD-001659.ipt", config_path=_write_config(tmp_path), recursive=False))

    assert fake_api.signed_in, "omitting api/vault_id must sign in as before"
    assert seen_vault_id["value"] == "V-SIGNED-IN"


def test_check_file_name_raises_when_vault_id_is_blank_despite_an_api():
    """api set, vault_id blank is not a hypothetical slip — it is reachable
    from a sign-in that reported success but returned a payload this module
    didn't recognise. Silently re-authenticating would hide that bug behind
    a quiet no-op, so this must raise instead of falling through to a sign-in.
    """
    class BoomAPI:
        async def create_session(self, **_kw):
            raise AssertionError("must raise before ever touching the api")

    with pytest.raises(ValueError,
                       match="api and vault_id must be supplied together"):
        asyncio.run(cfp.check_file_name(
            "CD-001659.ipt", api=BoomAPI(), vault_id="", recursive=False))


def test_check_file_name_raises_when_api_is_none_despite_a_vault_id():
    """vault_id set, api blank must also raise rather than sign in — half a
    session is an error, not a hint to authenticate a fresh one."""
    with pytest.raises(ValueError,
                       match="api and vault_id must be supplied together"):
        asyncio.run(cfp.check_file_name(
            "CD-001659.ipt", api=None, vault_id="V1", recursive=False))
