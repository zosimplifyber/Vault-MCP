"""
Check a Vault FILE's properties (iProperties) against a rule set.

Type a file name — ``CD-001659.iam`` — and this pulls that file's properties
from Vault and reports which ones are out of compliance.

Rules come from ``file_property_rules.json`` at the project root, keyed by the
file's Category Name ("Assembly - Engineering", "Part - Engineering",
"Part - Purchased", …). That file is re-read on every run, so edits take effect
immediately.

This is the file-side twin of ``check_item_properties.py``. The rule engine is
shared — imported from that module, not duplicated — but the fetch layer and
property names are different, because Vault files and Vault items expose
different properties (files have ``Description (File)`` and ``CAD Category``;
items have ``Description (Item,CO)`` and ``Units``).

Run:
    python scripts/check_file_properties.py CD-001659.iam
    python scripts/check_file_properties.py CD-001659.iam --recursive
    python scripts/check_file_properties.py CD-001659.iam --json
    python scripts/check_file_properties.py            # launches the GUI

Exit codes: 0 = everything passed, 1 = one or more failures, 2 = no rule set
matched the file's category.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# Make the project root and scripts/ importable so we can reuse VaultRestAPI
# and the item tool's rule engine.
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

from vault_rest_api import VaultRestAPI  # noqa: E402

# The rule engine is shared with the item checker. One implementation, one
# place to fix it. check_item_properties does not import this module, so there
# is no import cycle.
from check_item_properties import (  # noqa: E402
    check_properties,
    load_json,
    resolve_category,
)


CONFIG_PATH = PROJECT_ROOT / "config.json"
DEFAULT_RULES_PATH = PROJECT_ROOT / "file_property_rules.json"

# One re-fetch per child is issued when grading at the latest version, and
# vault_rest_api opens a fresh httpx client per request — so nothing else caps
# this. Matches vault_state.MAX_CONCURRENCY: enough to finish a 300-part BOM in
# seconds, few enough that Vault isn't hit with 300 sockets at once.
MAX_CONCURRENCY = 8


# ---------------------------------------------------------------------------
# Property extraction from Vault file responses
# ---------------------------------------------------------------------------
# A file-version record carries its user properties as
#     "properties": [{"propertyDefinitionId": "77", "value": "Released"}, ...]
# with no names attached. The names live in the response's
#     "included": {"propertyDefinition": {"77": {"displayName": "State", ...}}}
# map. Everything below turns that pair into a flat {display name: value} dict
# keyed the same way file_property_rules.json is.

# System fields that appear at the top level of a file-version record. Mapped
# onto the display names the rules use so a file still reports a sane Category
# Name / Revision / State even if the properties array is missing.
_RECORD_FIELD_NAMES: dict[str, str] = {
    "name": "File Name",
    "category": "Category Name",
    "revision": "Revision",
    "state": "State",
}


def extract_definition_index(payload: Any) -> dict[str, dict[str, Any]]:
    """Return ``{propertyDefinitionId: definition}`` from a response's ``included``.

    Vault embeds the definition of every property it returned, so a single
    request yields both the values and their names.
    """
    if not isinstance(payload, dict):
        return {}
    included = payload.get("included")
    if not isinstance(included, dict):
        return {}
    defs = included.get("propertyDefinition")
    if not isinstance(defs, dict):
        return {}
    return {str(k): v for k, v in defs.items() if isinstance(v, dict)}


def flatten_file_properties(
    record: dict[str, Any],
    defs: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """Return a flat ``{display name: value}`` dict for one file-version record.

    Values from the ``properties`` array win over the record's top-level system
    fields, and a populated value always wins over an empty one — Vault
    sometimes reports the same property twice (current and historical).
    """
    flat: dict[str, Any] = {}

    def absorb(name: Any, value: Any) -> None:
        if not isinstance(name, str) or not name.strip():
            return
        key = name.strip()
        if key in flat and _is_blank(value) and not _is_blank(flat[key]):
            return
        flat[key] = value

    for field, name in _RECORD_FIELD_NAMES.items():
        if field in record:
            absorb(name, record.get(field))

    for entry in record.get("properties") or []:
        if not isinstance(entry, dict):
            continue
        pid = str(entry.get("propertyDefinitionId", ""))
        definition = defs.get(pid) or {}
        display = definition.get("displayName")
        if not display:
            # No definition came back for this id — skip rather than invent a
            # name. A rule keyed on it will report "missing", which is honest.
            continue
        # Historical twins ("State (Historical)") are separate definitions, so
        # they land under their own names and never shadow the live value.
        absorb(display, entry.get("value"))
        system = definition.get("systemName")
        if system and system != display and system not in flat:
            flat[system] = entry.get("value")

    _fill_in_historical_state(record, flat)
    return flat


def _fill_in_historical_state(record: dict[str, Any], flat: dict[str, Any]) -> None:
    """Recover ``State`` on a pinned (non-latest) file version.

    A CAD BOM child is whatever version its parent assembly pins, which is
    often not the latest. Vault returns those with the live ``State`` blank —
    both the top-level field and property 77 — and reports the state that
    version is actually in via ``lifecycleState`` and the ``State (Historical)``
    twin (property 78).

    Without this, every pinned child reports "State missing (required)" while
    the same file checked on its own reports Released.
    """
    if not _is_blank(flat.get("State")):
        return

    lifecycle = record.get("lifecycleState")
    if isinstance(lifecycle, dict):
        for key in ("name", "displayName"):
            if not _is_blank(lifecycle.get(key)):
                flat["State"] = lifecycle[key]
                return

    historical = flat.get("State (Historical)")
    if not _is_blank(historical):
        flat["State"] = historical


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


# ---------------------------------------------------------------------------
# Vault lookup
# ---------------------------------------------------------------------------

def _results(payload: Any) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    rows = payload.get("results")
    if isinstance(rows, list):
        return [r for r in rows if isinstance(r, dict)]
    return []


def select_file_record(
    records: list[dict[str, Any]],
    file_name: str,
) -> tuple[dict[str, Any], str | None]:
    """Pick the record matching ``file_name`` and return ``(record, note)``.

    An exact case-insensitive name match always wins. Failing that, the first
    result is used and ``note`` explains the ambiguity so the report can say so
    rather than quietly checking the wrong file.
    """
    if not records:
        raise RuntimeError(f"No files found matching '{file_name}'.")

    wanted = file_name.strip().lower()
    exact = [r for r in records if str(r.get("name", "")).strip().lower() == wanted]
    if exact:
        note = None
        if len(exact) > 1:
            note = (
                f"{len(exact)} versions matched '{file_name}' exactly; "
                "checking the first (latest)."
            )
        return exact[0], note

    note = (
        f"No file is named exactly '{file_name}'. Checking "
        f"'{records[0].get('name')}'"
        + (f" ({len(records)} results matched)." if len(records) > 1 else ".")
    )
    return records[0], note


async def fetch_file(
    api: VaultRestAPI,
    vault_id: str,
    file_name: str,
) -> dict[str, Any]:
    """Look up a file by name and return its resolved properties.

    Returns ``{"record", "file_version_id", "file_id", "properties", "note"}``.
    Raises ``RuntimeError`` when the file cannot be found or Vault errors.
    """
    resp = await api.search_file_versions(
        vault_id=vault_id, query=file_name, prop_def_ids="all", limit=25,
    )
    if resp["error"]:
        raise RuntimeError(f"File search failed: {resp['data']}")

    payload = resp.get("data") or {}
    record, note = select_file_record(_results(payload), file_name)
    defs = extract_definition_index(payload)

    if not defs and (record.get("properties") or []):
        # Vault returned values without their definitions. Fall back to the
        # standalone definitions endpoint rather than report a file whose
        # properties we cannot name as compliant.
        defs = await fetch_definition_index(api, vault_id)

    return {
        "record": record,
        "file_version_id": str(record.get("id") or ""),
        "file_id": str((record.get("file") or {}).get("id") or ""),
        "properties": flatten_file_properties(record, defs),
        "note": note,
    }


_DEFS_CACHE: dict[str, dict[str, dict[str, Any]]] = {}


async def fetch_definition_index(
    api: VaultRestAPI, vault_id: str
) -> dict[str, dict[str, Any]]:
    """Fallback ``{id: definition}`` map from ``/property-definitions``, cached.

    Only used when a response omits ``included.propertyDefinition``.
    """
    cached = _DEFS_CACHE.get(vault_id)
    if cached is not None:
        return cached
    resp = await api.get_property_definitions(vault_id=vault_id, limit=500)
    if resp["error"]:
        raise RuntimeError(
            f"Could not load property definitions, so file properties cannot "
            f"be named: {resp['data']}"
        )
    rows = _results(resp.get("data") or {})
    index = {str(r["id"]): r for r in rows if r.get("id")}
    _DEFS_CACHE[vault_id] = index
    return index


async def fetch_cad_children(
    api: VaultRestAPI,
    vault_id: str,
    file_version_id: str,
    *,
    limit: int = 500,
    use_latest: bool = True,
) -> list[dict[str, Any]]:
    """Return one entry per child file in the CAD BOM.

    ``/uses`` enriches every child with its properties in the same call when
    ``prop_def_ids`` is passed, so discovering the children is a single request
    no matter how many there are.

    The versions ``/uses`` returns are the ones the parent assembly **pins**,
    which are often not the newest. Checking those answers "is this assembly,
    as pinned, compliant?" — but it means a file you just fixed keeps reporting
    its old failures, because the parent still references the version from
    before the fix. With ``use_latest`` (the default) each child is re-fetched
    at its newest version, so the report answers "which files still need
    fixing?" instead. Pass ``use_latest=False`` to audit exactly what the
    assembly consumes.

    Each entry: ``{"file_name", "file_version_id", "file_id", "assoc_type",
    "properties", "pinned_revision", "error"}``.
    """
    resp = await api.get_file_uses(
        vault_id=vault_id,
        file_version_id=file_version_id,
        prop_def_ids="all",
        limit=limit,
    )
    if resp["error"]:
        raise RuntimeError(f"CAD BOM walk failed: {resp['data']}")

    payload = resp.get("data") or {}
    defs = extract_definition_index(payload)
    rows = _results(payload)
    if rows and not defs:
        defs = await fetch_definition_index(api, vault_id)

    children: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        child = row.get("childFile")
        if not isinstance(child, dict):
            continue
        version_id = str(child.get("id") or "")
        master_id = str((child.get("file") or {}).get("id") or "")
        # A child used more than once in the assembly comes back once per
        # occurrence; check it once. Key on the File master where we have it,
        # so the same file pinned at two versions still collapses to one row.
        key = master_id or version_id
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        properties = flatten_file_properties(child, defs)
        children.append({
            "file_name": str(child.get("name") or "(unknown)"),
            "file_version_id": version_id,
            "file_id": master_id,
            "assoc_type": str(row.get("fileAssocType") or ""),
            "properties": properties,
            "pinned_revision": str(properties.get("Revision") or ""),
            "error": None,
        })

    if use_latest and children:
        sem = asyncio.Semaphore(MAX_CONCURRENCY)

        async def upgrade(child: dict[str, Any]) -> None:
            async with sem:
                await _upgrade_child_to_latest(api, vault_id, child)

        await asyncio.gather(*[upgrade(c) for c in children])

    children.sort(key=lambda c: c["file_name"].lower())
    return children


async def _upgrade_child_to_latest(
    api: VaultRestAPI, vault_id: str, child: dict[str, Any]
) -> None:
    """Replace a child's pinned properties with its newest version, in place.

    Sets ``child["error"]`` if the newest version can't be resolved. Failing
    loudly beats silently reporting the pinned version as though it were
    current — that mismatch is the whole reason this exists.
    """
    name = child["file_name"]
    resp = await api.search_file_versions(
        vault_id=vault_id, query=name, prop_def_ids="all",
        latest_only=True, limit=25,
    )
    if resp["error"]:
        child["error"] = f"Could not load the latest version of {name}: {resp['data']}"
        return

    payload = resp.get("data") or {}
    records = _results(payload)
    # Prefer the record belonging to the same File master. Two files in
    # different folders can share a name; the master id can't collide.
    master_id = child.get("file_id") or ""
    if master_id:
        same_file = [
            r for r in records
            if str((r.get("file") or {}).get("id") or "") == master_id
        ]
        if same_file:
            records = same_file

    try:
        record, _note = select_file_record(records, name)
    except RuntimeError as exc:
        child["error"] = f"Could not resolve the latest version of {name}: {exc}"
        return

    defs = extract_definition_index(payload)
    if not defs and (record.get("properties") or []):
        defs = await fetch_definition_index(api, vault_id)

    child["properties"] = flatten_file_properties(record, defs)
    child["file_version_id"] = str(record.get("id") or child["file_version_id"])


# ---------------------------------------------------------------------------
# Pipeline (shared by CLI and GUI)
# ---------------------------------------------------------------------------

def evaluate_against_rules(
    properties: dict[str, Any],
    raw_category: str,
    rules: dict[str, Any],
) -> dict[str, Any]:
    """Resolve the rule set for ``raw_category`` and run it against ``properties``.

    ``report`` is None when no rule set matches — which reports as SKIP, never
    as a pass.
    """
    category = resolve_category(raw_category, rules)
    report = None
    if category:
        report = check_properties(properties, rules["categories"][category])
    return {
        "category_raw": raw_category,
        "category_resolved": category,
        "report": report,
    }


async def check_file_name(
    file_name: str,
    *,
    config_path: Path = CONFIG_PATH,
    rules_path: Path = DEFAULT_RULES_PATH,
    category_override: str = "",
    recursive: bool = False,
    bom_limit: int = 500,
    api: Any = None,
    vault_id: str = "",
) -> dict[str, Any]:
    """Sign in, look up the file, run the rules. Returns a result dict.

    Keys: ``file_name``, ``info``, ``rules``, ``category_raw``,
    ``category_resolved``, ``report``, ``children``, ``children_error``,
    ``recursive``. Raises ``RuntimeError`` for any fatal Vault / config error
    so the caller can surface the message.

    Pass ``api`` and ``vault_id`` together to reuse an already-authenticated
    session (the GUI has one); omit both to sign in from ``config_path`` as
    before.
    """
    rules = load_json(rules_path)

    if api is not None and vault_id:
        # Caller handed us a live session — don't sign in a second time.
        pass
    else:
        cfg = load_json(config_path)
        vault_cfg = cfg.get("vault") or {}
        for key in ("servername", "username", "password", "database"):
            if not vault_cfg.get(key):
                raise RuntimeError(f"config.json is missing vault.{key}")

        api = VaultRestAPI(servername=vault_cfg["servername"])
        sign_in = await api.create_session(
            database=vault_cfg["database"],
            username=vault_cfg["username"],
            password=vault_cfg["password"],
        )
        if sign_in["error"]:
            raise RuntimeError(f"Vault sign-in failed: {sign_in['data']}")

        vault_id = str(
            (sign_in["data"].get("vaultInformation") or {}).get("id", "")
            or sign_in["data"].get("vaultId", "")
            or ""
        )

    info = await fetch_file(api, vault_id, file_name)

    raw_category = category_override or info["properties"].get("Category Name") or ""
    eval_top = evaluate_against_rules(info["properties"], raw_category, rules)

    children: list[dict[str, Any]] = []
    children_error: str | None = None
    if recursive:
        if not info["file_version_id"]:
            children_error = (
                "Cannot walk the CAD BOM — Vault returned no file-version ID."
            )
        else:
            try:
                rows = await fetch_cad_children(
                    api, vault_id, info["file_version_id"], limit=bom_limit
                )
            except RuntimeError as exc:
                children_error = str(exc)
                rows = []

            for row in rows:
                base = {
                    "file_name": row["file_name"],
                    "file_version_id": row["file_version_id"],
                    "file_id": row.get("file_id", ""),
                    "assoc_type": row["assoc_type"],
                    "pinned_revision": row.get("pinned_revision", ""),
                    "properties": row["properties"],
                }
                if row.get("error"):
                    # Couldn't resolve this child — report it rather than
                    # grading whatever stale data we happen to hold.
                    children.append({
                        **base,
                        "category_raw": row["properties"].get("Category Name") or "",
                        "category_resolved": None,
                        "report": None,
                        "error": row["error"],
                    })
                    continue

                child_cat = row["properties"].get("Category Name") or ""
                children.append({
                    **base,
                    **evaluate_against_rules(row["properties"], child_cat, rules),
                    "error": None,
                })

    return {
        "file_name": file_name,
        "info": info,
        "rules": rules,
        "category_raw": eval_top["category_raw"],
        "category_resolved": eval_top["category_resolved"],
        "report": eval_top["report"],
        "children": children,
        "children_error": children_error,
        "recursive": recursive,
    }


# ---------------------------------------------------------------------------
# Status helpers
# ---------------------------------------------------------------------------

def child_status(child: dict[str, Any]) -> str:
    if child.get("error"):
        return "ERROR"
    if not child.get("category_resolved"):
        return "SKIP"
    return "PASS" if (child.get("report") or {}).get("failed", 0) == 0 else "FAIL"


def result_exit_code(result: dict[str, Any]) -> int:
    report = result.get("report")
    if report and report.get("failed", 0) > 0:
        return 1
    if any(child_status(c) in ("FAIL", "ERROR") for c in result.get("children") or []):
        return 1
    if not result.get("category_resolved"):
        return 2
    return 0


def _display(value: Any) -> str:
    return "(empty)" if _is_blank(value) else str(value)


# ---------------------------------------------------------------------------
# Terminal reporting
# ---------------------------------------------------------------------------

def _supports_color() -> bool:
    return sys.stdout.isatty()


def _c(code: str, text: str) -> str:
    if not _supports_color():
        return text
    return f"\x1b[{code}m{text}\x1b[0m"


def print_report(result: dict[str, Any]) -> None:
    info = result["info"]
    props = info["properties"]
    report = result["report"]

    print()
    print(_c("1;36", f"File: {result['file_name']}"))
    print(f"  Name          : {props.get('File Name', '(unknown)')}")
    print(f"  Title         : {_display(props.get('Title'))}")
    print(f"  Description   : {_display(props.get('Description (File)'))}")
    print(f"  Category      : {props.get('Category Name', '(unknown)')}")
    print(f"  Revision/State: {props.get('Revision', '?')} / {props.get('State', '?')}")
    if info.get("note"):
        print(_c("33", f"  Note: {info['note']}"))
    print()

    if not result["category_resolved"]:
        print(_c("31", "No rule set matched this file's category "
                       f"({result['category_raw'] or 'none'})."))
        print("  Edit file_property_rules.json or pass --category to override.")
        return

    print(_c("1", f"Checking against rules: {result['category_resolved']}"))
    print()

    results = report["results"]
    name_w = max((len(r["property"]) for r in results), default=18)
    name_w = max(name_w, 18)

    for r in results:
        mark = _c("32", "PASS") if r["passed"] else _c("31", "FAIL")
        value_str = _display(r["value"])
        if value_str == "(empty)":
            value_str = _c("90", value_str)
        print(f"  [{mark}]  {r['property']:<{name_w}}  {value_str}")
        for f in r["failures"]:
            print(f"          {_c('31', '> ' + f)}")

    print()
    summary = f"{report['passed']}/{report['total']} properties passed"
    print(_c("1;32" if report["failed"] == 0 else "1;31", summary))


def print_children_report(children: list[dict[str, Any]], show_all: bool) -> None:
    if not children:
        print(_c("33", "No CAD BOM children found."))
        return

    statuses = [child_status(c) for c in children]
    print()
    print(_c("1", f"Children ({len(children)} files)"))
    print()

    for c, status in zip(children, statuses):
        tag = {"PASS": "32", "FAIL": "31", "SKIP": "33", "ERROR": "31"}[status]
        rep = c.get("report") or {}
        score = (
            f"{rep.get('passed', 0)}/{rep.get('total', 0)}"
            if status in ("PASS", "FAIL") else ""
        )
        cat = c.get("category_resolved") or c.get("category_raw") or "(no rule set)"
        # Pad before colouring — ANSI escapes count toward a format width and
        # would blow the column alignment apart.
        print(f"  [{_c(tag, f'{status:<5}')}] {c['file_name']:<24} "
              f"{cat:<24} {score:>7}")

        if c.get("error"):
            print(f"          {_c('31', '> ' + c['error'])}")
            continue

        if status == "FAIL" or show_all:
            for r in rep.get("results") or []:
                if r["passed"] and not show_all:
                    continue
                inner_tag = "32" if r["passed"] else "31"
                inner_mark = "PASS" if r["passed"] else "FAIL"
                print(f"          [{_c(inner_tag, inner_mark)}] "
                      f"{r['property']:<24} {_display(r['value'])}")
                for f in r["failures"]:
                    print(f"                 {_c('31', '> ' + f)}")

    print()
    summary = (
        f"Children: {statuses.count('PASS')} pass, {statuses.count('FAIL')} fail, "
        f"{statuses.count('SKIP')} skipped (no rule set), "
        f"{statuses.count('ERROR')} errored."
    )
    failed = statuses.count("FAIL") + statuses.count("ERROR")
    print(_c("1;31" if failed else "1;32", summary))


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------

def _escape(text: str) -> str:
    """Escape pipes so a value or message can't break out of a table cell.

    Failure messages quote the rule that fired, and alternation regexes like
    ``(CD|SF|MFG|DT)`` are full of pipes.
    """
    return str(text).replace("|", "\\|")


def _short(value: Any, *, width: int = 40) -> str:
    if _is_blank(value):
        return "_(empty)_"
    s = str(value).strip()
    if len(s) > width:
        s = s[: width - 1] + "…"
    return _escape(s)


def _problem(rule_result: dict[str, Any]) -> str:
    return _escape("; ".join(rule_result.get("failures") or []) or "non-compliant")


def format_markdown_report(result: dict[str, Any]) -> str:
    """Render the file compliance check as Markdown."""
    info = result.get("info") or {}
    props = info.get("properties") or {}
    report = result.get("report") or {}
    children = result.get("children") or []
    category = (
        result.get("category_resolved")
        or result.get("category_raw")
        or "(unresolved)"
    )

    lines: list[str] = []
    lines.append(f"# File Property Compliance — `{result.get('file_name')}`")
    lines.append("")
    lines.append("| Field | Value |")
    lines.append("|---|---|")
    lines.append(f"| Name | `{props.get('File Name', '(unknown)')}` |")
    lines.append(f"| Title | {_short(props.get('Title'), width=80)} |")
    lines.append(f"| Description | {_short(props.get('Description (File)'), width=80)} |")
    lines.append(f"| Category | {_short(category, width=40)} |")
    lines.append(f"| Revision / State | `{props.get('Revision', '?')}` / "
                 f"`{props.get('State', '?')}` |")
    if info.get("note"):
        lines.append(f"| Note | {info['note']} |")
    lines.append("")

    if not result.get("category_resolved"):
        lines.append("> No rule set matched this file's category. Add one to "
                     "`file_property_rules.json` or pass an explicit category.")
        lines.append("")
    elif report.get("failed", 0) == 0:
        lines.append(f"**PASS** — {report.get('passed', 0)}/{report.get('total', 0)} "
                     "properties compliant.")
        lines.append("")
    else:
        lines.append(f"**FAIL** — {report.get('passed', 0)}/{report.get('total', 0)} "
                     "properties compliant.")
        lines.append("")
        lines.append("| Property | Current Value | Problem |")
        lines.append("|---|---|---|")
        for r in report.get("results") or []:
            if r.get("passed"):
                continue
            lines.append(f"| `{r['property']}` | {_short(r.get('value'))} "
                         f"| {_problem(r)} |")
        lines.append("")

    if result.get("recursive"):
        lines.append("## CAD BOM children")
        lines.append("")
        if result.get("children_error"):
            lines.append(f"> {result['children_error']}")
            lines.append("")
        elif not children:
            lines.append("_(No child files found.)_")
            lines.append("")
        else:
            statuses = [child_status(c) for c in children]
            lines.append(
                f"**{len(children)} child file(s):** {statuses.count('PASS')} pass, "
                f"{statuses.count('FAIL')} fail, "
                f"{statuses.count('SKIP')} skipped (no rule set), "
                f"{statuses.count('ERROR')} errored."
            )
            lines.append("")
            lines.append("| Status | File | Category | Score |")
            lines.append("|---|---|---|---|")
            for c, st in zip(children, statuses):
                rep = c.get("report") or {}
                score = (
                    f"{rep.get('passed', 0)}/{rep.get('total', 0)}"
                    if st in ("PASS", "FAIL") else "—"
                )
                cat = c.get("category_resolved") or c.get("category_raw") or "(no rule set)"
                lines.append(f"| {st} | `{c.get('file_name')}` | {cat} | {score} |")
            lines.append("")

            offenders = [c for c, st in zip(children, statuses) if st in ("FAIL", "ERROR")]
            if offenders:
                lines.append("### Child failures")
                lines.append("")
                for c in offenders:
                    if c.get("error"):
                        lines.append(f"- **`{c['file_name']}`** — error: {c['error']}")
                        continue
                    bad = [r for r in ((c.get("report") or {}).get("results") or [])
                           if not r.get("passed")]
                    if not bad:
                        continue
                    lines.append(f"- **`{c['file_name']}`** "
                                 f"({c.get('category_resolved') or '?'})")
                    for r in bad:
                        lines.append(f"    - `{r['property']}` = "
                                     f"{_short(r.get('value'))} → {_problem(r)}")
                lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
# Palette matches bom_purchasing.py so the compliance workbook sits alongside
# the purchasing sheets without looking like it came from somewhere else.

XL_DARK_BLUE = "1F3864"
XL_MID_BLUE = "2E75B6"
XL_PALE_BLUE = "EAF3FB"
XL_LIGHT_GRAY = "F2F2F2"
XL_GRAY_BDR = "CCCCCC"
XL_DARK_GRAY = "888888"
XL_WHITE = "FFFFFF"
XL_PASS_FILL = "D8E4BC"      # olive — matches the purchasing sheet's OK state
XL_FAIL_FILL = "F8CBAD"      # rust-tinted, legible behind black text
XL_SKIP_FILL = "FCE4D6"

# Sheet 1 lists one row per property checked; sheet 2 rolls up the children.
_DETAIL_COLUMNS = [
    ("File", 26), ("Category", 24), ("Property", 22), ("Status", 9),
    ("Current Value", 46), ("Problem", 62),
]
_SUMMARY_COLUMNS = [
    ("File", 26), ("Category", 24), ("Description", 46), ("Status", 9),
    ("Failures", 52),
]


def default_export_path(file_name: str, *, directory: Path | None = None) -> Path:
    """Where an export lands when the caller doesn't name a file.

    Defaults to the user's Downloads folder, matching the MFG package builder
    and the purchasing sheet. ``export_to_excel`` creates the folder if it
    isn't there.
    """
    stem = Path(file_name).stem or "property-check"
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    folder = directory or (Path.home() / "Downloads")
    return Path(folder) / f"property-check_{stem}_{stamp}.xlsx"


def _export_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    """Flatten a result into one row per property checked, file by file.

    The top-level file first, then each child. Files with no rule set
    contribute a single SKIP row so they can't be mistaken for compliant.
    """
    rows: list[dict[str, Any]] = []

    def add(file_name: str, category: str, entry: dict[str, Any] | None,
            report: dict[str, Any] | None, error: str | None) -> None:
        if error:
            rows.append({
                "File": file_name, "Category": category, "Property": "—",
                "Status": "ERROR", "Current Value": "", "Problem": error,
            })
            return
        if report is None:
            rows.append({
                "File": file_name, "Category": category or "(unknown)",
                "Property": "—", "Status": "SKIP", "Current Value": "",
                "Problem": "No rule set for this category — nothing was checked.",
            })
            return
        for r in report.get("results") or []:
            rows.append({
                "File": file_name,
                "Category": category,
                "Property": r["property"],
                "Status": "PASS" if r["passed"] else "FAIL",
                "Current Value": "" if _is_blank(r["value"]) else str(r["value"]),
                "Problem": "; ".join(r.get("failures") or []),
            })

    info = result.get("info") or {}
    props = info.get("properties") or {}
    add(
        str(props.get("File Name") or result.get("file_name") or ""),
        result.get("category_resolved") or result.get("category_raw") or "",
        info, result.get("report"), None,
    )
    for child in result.get("children") or []:
        add(
            child.get("file_name", ""),
            child.get("category_resolved") or child.get("category_raw") or "",
            child, child.get("report"), child.get("error"),
        )
    return rows


def _summary_rows(result: dict[str, Any]) -> list[dict[str, Any]]:
    """One roll-up row per file: what it is, its status, and what failed.

    The Description column carries the file's own ``Description (File)``
    property, so the sheet says what each part actually *is* rather than only
    naming it. It is reported whether or not the description is currently
    gated.
    """
    rows: list[dict[str, Any]] = []

    def add(file_name: str, category: str, properties: dict[str, Any],
            status: str, report: dict[str, Any] | None,
            problem: str = "") -> None:
        failures = problem
        if not failures and report:
            failures = ", ".join(
                r["property"] for r in report.get("results") or []
                if not r["passed"]
            )
        description = (properties or {}).get("Description (File)")
        rows.append({
            "File": file_name,
            "Category": category or "(unknown)",
            "Description": "" if _is_blank(description) else str(description),
            "Status": status,
            "Failures": failures,
        })

    info = result.get("info") or {}
    props = info.get("properties") or {}
    report = result.get("report")
    if not result.get("category_resolved"):
        top_status = "SKIP"
    else:
        top_status = "PASS" if (report or {}).get("failed", 0) == 0 else "FAIL"
    add(
        str(props.get("File Name") or result.get("file_name") or ""),
        result.get("category_resolved") or result.get("category_raw") or "",
        props, top_status, report,
        "" if result.get("category_resolved") else "No rule set for this category.",
    )

    for child in result.get("children") or []:
        status = child_status(child)
        add(
            child.get("file_name", ""),
            child.get("category_resolved") or child.get("category_raw") or "",
            child.get("properties") or {},
            status, child.get("report"),
            child.get("error") or (
                "No rule set for this category." if status == "SKIP" else ""
            ),
        )
    return rows


def export_to_excel(result: dict[str, Any], output_path: str | Path) -> str:
    """Write the compliance report to a formatted .xlsx and return its path.

    Two sheets: **Summary** (one row per file) and **Detail** (one row per
    property checked). Both carry an autofilter and frozen headers so a long
    BOM walk stays navigable, and FAIL / SKIP rows are colour-coded.

    Raises ``RuntimeError`` if openpyxl is unavailable or the file cannot be
    written (most often: it's already open in Excel).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:                      # pragma: no cover
        raise RuntimeError(
            f"Excel export needs openpyxl — pip install openpyxl ({exc})"
        ) from exc

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    thin = Side(style="thin", color=XL_GRAY_BDR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_fills = {
        "PASS": PatternFill("solid", fgColor=XL_PASS_FILL),
        "FAIL": PatternFill("solid", fgColor=XL_FAIL_FILL),
        "SKIP": PatternFill("solid", fgColor=XL_SKIP_FILL),
        "ERROR": PatternFill("solid", fgColor=XL_FAIL_FILL),
    }

    def write_sheet(ws, columns, rows, title) -> None:
        n_cols = len(columns)
        last_col = get_column_letter(n_cols)

        ws.merge_cells(f"A1:{last_col}1")
        cell = ws["A1"]
        cell.value = title
        cell.font = Font(name="Arial", bold=True, color=XL_WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=XL_DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        ws.merge_cells(f"A2:{last_col}2")
        cell = ws["A2"]
        cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
        cell.font = Font(name="Arial", size=9, color=XL_DARK_GRAY, italic=True)
        cell.fill = PatternFill("solid", fgColor=XL_LIGHT_GRAY)
        cell.alignment = Alignment(horizontal="right", vertical="center")

        hdr_row = 3
        for ci, (name, width) in enumerate(columns, start=1):
            cell = ws.cell(row=hdr_row, column=ci, value=name)
            cell.font = Font(name="Arial", bold=True, color=XL_WHITE, size=10)
            cell.fill = PatternFill("solid", fgColor=XL_DARK_BLUE)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = width

        for ri, row in enumerate(rows, start=hdr_row + 1):
            fill = status_fills.get(row.get("Status", ""))
            for ci, (name, _w) in enumerate(columns, start=1):
                cell = ws.cell(row=ri, column=ci, value=row.get(name, ""))
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(
                    horizontal="center" if name == "Status" else "left",
                    vertical="center",
                    wrap_text=name in ("Problem", "Failures"),
                )
                cell.border = border
                # Colour the whole row so a failure is visible without
                # reading across to the Status column.
                if fill is not None:
                    cell.fill = fill
                elif ri % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=XL_PALE_BLUE)

        last_row = hdr_row + len(rows)
        if rows:
            ws.auto_filter.ref = f"A{hdr_row}:{last_col}{last_row}"
        ws.freeze_panes = f"A{hdr_row + 1}"

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    write_sheet(
        summary, _SUMMARY_COLUMNS, _summary_rows(result),
        f"Property Compliance — {result.get('file_name', '')}",
    )
    write_sheet(
        wb.create_sheet("Detail"), _DETAIL_COLUMNS, _export_rows(result),
        f"Property Compliance Detail — {result.get('file_name', '')}",
    )

    try:
        wb.save(str(output_path))
    except PermissionError as exc:
        raise RuntimeError(
            f"Could not write {output_path} — it's probably open in Excel. "
            f"Close it and try again. ({exc})"
        ) from exc
    except OSError as exc:
        raise RuntimeError(f"Could not write {output_path}: {exc}") from exc

    return str(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

async def run_cli(args: argparse.Namespace) -> int:
    try:
        result = await check_file_name(
            args.file_name,
            config_path=args.config,
            rules_path=args.rules,
            category_override=args.category,
            recursive=args.recursive,
        )
    except (RuntimeError, FileNotFoundError, ValueError) as exc:
        sys.exit(f"[ERROR] {exc}")

    if args.markdown:
        print(format_markdown_report(result))
    elif args.json:
        print(json.dumps({
            "file_name": args.file_name,
            "category_raw": result["category_raw"],
            "category_resolved": result["category_resolved"],
            "properties": result["info"]["properties"],
            "note": result["info"].get("note"),
            "report": result["report"],
            "children": [
                {k: v for k, v in c.items()
                 if k != "properties" or args.show_all_props}
                for c in result["children"]
            ],
            "children_error": result["children_error"],
            "available_categories": sorted(
                (result["rules"].get("categories") or {}).keys()
            ),
        }, indent=2, default=str))
    else:
        print_report(result)
        if args.show_all_props:
            print()
            print(_c("1", "All properties returned by Vault:"))
            for k in sorted(result["info"]["properties"]):
                print(f"  {k:<34} = {result['info']['properties'][k]}")
        if args.recursive:
            if result["children_error"]:
                print()
                print(_c("31", result["children_error"]))
            else:
                print_children_report(result["children"], args.show_all_props)

    if args.excel is not None:
        # --excel with no value means "pick a name for me".
        target = Path(args.excel) if args.excel else default_export_path(
            args.file_name)
        try:
            written = export_to_excel(result, target)
        except RuntimeError as exc:
            print()
            print(_c("31", f"Excel export failed: {exc}"))
            return 1
        print()
        print(_c("1;32", f"Excel report written to {written}"))

    return result_exit_code(result)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Check a Vault file's properties against "
                    "file_property_rules.json.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("file_name", nargs="?", default="",
                   help="Vault file name, e.g. CD-001659.iam. "
                        "Omit to launch the GUI.")
    p.add_argument("--config", type=Path, default=CONFIG_PATH,
                   help="Path to config.json.")
    p.add_argument("--rules", type=Path, default=DEFAULT_RULES_PATH,
                   help="Path to the rules JSON.")
    p.add_argument("--category", default="",
                   help="Override the file's Category Name "
                        "(e.g. 'Assembly - Engineering').")
    p.add_argument("--json", action="store_true",
                   help="Emit a machine-readable JSON report.")
    p.add_argument("--markdown", "-m", action="store_true",
                   help="Emit a Markdown report.")
    p.add_argument("--excel", "-x", nargs="?", const="", default=None,
                   metavar="PATH",
                   help="Also write the report to an .xlsx. Give a path, or "
                        "pass the flag bare to drop a timestamped file in "
                        "your Downloads folder.")
    p.add_argument("--show-all-props", action="store_true",
                   help="Also dump every property Vault returned.")
    p.add_argument("--recursive", "-r", action="store_true",
                   help="Walk the CAD BOM and check every child file too.")
    p.add_argument("--gui", action="store_true",
                   help="Launch the GUI even if a file name is supplied.")
    p.add_argument("--no-gui", action="store_true",
                   help="Force CLI mode — error out if no file name is given.")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    if args.gui or (not args.file_name and not args.no_gui):
        from gui.file_property_check import run_gui  # noqa: PLC0415
        run_gui(default_config=args.config, default_rules=args.rules)
        sys.exit(0)
    if not args.file_name:
        sys.exit("[ERROR] file_name is required in --no-gui mode.")
    sys.exit(asyncio.run(run_cli(args)))


if __name__ == "__main__":
    main()
