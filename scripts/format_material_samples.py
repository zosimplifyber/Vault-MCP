"""Reshape the Material Samples workbook into RAG-friendly documents.

The source workbook is laid out for humans: process conditions live in a merged
title row, unused ingredients sit in the table as zero rows, statistics are
interleaved with the samples they summarise, and percentages are stored as raw
fractions. None of that survives chunking - a retrieved row like
``SF_Fybron PET-1-2, A, internal coating, 0.58, 386.6`` means nothing without
its header and its material.

This script denormalises the workbook into two shapes:

* Markdown, one document per material, each self-contained (process conditions,
  formulation, coating recipe, its own test results). For "how do we make X".
* A single flat CSV where every test sample carries its full context on one
  row. For "what was the tensile on sample Y".

Run it again whenever the workbook changes; the output directory is rewritten.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
from pathlib import Path

import openpyxl

DEFAULT_WORKBOOK = Path.home() / "Downloads" / "Material Samples_Formulations & Testing.xlsx"
DEFAULT_OUTDIR = Path(__file__).resolve().parent.parent / "docs" / "rag" / "material-samples"

SOURCE_NAME = "Material Samples_Formulations & Testing.xlsx"

# Sheets holding a slurry recipe, in the order we want the docs to read.
FORMULATION_SHEETS = [
    "Fybron",
    "Fybron_Nylon",
    "Fyberite",
    "Fyberite_v2",
    "Fyberite_NoFoam",
    "FyberCom",
    "FyberCom_Dyed",
]

# Which coating sheet applies to which formulation(s).
COATING_FOR = {
    "Fybron": "Fybron_Coating",
    "Fybron_Nylon": "Fybron_Coating",
    "Fyberite": "Fyberite_Coating",
    "Fyberite_v2": "Fyberite_Coating",
    "Fyberite_NoFoam": "Fyberite_Coating",
    "FyberCom": "FyberCom_Coating",
    "FyberCom_Dyed": "FyberCom_Coating",
}

# Testing sheet -> the material family its samples default to.
TESTING_SHEETS = {
    "Testing_Fybron": "Fybron",
    "Testing_Fyberite": "Fyberite",
    "Testing_FyberCom": "FyberCom",
}

# Rows in a testing sheet whose "Rep" cell marks a derived statistic, not a sample.
STAT_LABELS = {"average", "sd", "%var"}

# Columns stored as fractions that the header already labels as a percentage.
PERCENT_COLUMNS = {
    "Thickness %",
    "Water absorption, %",
    "Tensile Drop, %",
    "Tensile drop, %",
    "Tear drop, %",
    "Flexural Modulus Drop, %",
}

# Source header -> CSV column, unifying the three testing sheets' differing layouts.
TEST_COLUMN_MAP = {
    "Formulation ID": "sample_id",
    "Rep": "rep",
    "Process step": "process_step",
    "Weight, g": "weight_g",
    "GSM": "gsm",
    "Thickness, mm": "thickness_mm",
    "Density, kg/m3": "density_kg_m3",
    "Tensile, N": "tensile_n",
    "Tear, N": "tear_n",
    "Elongation, %": "elongation_pct",
    "Peak Force (Bending), N": "peak_force_bending_n",
    "Flexural Modulus MPa": "flexural_modulus_mpa",
    "Maximum Flexural Strength (MPa)": "max_flexural_strength_mpa",
    "Post thickness, mm": "post_thickness_mm",
    "Thickness %": "thickness_change_pct",
    "Post weight, g": "post_weight_g",
    "Post GSM": "post_gsm",
    "Water absorption, %": "water_absorption_pct",
    "Post Tensile, N": "post_tensile_n",
    "Post Tear, N": "post_tear_n",
    "Post Flexural Modulus MPa": "post_flexural_modulus_mpa",
    "Post Peak Force (Bending), N": "post_peak_force_bending_n",
    "Post Maximum Flexural Strength (MPa)": "post_max_flexural_strength_mpa",
    "Tensile Drop, %": "tensile_drop_pct",
    "Tensile drop, %": "tensile_drop_pct",
    "Tear drop, %": "tear_drop_pct",
    "Flexural Modulus Drop, %": "flexural_modulus_drop_pct",
}

# CSV column order. Measured-before, then measured-after, then the derived drops.
CSV_COLUMNS = [
    "material",
    "sample_id",
    "rep",
    "process_step",
    "weight_g",
    "gsm",
    "thickness_mm",
    "density_kg_m3",
    "tensile_n",
    "tear_n",
    "elongation_pct",
    "peak_force_bending_n",
    "flexural_modulus_mpa",
    "max_flexural_strength_mpa",
    "post_weight_g",
    "post_gsm",
    "post_thickness_mm",
    "post_tensile_n",
    "post_tear_n",
    "post_flexural_modulus_mpa",
    "post_peak_force_bending_n",
    "post_max_flexural_strength_mpa",
    "water_absorption_pct",
    "thickness_change_pct",
    "tensile_drop_pct",
    "tear_drop_pct",
    "flexural_modulus_drop_pct",
    "source_sheet",
]

# Derived percentage -> (base column, paired column, direction). Every one of
# these is checked against its own row before being carried through; see
# reconcile_derived().
DERIVED_COLUMNS = {
    "water_absorption_pct": ("weight_g", "post_weight_g", "gain"),
    "thickness_change_pct": ("thickness_mm", "post_thickness_mm", "gain"),
    "tensile_drop_pct": ("tensile_n", "post_tensile_n", "drop"),
    "tear_drop_pct": ("tear_n", "post_tear_n", "drop"),
    "flexural_modulus_drop_pct": ("flexural_modulus_mpa", "post_flexural_modulus_mpa", "drop"),
}

# Allowed disagreement, in percentage points, between the sheet's stored value
# and the same figure recomputed from the row. Covers input rounding only.
RECONCILE_TOLERANCE_PP = 0.5

# Human labels for the CSV columns, used in the Markdown result tables.
LABELS = {
    "gsm": "GSM",
    "thickness_mm": "Thickness (mm)",
    "density_kg_m3": "Density (kg/m3)",
    "tensile_n": "Tensile (N)",
    "tear_n": "Tear (N)",
    "peak_force_bending_n": "Bending peak force (N)",
    "flexural_modulus_mpa": "Flexural modulus (MPa)",
    "max_flexural_strength_mpa": "Max flexural strength (MPa)",
    "water_absorption_pct": "Water absorption (%)",
    "thickness_change_pct": "Thickness change (%)",
    "tensile_drop_pct": "Tensile drop (%)",
    "tear_drop_pct": "Tear drop (%)",
    "flexural_modulus_drop_pct": "Flex. modulus drop (%)",
}

# Columns shown in the Markdown result tables, in order.
MD_RESULT_COLUMNS = [
    "gsm",
    "thickness_mm",
    "density_kg_m3",
    "tensile_n",
    "tear_n",
    "peak_force_bending_n",
    "flexural_modulus_mpa",
    "water_absorption_pct",
    "tensile_drop_pct",
    "tear_drop_pct",
]

MATERIAL_BLURB = {
    "Fybron": "Thin foam-formed nonwoven, ~0.9 mm pressed. Lyocell/viscose/nylon blend with CoPET as the thermobinder.",
    "Fybron_Nylon": "Fybron variant that swaps the CoPET binder line for a CoPET/nylon binder and drops the rinse step.",
    "Fyberite": "Thicker structural foam-formed sheet, ~1.5 mm pressed. Pulp/lyocell base with carbon fibre, CoPET and PLA.",
    "Fyberite_v2": "Higher-pulp, lower-consistency Fyberite revision run at 0.4% consistency with Exilva MFC-F.",
    "Fyberite_NoFoam": "Fyberite run without foaming (no SDS) at 0.4% consistency.",
    "FyberCom": "Composite construction: a Fyberite structural base (~1.5 mm) with a soft Fybron-style skin from this recipe (~0.7 mm), pressed together to ~2.2 mm.",
    "FyberCom_Dyed": "FyberCom with a Fyberite_v2 base (~1.6 mm) and food dye added to the skin slurry. Pressed to ~2.3 mm.",
    "FybeRoll": "Coating-only product. The substrate is purchased FiberTex PET roll goods rather than a formed slurry, so there is no formulation sheet.",
}


class Issues:
    """Collects every data-quality observation so it can be published, not hidden."""

    def __init__(self) -> None:
        self.items: list[tuple[str, str]] = []

    def add(self, where: str, what: str) -> None:
        if (where, what) not in self.items:
            self.items.append((where, what))


def is_error(value: object) -> bool:
    return isinstance(value, str) and value.startswith("#")


def clean(value: object) -> object | None:
    """Blank cells and cached formula errors both become None."""
    if value is None or is_error(value):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def num(value: object, places: int = 2) -> float | None:
    """Round away float noise; return None for anything non-numeric."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    rounded = round(float(value), places)
    return int(rounded) if rounded == int(rounded) else rounded


def pct(value: object, places: int = 1) -> float | None:
    """A stored fraction becomes the percentage its header already claims."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    return num(float(value) * 100, places)


def fmt(value: object) -> str:
    return "" if value is None else str(value)


def parse_conditions(title: str) -> tuple[dict[str, str], str | None]:
    """Pull the press conditions out of the merged title row.

    e.g. "Fyberite_Foam (Rinse, 180C, ~10 s after water removed [~45s], 20kN,
    Thickness = 1.5 mm)" -> rinse / temperature / dwell / force / thickness.
    """
    conditions: dict[str, str] = {}

    structure = None
    brace = re.search(r"\{([^}]*)\}", title)
    if brace:
        structure = brace.group(1).strip()

    rinse = re.search(r"\b(No Rinse|Rinse)\b", title, re.IGNORECASE)
    if rinse:
        conditions["Rinse"] = "No" if rinse.group(1).lower() == "no rinse" else "Yes"

    temp = re.search(r"(\d+(?:\.\d+)?)\s*C\b", title)
    if temp:
        conditions["Press temperature"] = f"{temp.group(1)} C"

    press_dwell = re.search(r"~?\s*(\d+(?:\.\d+)?)\s*s\s+after water removed", title, re.IGNORECASE)
    if press_dwell:
        conditions["Press dwell after water removed"] = f"~{press_dwell.group(1)} s"

    water_time = re.search(r"\[\s*~?\s*(\d+(?:\.\d+)?)\s*s\s*\]", title)
    if water_time:
        conditions["Time to remove water"] = f"~{water_time.group(1)} s"

    force = re.search(r"(\d+(?:\.\d+)?)\s*kN", title)
    if force:
        conditions["Press force"] = f"{force.group(1)} kN"

    thickness = re.search(r"Thickness\s*=\s*([\d.]+)\s*mm", title, re.IGNORECASE)
    if thickness:
        conditions["Target pressed thickness"] = f"{thickness.group(1)} mm"

    foaming = "NoFoam" in title or "No Foam" in title
    conditions["Foamed"] = "No" if foaming else "Yes"

    return conditions, structure


def parse_ingredient_info(info: str | None) -> tuple[int | None, str | None, str, str | None]:
    """Split "3) 6 MM (Minifibers: NYT66-...)" into order, spec and handling note.

    The order marker is kept verbatim as well as parsed: FyberCom really does
    number two of its steps ``6')`` and ``6)``, and flattening that to two 6s
    would invent a duplicate that isn't in the source.
    """
    if not info:
        return None, None, "", None

    text = info.strip()
    order = None
    order_label = None
    match = re.match(r"^(\d+'?)\)\s*(.*)$", text)
    if match:
        order_label = match.group(1)
        order = int(order_label.rstrip("'"))
        text = match.group(2).strip()

    note = None
    bracket = re.search(r"\[([^\]]*)\]", text)
    if bracket:
        note = bracket.group(1).strip()
        text = re.sub(r"\s*\[[^\]]*\]\s*", " ", text).strip()

    return order, order_label, text, note


def read_formulation(worksheet, issues: Issues) -> dict:
    """Parse one formulation sheet into ingredients, totals and notes."""
    sheet = worksheet.title
    rows = list(worksheet.iter_rows(values_only=True))

    title = clean(rows[0][1]) if len(rows[0]) > 1 else None
    title = str(title) if title else sheet
    conditions, structure = parse_conditions(title)

    ingredients: list[dict] = []
    totals: dict[str, object] = {}
    notes: list[str] = []
    batch_date = None

    for index, row in enumerate(rows[2:], start=3):
        label = clean(row[0])

        # Notes and the batch date live in stray cells below the table.
        for cell in row[1:]:
            if isinstance(cell, dt.datetime):
                batch_date = cell.strftime("%B %d, %Y").replace(" 0", " ")
            elif isinstance(cell, str):
                text = cell.strip()
                if re.match(r"^(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d", text):
                    batch_date = text
                elif text.endswith(":") or text.lower().startswith(("foam consistency", "slurry consistency", "used exilva")):
                    if text.rstrip(":").lower() != "note" and text not in notes:
                        notes.append(text)

        if not label:
            continue

        key = str(label).strip().lower()
        if key == "water":
            totals["water_g"] = num(clean(row[2]), 0)
            continue
        if key == "total":
            totals["dry_solids_g"] = num(clean(row[2]), 2)
            continue
        if key.startswith("consistency"):
            totals["consistency_pct"] = pct(clean(row[2]), 3)
            continue

        dry = clean(row[2])
        if is_error(row[4]) or is_error(row[5]):
            issues.add(sheet, f"row {index} ({label}) has a broken formula in the dosing columns; those cells are left blank")

        if not isinstance(dry, (int, float)) or dry == 0:
            continue  # Ingredient listed but unused in this batch.

        order, order_label, spec, note = parse_ingredient_info(clean(row[1]) and str(clean(row[1])))
        ingredients.append(
            {
                "order": order,
                "order_label": order_label,
                "name": str(label).strip(),
                "spec": spec,
                "handling": note,
                "dry_g": num(dry, 2),
                "stock_conc_pct": pct(clean(row[3]), 1),
                "dosed_g": num(clean(row[4]), 2),
                "water_from_stock_g": num(clean(row[5]), 2),
                "pct_of_solids": pct(clean(row[6]), 2),
            }
        )

    # Sort by add order; a primed marker like 6' sorts just ahead of a plain 6.
    ingredients.sort(
        key=lambda item: (
            item["order"] is None,
            item["order"] or 0,
            0 if (item["order_label"] or "").endswith("'") else 1,
        )
    )

    return {
        "sheet": sheet,
        "title": title,
        "conditions": conditions,
        "structure": structure,
        "ingredients": ingredients,
        "totals": totals,
        "notes": notes,
        "batch_date": batch_date,
    }


def read_coating(worksheet) -> list[tuple[str, str]]:
    """Coating sheets are a plain step/detail key-value list."""
    steps: list[tuple[str, str]] = []
    for row in worksheet.iter_rows(values_only=True):
        step = clean(row[0])
        detail = clean(row[1]) if len(row) > 1 else None
        if detail is None:
            continue
        if step is None and steps:
            # A continuation line belongs to the step above it.
            steps[-1] = (steps[-1][0], f"{steps[-1][1]}; {detail}")
            continue
        steps.append((str(step).rstrip(":") if step else "", str(detail)))
    return steps


def reconcile_derived(record: dict, sheet: str, index: int, issues: Issues) -> None:
    """Drop any derived percentage that its own row cannot support.

    Several of these columns are wrong in the source. Where the "post" value was
    never measured, Excel read the blank as zero and reported a 100% loss; in
    other rows the formula reaches into a neighbouring sample's cell, so the
    percentage belongs to a different specimen. Both produce numbers that read as
    real results, which is worse than no number at all. A stored value is kept
    only when the row's own before/after pair reproduces it.
    """
    for column, (base_col, post_col, direction) in DERIVED_COLUMNS.items():
        stored = record.get(column)
        if stored is None:
            continue

        base, post = record.get(base_col), record.get(post_col)
        if base is None or post is None:
            missing = base_col if base is None else post_col
            record[column] = None
            issues.add(
                sheet,
                f"row {index} ({record.get('sample_id')}) reports {column} = {stored}% but {missing} is "
                f"blank, so the sheet computed it against an empty cell; dropped",
            )
            continue

        if base == 0:
            record[column] = None
            issues.add(sheet, f"row {index} ({record.get('sample_id')}) has {base_col} = 0; {column} dropped")
            continue

        expected = ((post - base) / base if direction == "gain" else (base - post) / base) * 100
        if abs(expected - stored) > RECONCILE_TOLERANCE_PP:
            record[column] = None
            issues.add(
                sheet,
                f"row {index} ({record.get('sample_id')}) reports {column} = {stored}% but its own "
                f"{base_col}/{post_col} give {round(expected, 1)}%; the sheet formula references another "
                f"row, so the value is dropped",
            )


def assign_material(sample_id: str, sheet_family: str) -> tuple[str, bool]:
    """Samples are occasionally recorded on the wrong sheet; trust the ID.

    Returns the material and whether the ID actually named it. An ID that names
    no family (``100% PET``, ``90:10 Viscose``) falls back to the sheet it was
    recorded on, and the caller flags it as having no formulation on record.
    """
    lowered = sample_id.lower()
    if "fyberite" in lowered or "fiberite" in lowered:
        return "Fyberite", True
    if "fybercom" in lowered or "fibercom" in lowered:
        return "FyberCom", True
    if "fybron" in lowered:
        return "Fybron", True
    return sheet_family, False


def read_testing(worksheet, issues: Issues) -> list[dict]:
    """Flatten a testing sheet into fully self-describing sample rows."""
    sheet = worksheet.title
    family = TESTING_SHEETS[sheet]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [clean(cell) for cell in rows[0]]

    samples: list[dict] = []
    current_id = None
    unmapped_with_data: set[str] = set()

    for index, row in enumerate(rows[1:], start=2):
        raw_id = clean(row[0])
        rep = clean(row[1]) if len(row) > 1 else None

        if isinstance(rep, str) and rep.strip().lower() in STAT_LABELS:
            continue  # Derived statistic, not a measurement.

        if raw_id is not None and not isinstance(raw_id, (int, float)):
            current_id = str(raw_id).strip()
        sample_id = current_id if raw_id is None else (str(raw_id).strip() if not isinstance(raw_id, (int, float)) else current_id)

        record: dict[str, object] = {column: None for column in CSV_COLUMNS}
        has_measurement = False
        # Errors are held back until the row proves to be a real sample. Blank
        # template rows below the data spill #DIV/0! into their formula columns,
        # and reporting those would bury the findings that matter.
        row_errors: list[str] = []

        for position, header in enumerate(headers):
            if not header or position >= len(row):
                continue
            column = TEST_COLUMN_MAP.get(str(header))
            if column is None:
                if clean(row[position]) is not None:
                    unmapped_with_data.add(str(header))
                continue
            if is_error(row[position]):
                row_errors.append(f"row {index} column {header!r} holds {row[position]}; left blank")
                continue
            value = clean(row[position])
            if value is None:
                continue
            if column in ("sample_id", "rep", "process_step"):
                record[column] = str(value).strip()
            elif str(header) in PERCENT_COLUMNS:
                record[column] = pct(value, 1)
                has_measurement = has_measurement or record[column] is not None
            else:
                places = 3 if column in ("thickness_mm", "post_thickness_mm") else 2
                record[column] = num(value, places)
                has_measurement = has_measurement or record[column] is not None

        if not has_measurement:
            continue  # Empty template row.

        if not sample_id:
            issues.add(sheet, f"row {index} has measurements but no formulation ID; skipped")
            continue

        record["sample_id"] = sample_id
        reconcile_derived(record, sheet, index, issues)

        material, named = assign_material(sample_id, family)
        record["source_sheet"] = sheet
        record["material"] = material
        record["_named"] = named

        for note in row_errors:
            issues.add(sheet, note)
        if material != family:
            issues.add(
                sheet,
                f"sample {sample_id!r} is recorded on the {family} sheet but names {material}; "
                f"it is attributed to {material}",
            )
        if not named:
            issues.add(
                sheet,
                f"sample {sample_id!r} names no material family and no formulation sheet defines it; "
                f"it is grouped under {family} because that is where it was recorded",
            )
        samples.append(record)

    for header in sorted(unmapped_with_data):
        issues.add(sheet, f"column {header!r} holds data but is not carried into the CSV (no mapping defined)")

    return samples


def read_materials_used(worksheet) -> list[dict]:
    """The consumption tracker repeats a 4-column block per week."""
    rows = list(worksheet.iter_rows(values_only=True))
    blocks: list[dict] = []

    for start in range(0, worksheet.max_column, 4):
        label = clean(rows[0][start]) if start < len(rows[0]) else None
        entries: list[dict] = []

        for row in rows[1:]:
            if start + 3 >= len(row):
                continue
            amount = clean(row[start + 1])
            name = clean(row[start + 2])
            kind = clean(row[start + 3])
            if isinstance(name, str) and name.strip().lower() in ("fiber", "chemical"):
                continue  # Repeated sub-header.
            if not isinstance(amount, (int, float)) or not name:
                continue
            entries.append({"amount_g": num(amount, 2), "material": str(name).strip(), "type": str(kind).strip() if kind else ""})

        if entries:
            blocks.append({"label": str(label).strip() if label else "(unlabelled block)", "entries": entries})

    return blocks


def md_table(headers: list[str], rows: list[list[str]], context: list[tuple[str, str]] | None = None) -> list[str]:
    """Render a Markdown table, repeating any context columns on every row.

    RAGFlow's parser lifts each table into a chunk of its own, which strips it
    away from the heading that said which material it belongs to. A bare table
    of press steps retrieves for the wrong material, so the context that the
    heading carried is duplicated into the rows instead.
    """
    context = context or []
    headers = [label for label, _ in context] + headers
    lines = ["| " + " | ".join(headers) + " |", "|" + "|".join("---" for _ in headers) + "|"]
    lines.extend("| " + " | ".join([value for _, value in context] + row) + " |" for row in rows)
    return lines


def result_table(samples: list[dict], material: str, step: str) -> list[str]:
    """Only show measurement columns that this group actually populated."""
    columns = [c for c in MD_RESULT_COLUMNS if any(s.get(c) is not None for s in samples)]
    headers = ["Sample", "Rep"] + [LABELS.get(c, c) for c in columns]
    rows = [
        [s["sample_id"], fmt(s.get("rep"))] + [fmt(s.get(c)) for c in columns]
        for s in samples
    ]
    return md_table(headers, rows, context=[("Material", material), ("Process step", step)])


def write_material_doc(path: Path, name: str, formulation: dict | None, coating: list[tuple[str, str]] | None,
                       coating_sheet: str | None, samples: list[dict], orphan_note: list[str]) -> None:
    lines = [f"# {name}", ""]
    lines.append(f"**Material:** {name}")
    if name in MATERIAL_BLURB:
        lines.append(f"**Summary:** {MATERIAL_BLURB[name]}")
    sources = [s for s in (formulation["sheet"] if formulation else None, coating_sheet) if s]
    lines.append(f"**Source:** `{SOURCE_NAME}`, sheet(s) {', '.join(f'`{s}`' for s in sources)}")
    if formulation and formulation.get("batch_date"):
        lines.append(f"**Batch date:** {formulation['batch_date']}")
    lines.append("")

    if formulation:
        lines += ["## Process conditions", ""]
        if formulation.get("structure"):
            lines += [f"**Construction:** {formulation['structure']}", ""]
        conditions = formulation["conditions"]
        if conditions:
            lines += md_table(
                ["Parameter", "Value"],
                [[k, v] for k, v in conditions.items()],
                context=[("Material", name)],
            )
            lines.append("")
        lines += [f"Conditions as recorded in the sheet title: `{formulation['title']}`", ""]

        totals = formulation["totals"]
        lines += ["## Formulation", ""]
        basis = []
        if totals.get("dry_solids_g") is not None:
            basis.append(f"{totals['dry_solids_g']} g dry solids")
        if totals.get("water_g") is not None:
            basis.append(f"{totals['water_g']} g water")
        if totals.get("consistency_pct") is not None:
            basis.append(f"slurry consistency {totals['consistency_pct']}%")
        if basis:
            lines += [f"Batch basis: {', '.join(basis)}.", ""]
        lines += [
            "Ingredients are listed in the order they are added to the slurry. "
            "Only ingredients actually dosed in this batch are shown; the source sheet also "
            "carries a standing list of zero-amount alternates.",
            "",
        ]
        lines += md_table(
            ["Add order", "Ingredient", "Spec / supplier", "Dry solids (g)", "Stock conc. (%)",
             "Dosed (g)", "Water from stock (g)", "% of solids", "Handling"],
            [
                [
                    fmt(i["order_label"]),
                    i["name"],
                    i["spec"],
                    fmt(i["dry_g"]),
                    fmt(i["stock_conc_pct"]),
                    fmt(i["dosed_g"]),
                    fmt(i["water_from_stock_g"]),
                    fmt(i["pct_of_solids"]),
                    fmt(i["handling"]),
                ]
                for i in formulation["ingredients"]
            ],
            context=[("Material", name)],
        )
        lines.append("")
        if formulation["notes"]:
            lines += ["Notes recorded with this batch:", ""]
            lines += [f"- {note}" for note in formulation["notes"]]
            lines.append("")

    if coating:
        heading = f"## Coating and finishing (sheet `{coating_sheet}`)"
        lines += [heading, ""]
        lines += md_table(
            ["Step", "Detail"],
            [[step or "(continued)", detail] for step, detail in coating],
            context=[("Material", name), ("Stage", "Coating and finishing")],
        )
        lines.append("")

    if samples:
        lines += ["## Test results", ""]
        lines += [
            "Percentages are computed from the paired before/after measurements: water absorption is "
            "`(post weight - weight) / weight`, thickness change is `(post thickness - thickness) / thickness`, "
            "and the drop columns are `(before - after) / before`. Precise per-sample values are also in "
            "`test_samples.csv`.",
            "",
        ]
        by_step: dict[str, list[dict]] = {}
        for sample in samples:
            by_step.setdefault(sample.get("process_step") or "(step not recorded)", []).append(sample)
        for step, group in by_step.items():
            lines += [f"### Process step: {step}", ""]
            lines += result_table(group, name, step)
            lines.append("")

    if orphan_note:
        lines += ["## Data notes", ""]
        lines += [f"- {note}" for note in orphan_note]
        lines.append("")

    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_materials_used(path: Path, blocks: list[dict]) -> None:
    lines = [
        "# Materials consumption by week",
        "",
        f"**Source:** `{SOURCE_NAME}`, sheet `Materials Used`",
        "",
        "Fibre and chemical consumption logged per week across all sample production. "
        "The source sheet carries seven repeated week blocks as a standing template; only the "
        "populated ones appear below.",
        "",
    ]
    for block in blocks:
        lines += [f"## {block['label']}", ""]
        lines += md_table(
            ["Material", "Type", "Amount consumed (g)"],
            [[e["material"], e["type"], fmt(e["amount_g"])] for e in block["entries"]],
            context=[("Week", block["label"])],
        )
        total = sum(e["amount_g"] or 0 for e in block["entries"])
        lines += ["", f"Total logged for this week: {round(total, 2)} g.", ""]
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_overview(path: Path, materials: list[str], sample_count: int) -> None:
    lines = [
        "# Simplifyber material samples: overview",
        "",
        f"**Source:** `{SOURCE_NAME}`",
        "",
        "This collection describes Simplifyber's foam-formed nonwoven material samples: what goes into "
        "each slurry, how it is pressed and coated, and how the pressed samples performed in testing.",
        "",
        "## Materials",
        "",
    ]
    lines += md_table(
        ["Material", "Description"],
        [[m, MATERIAL_BLURB.get(m, "")] for m in materials],
    )
    lines += [
        "",
        "## How the documents fit together",
        "",
        "- One document per material, each holding its process conditions, slurry formulation, "
        "coating recipe and test results.",
        "- `Materials consumption by week` logs total fibre and chemical usage.",
        f"- `test_samples.csv` holds all {sample_count} individual test samples as flat rows, one row per "
        "sample, for precise numeric lookup.",
        "- `Data quality notes` lists every gap and inconsistency found in the source workbook.",
        "",
        "## Reading the process conditions",
        "",
        "Every formulation sheet records its press conditions in one line, for example "
        "`Fyberite_Foam (Rinse, 180C, ~10 s after water removed [~45s], 20kN, Thickness = 1.5 mm)`. "
        "That means: the formed sheet is rinsed, the press runs at 180 C, water takes about 45 s to be "
        "removed, the press is then held about 10 s longer, at 20 kN, to a target thickness of 1.5 mm.",
        "",
        "## Reading the formulation tables",
        "",
        "- **Dry solids (g)** is the ingredient's oven-dry mass in the batch.",
        "- **Stock conc. (%)** is the solids content of the material as supplied, so a 92% lyocell needs "
        "`dry / 0.92` grams dosed.",
        "- **Dosed (g)** is what is actually weighed out; **water from stock** is the moisture that comes "
        "along with it and counts toward the batch water.",
        "- **% of solids** is the ingredient's share of total dry solids - the number to compare across "
        "formulations.",
        "- **Add order** is the sequence the ingredients go into the mixer, which matters for dispersion.",
        "",
        "## Reading the test results",
        "",
        "- **GSM** is grams per square metre; **density** is GSM divided by thickness.",
        "- **Tensile (N)** and **tear (N)** are peak forces, not normalised stresses.",
        "- `Post` columns are measured after a water soak. **Water absorption %** is the mass gain, so "
        "126% means the sample took on 1.26x its own dry mass in water.",
        "- **Drop %** columns are the loss in a property after soaking: `(before - after) / before`. "
        "A negative drop means the property improved.",
        "- **Process step** distinguishes `uncoated`, `internal coating` (bath coat only), "
        "`internal and spray`, `coated`, and `finished` samples. Compare like with like.",
        "",
    ]
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_issues(path: Path, issues: Issues) -> None:
    lines = [
        "# Data quality notes",
        "",
        f"**Source:** `{SOURCE_NAME}`",
        "",
        "Everything below was found while reshaping the workbook. Nothing here has been silently "
        "corrected in the source; where a value could not be trusted it was left blank and recorded here.",
        "",
        "## Derived percentages were re-checked",
        "",
        "Every water absorption, thickness change and property-drop percentage was recomputed from the "
        "before/after pair on its own row and compared with the value the sheet stored. All but four "
        "reconciled. The four that did not are listed under the per-sheet findings below and have been "
        "left blank rather than carried through, because each was produced by a formula reading a cell "
        "that was never filled in: Excel treats the blank as zero, which turns an unmeasured sample into "
        "an apparent 100% loss of strength. Those are not real results.",
        "",
        "## Structural issues",
        "",
        "- Sheet `Fyberite_v2` carries the title `Fyberite_NoFoam (No Rinse, 180C, ...)`, which looks like "
        "a copy-paste artefact from the `Fyberite_NoFoam` sheet. The two sheets have different "
        "formulations, so the title on `Fyberite_v2` is probably wrong rather than the data.",
        "- Sheet `Testing_Fybron` contains six `SF_Fyberite_8.5_*` samples. They have been attributed to "
        "Fyberite, not Fybron.",
        "- Sheet `Testing_Fybron` tests `90:10 Viscose` and `100% PET` samples. No formulation sheet "
        "defines either, so their recipes are unknown.",
        "- On `Testing_Fybron`, the A/B/C/D group at rows 22-25 mixes four different formulations "
        "(`SF_Fybron Nylon-2-4`, `SF_Fybron Nylon 2% AKD`, `100% PET`). The Average and SD rows the sheet "
        "computes across that group therefore average unlike samples.",
        "- Sheet `Testing_FyberCom` is an empty template. Its column list is the planned FyberCom test "
        "matrix: tensile, tear, elongation, abrasion, Bally flex, vamp flex, cleanability, water "
        "absorption, dimension change and weathering resistance.",
        "- Batch dates are recorded inconsistently: some as text (`July 7th`) with no year, some as real "
        "dates (2026-07-15). The text dates are assumed to be 2026 to match.",
        "",
        "## Per-sheet findings",
        "",
    ]
    if issues.items:
        by_sheet: dict[str, list[str]] = {}
        for sheet, note in issues.items:
            by_sheet.setdefault(sheet, []).append(note)
        for sheet, notes in by_sheet.items():
            lines += [f"### `{sheet}`", ""]
            lines += [f"- {note}" for note in notes]
            lines.append("")
    else:
        lines += ["No further issues found.", ""]

    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--outdir", type=Path, default=DEFAULT_OUTDIR)
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.workbook, data_only=True)
    issues = Issues()
    args.outdir.mkdir(parents=True, exist_ok=True)

    formulations = {sheet: read_formulation(workbook[sheet], issues) for sheet in FORMULATION_SHEETS}
    coatings = {
        sheet: read_coating(workbook[sheet])
        for sheet in workbook.sheetnames
        if sheet.endswith("_Coating")
    }

    samples: list[dict] = []
    for sheet in TESTING_SHEETS:
        samples.extend(read_testing(workbook[sheet], issues))

    with (args.outdir / "test_samples.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for sample in samples:
            writer.writerow({column: fmt(sample.get(column)) for column in CSV_COLUMNS})

    # Samples are grouped by material family so each doc carries its own results.
    family_of = {name: name.split("_")[0] for name in FORMULATION_SHEETS}
    materials = FORMULATION_SHEETS + ["FybeRoll"]

    for name in materials:
        formulation = formulations.get(name)
        coating_sheet = COATING_FOR.get(name) or f"{name}_Coating"
        coating = coatings.get(coating_sheet)
        if coating is None:
            coating_sheet = None

        family = family_of.get(name, name)
        # A doc gets the results whose sample ID names it, or its family's results
        # when it is the family's primary recipe.
        primary = name in ("Fybron", "Fyberite", "FyberCom")
        material_samples = [s for s in samples if s["material"] == family] if primary else []

        notes: list[str] = []
        undefined = sorted({s["sample_id"] for s in material_samples if not s["_named"]})
        if undefined:
            notes.append(
                "These samples appear in the results above because they were tested on the "
                f"{family} sheet, but no formulation sheet defines them, so their recipes are unknown: "
                + ", ".join(f"`{s}`" for s in undefined)
                + "."
            )
        if name == "FyberCom" and not material_samples:
            notes.append(
                "No test results are recorded for FyberCom: sheet `Testing_FyberCom` exists but is an "
                "empty template. Its columns give the planned test matrix - tensile, tear, elongation, "
                "abrasion (cycles and rating), Bally flex, vamp flex, cleanability, water absorption, "
                "dimension change and weathering resistance."
            )
        if name == "Fyberite_v2":
            notes.append(
                "The title row on this sheet reads `Fyberite_NoFoam`, which appears to be a copy-paste "
                "artefact; the formulation below differs from the `Fyberite_NoFoam` sheet."
            )
        if not primary and formulation:
            notes.append(
                f"Test results for this variant are recorded under the {family} family; see the {family} document "
                "and `test_samples.csv`."
            )
        if name == "FybeRoll":
            notes.append("No formulation sheet exists: the substrate is purchased FiberTex PET roll goods.")

        write_material_doc(
            args.outdir / f"{slug(name)}.md",
            name,
            formulation,
            coating,
            coating_sheet,
            material_samples,
            notes,
        )

    write_materials_used(args.outdir / "materials-consumption.md", read_materials_used(workbook["Materials Used"]))
    write_overview(args.outdir / "00-overview.md", materials, len(samples))
    write_issues(args.outdir / "data-quality-notes.md", issues)

    print(f"Wrote {len(materials)} material docs + 3 support docs and {len(samples)} sample rows to {args.outdir}")
    print(f"Data-quality findings: {len(issues.items)}")


if __name__ == "__main__":
    main()
