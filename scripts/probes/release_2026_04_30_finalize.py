"""
Finalizer for the 2026-04-30 13-CD release batch.

Polls the 26 PDF/STEP jobs queued via the MCP, then for each CD-number:
  - finds the freshly published CD-XXX.pdf in Vault and saves it (clean, no watermark)
  - finds CD-XXX.stp (or .step) in Vault and saves it
Then:
  - generates a Manufacturing BOM .xlsx
  - combines every CD-XXX.pdf into one file and stamps every page "RELEASED"

Output: ~/Downloads/CD-Release-2026-04-30/
        |- CD-Manufacturing-BOM.xlsx
        |- CD-Combined-RELEASED.pdf
        |- PDFs\\CD-XXX.pdf       (clean individual PDFs)
        \\- STEP\\CD-XXX.stp      (individual STEP files)
"""

import json
import sys
import time
from io import BytesIO
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pypdf import PdfReader, PdfWriter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from pdf_watermark import apply_watermark


CONFIG = json.loads(
    (Path(__file__).resolve().parent.parent / "config.json").read_text("utf-8")
)["vault"]
SERVER = CONFIG["servername"].rstrip("/")
DB = CONFIG["database"]
USER = CONFIG["username"]
PWD = CONFIG["password"]
BASE = f"{SERVER}/AutodeskDM/Services/api/vault/v2"
VAULT_ID = "1"

# (cd, sf, description, type, pdf_job_id, step_job_id)
JOBS = [
    ("CD-001107", "SF-001548", "bolt retainer and hard stops",                "part",     24862, 24875),
    ("CD-001141", "SF-001580", "kft90 lower hot press adapter plate",         "part",     24863, 24876),
    ("CD-001277", "SF-001717", "kft 90 lower heater core adapter plate asm",  "assembly", 24864, 24877),
    ("CD-001308", "SF-001743", "laptop case deckle plate",                    "part",     24865, 24878),
    ("CD-001328", "SF-001729", "kft 90 upper heater core adapter plate asm",  "assembly", 24866, 24879),
    ("CD-001331", "SF-001730", "kft90 upper hot press adapter plate",         "part",     24867, 24880),
    ("CD-001358", "SF-001741", "laptop case vacuum tool tine",                "part",     24868, 24881),
    ("CD-001360", "SF-001767", "20260407-sf01_mold_b",                        "part",     24869, 24882),
    ("CD-001361", "SF-001781", "a cavity smiley-emoji",                       "part",     24870, 24883),
    ("CD-001365", "SF-001737", "laptop case mold cavity",                     "part",     24871, 24884),
    ("CD-001369", "SF-001752", "laptop case mold cavity 2",                   "part",     24872, 24885),
    ("CD-001370", "SF-001753", "laptop case vacuum tool tine 2",              "part",     24873, 24886),
    ("CD-001386", "SF-001768", "a cavity bubbles-doodle",                     "part",     24874, 24887),
]

OUT_DIR = Path.home() / "Downloads" / "CD-Release-2026-04-30"
PDF_DIR = OUT_DIR / "PDFs"
STEP_DIR = OUT_DIR / "STEP"
BOM_PATH = OUT_DIR / "CD-Manufacturing-BOM.xlsx"
COMBINED_PDF = OUT_DIR / "CD-Combined-RELEASED.pdf"

POLL_INTERVAL_S = 20
POLL_TIMEOUT_S = 60 * 90   # 90 min total
LABELS = {0: "Ready", 1: "Running", 2: "Success", 3: "Failure"}


def sign_in(client: httpx.Client) -> dict:
    r = client.post(
        f"{BASE}/sessions",
        json={"input": {"vault": DB, "userName": USER, "password": PWD, "appCode": ""}},
    )
    r.raise_for_status()
    token = r.json()["accessToken"]
    if not token.startswith("Bearer "):
        token = f"Bearer {token}"
    return {"Authorization": token}


def get_job(client, headers, job_id):
    r = client.get(f"{BASE}/vaults/{VAULT_ID}/jobs/{job_id}", headers=headers)
    return r.status_code, (r.json() if r.status_code in (200, 404) else None)


def poll_until_done(client, headers, all_job_ids, timeout_s=POLL_TIMEOUT_S):
    start = time.time()
    final = {}
    pending = set(all_job_ids)
    last_summary = ""
    while pending and time.time() - start < timeout_s:
        for jid in list(pending):
            sc, payload = get_job(client, headers, jid)
            if sc == 404:
                final[jid] = ("Deleted", None); pending.discard(jid)
            elif sc == 200:
                st = payload.get("status")
                if st == 2:
                    final[jid] = ("Success", payload); pending.discard(jid)
                elif st == 3:
                    final[jid] = ("Failure", payload); pending.discard(jid)
        counts = {"Success": 0, "Failure": 0, "Deleted": 0, "Pending": len(pending)}
        for v, _ in final.values():
            counts[v] = counts.get(v, 0) + 1
        msg = f"  [t+{int(time.time()-start)}s] {counts}"
        if msg != last_summary:
            print(msg, flush=True)
            last_summary = msg
        if pending:
            time.sleep(POLL_INTERVAL_S)
    if pending:
        for jid in pending:
            final[jid] = ("Timeout", None)
    return final


def search(client, headers, q, limit=20):
    r = client.get(
        f"{BASE}/vaults/{VAULT_ID}/search-results",
        headers=headers,
        params={"q": q, "limit": limit, "option.latestOnly": "true",
                "option.searchSubFolders": "true"},
    )
    if r.status_code != 200:
        return []
    return r.json().get("results", [])


def find_file(client, headers, cd, exts):
    for r in search(client, headers, cd, limit=20):
        if r.get("entityType") != "FileVersion":
            continue
        nm = r.get("name") or ""
        if not nm.startswith(cd + "."):
            continue
        ext = nm.rsplit(".", 1)[-1].lower()
        if ext in exts:
            return r
    return None


def download_version_content(client, headers, file_version_id) -> bytes:
    h = dict(headers); h.pop("Content-Type", None)
    r = client.get(
        f"{BASE}/vaults/{VAULT_ID}/file-versions/{file_version_id}/content",
        headers=h, follow_redirects=True, timeout=180.0,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"download failed http {r.status_code}: {r.text[:200]}")
    return r.content


def write_bom_xlsx(rows, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Manufacturing BOM"

    headers = [
        "CD-Number", "SF-Number", "Description", "Type", "Revision", "State",
        "PDF Status", "STEP Status",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    widths = [14, 12, 44, 10, 10, 14, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 32

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def build_combined_watermarked(pdf_paths, out_path: Path):
    """Concatenate pdf_paths into one PDF, stamp every page with RELEASED."""
    if not pdf_paths:
        print("  WARN: no PDFs to combine"); return False

    writer = PdfWriter()
    for p in pdf_paths:
        try:
            r = PdfReader(str(p))
            for page in r.pages:
                writer.add_page(page)
        except Exception as e:
            print(f"  WARN: skip {p.name}: {e}")

    buf = BytesIO()
    writer.write(buf)
    combined = buf.getvalue()

    stamped = apply_watermark(
        combined, "RELEASED",
        font_size=140, color="#C0392B", opacity=0.35, rotation=45.0,
    )
    out_path.write_bytes(stamped)
    return True


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    STEP_DIR.mkdir(parents=True, exist_ok=True)

    all_ids = [pdf for *_, pdf, _ in JOBS] + [stp for *_, _, stp in JOBS]
    print(f"Polling {len(all_ids)} jobs (PDF + STEP). interval={POLL_INTERVAL_S}s, timeout={POLL_TIMEOUT_S//60}m\n", flush=True)

    bom_rows = []
    saved_pdfs = []

    with httpx.Client(verify=False, timeout=60.0) as client:
        headers = sign_in(client)
        finals = poll_until_done(client, headers, all_ids)

        print("\n=== Polling complete. Downloading outputs ===\n", flush=True)

        for cd, sf, desc, typ, pdf_id, step_id in JOBS:
            pdf_state, _ = finals.get(pdf_id, ("Unknown", None))
            step_state, _ = finals.get(step_id, ("Unknown", None))
            print(f"\n--- {cd}  PDF={pdf_state}  STEP={step_state} ---", flush=True)

            row = {
                "CD-Number": cd, "SF-Number": sf, "Description": desc,
                "Type": typ, "Revision": "", "State": "",
                "PDF Status": pdf_state, "STEP Status": step_state,
            }

            # PDF
            if pdf_state in ("Success", "Deleted"):
                pdf_rec = find_file(client, headers, cd, ("pdf",))
                if pdf_rec:
                    try:
                        raw = download_version_content(client, headers, pdf_rec["id"])
                        out = PDF_DIR / f"{cd}.pdf"
                        out.write_bytes(raw)
                        saved_pdfs.append(out)
                        row["PDF Status"] = f"OK ({len(raw)//1024} KB)"
                        row["Revision"] = pdf_rec.get("revision", "")
                        row["State"] = pdf_rec.get("state", "")
                        print(f"  PDF  OK  {out.name} ({len(raw)//1024} KB)", flush=True)
                    except Exception as e:
                        row["PDF Status"] = f"download err: {e}"
                        print(f"  PDF  ERR {e}", flush=True)
                else:
                    row["PDF Status"] = "no .pdf in vault"
                    print(f"  PDF  --  no CD-XXX.pdf yet", flush=True)

            # STEP
            if step_state in ("Success", "Deleted"):
                stp_rec = find_file(client, headers, cd, ("stp", "step"))
                if stp_rec:
                    try:
                        raw = download_version_content(client, headers, stp_rec["id"])
                        out = STEP_DIR / stp_rec["name"]
                        out.write_bytes(raw)
                        row["STEP Status"] = f"OK ({len(raw)//1024} KB)"
                        print(f"  STEP OK  {stp_rec['name']} ({len(raw)//1024} KB)", flush=True)
                    except Exception as e:
                        row["STEP Status"] = f"download err: {e}"
                        print(f"  STEP ERR {e}", flush=True)
                else:
                    row["STEP Status"] = "no .stp/.step in vault"
                    print(f"  STEP --  no CD-XXX.stp yet", flush=True)

            bom_rows.append(row)

    # BOM
    write_bom_xlsx(bom_rows, BOM_PATH)
    print(f"\nBOM written -> {BOM_PATH}", flush=True)

    # Combined watermarked PDF (sorted by CD number)
    saved_pdfs_sorted = sorted(saved_pdfs, key=lambda p: p.name)
    if build_combined_watermarked(saved_pdfs_sorted, COMBINED_PDF):
        print(f"Combined+watermarked -> {COMBINED_PDF} "
              f"({COMBINED_PDF.stat().st_size//1024} KB, {len(saved_pdfs_sorted)} files)", flush=True)

    # Summary
    print(f"\n=== Output folder: {OUT_DIR} ===", flush=True)
    pdf_ok = sum(1 for r in bom_rows if r["PDF Status"].startswith("OK"))
    stp_ok = sum(1 for r in bom_rows if r["STEP Status"].startswith("OK"))
    print(f"PDFs saved : {pdf_ok}/{len(JOBS)}", flush=True)
    print(f"STEPs saved: {stp_ok}/{len(JOBS)}", flush=True)


if __name__ == "__main__":
    main()
