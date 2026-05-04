"""
Rewrite the CD-Release-2026-04-30 manufacturing BOM in the green-themed
format used elsewhere in the company:
    ID | SF-Part Number | Quantity | Serial Number | Material |
    File Name (.stp) | Description | Notes

Material comes from each item's Vault properties; quantities are tallied
from MFG-00044-OrderFiles.iam (cd_quantities.json — produced by
count_cd_quantities.py).
"""

from __future__ import annotations

import asyncio
import json
import sys
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

QTY_PATH = Path(__file__).parent / "cd_quantities.json"

CFG = json.loads((PROJECT_ROOT / "config.json").read_text("utf-8"))["vault"]
BASE = CFG["servername"].rstrip("/") + "/AutodeskDM/Services/api/vault/v2"
VAULT_ID = "1"

# Latest .ipt / .iam file-version IDs for each CD (resolved earlier in this session).
MODEL_VERSION_IDS = {
    "CD-001107": "114396", "CD-001141": "114378", "CD-001277": "114362",
    "CD-001308": "114438", "CD-001328": "114369", "CD-001331": "114407",
    "CD-001358": "114439", "CD-001360": "114403", "CD-001361": "114398",
    "CD-001365": "114440", "CD-001369": "114448", "CD-001370": "114446",
    "CD-001386": "114400",
}

# (CD, SF, Description, Type)
ROWS = [
    ("CD-001107", "SF-001548", "bolt retainer and hard stops",                "part"),
    ("CD-001141", "SF-001580", "kft90 lower hot press adapter plate",         "part"),
    ("CD-001277", "SF-001717", "kft 90 lower heater core adapter plate asm",  "assembly"),
    ("CD-001308", "SF-001743", "laptop case deckle plate",                    "part"),
    ("CD-001328", "SF-001729", "kft 90 upper heater core adapter plate asm",  "assembly"),
    ("CD-001331", "SF-001730", "kft90 upper hot press adapter plate",         "part"),
    ("CD-001358", "SF-001741", "laptop case vacuum tool tine",                "part"),
    ("CD-001360", "SF-001767", "laptop case mold side b",                     "part"),
    ("CD-001361", "SF-001781", "a cavity smiley-emoji",                       "part"),
    ("CD-001365", "SF-001737", "laptop case mold cavity",                     "part"),
    ("CD-001369", "SF-001752", "laptop case mold cavity 2",                   "part"),
    ("CD-001370", "SF-001753", "laptop case vacuum tool tine 2",              "part"),
    ("CD-001386", "SF-001768", "a cavity bubbles-doodle",                     "part"),
]

OUT_PATH = Path.home() / "Downloads" / "CD-Release-2026-04-30" / "CD-Manufacturing-BOM.xlsx"

# Style palette (Office green, matches reference sheet)
HEADER_FILL = PatternFill("solid", fgColor="548235")
ALT_ROW_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(color="375623", size=10)
THIN = Side(style="thin", color="A9D08E")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    "ID", "SF-Part Number", "Quantity", "Serial Number",
    "Material", "File Name (.stp)", "Description", "Notes",
]
WIDTHS = [6, 16, 12, 14, 16, 18, 60, 24]


def sign_in(client: httpx.Client) -> dict:
    r = client.post(
        f"{BASE}/sessions",
        json={"input": {"vault": CFG["database"], "userName": CFG["username"],
                        "password": CFG["password"], "appCode": ""}},
    )
    r.raise_for_status()
    tok = r.json()["accessToken"]
    if not tok.startswith("Bearer "):
        tok = f"Bearer {tok}"
    return {"Authorization": tok}


def material_from_file_version(client, headers, file_version_id: str) -> str:
    r = client.get(
        f"{BASE}/vaults/{VAULT_ID}/file-versions/{file_version_id}",
        headers=headers,
        params={"option.includeProperties": "true"},
        timeout=30.0,
    )
    if r.status_code != 200:
        return ""
    for p in (r.json().get("properties") or []):
        defn = p.get("definition") or {}
        if defn.get("systemName") == "Material" or defn.get("displayName") == "Material":
            v = p.get("value")
            if v not in (None, "", "None"):
                return str(v).strip()
    return ""


def fetch_materials() -> dict[str, str]:
    out: dict[str, str] = {}
    with httpx.Client(verify=False, timeout=30.0) as c:
        h = sign_in(c)
        for cd, fv_id in MODEL_VERSION_IDS.items():
            try:
                mat = material_from_file_version(c, h, fv_id)
            except Exception as exc:
                print(f"  WARN: {cd} material lookup failed: {exc}")
                mat = ""
            out[cd] = mat
            print(f"  {cd}: Material = {mat!r}")
    return out


def build_workbook(materials: dict[str, str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Manufacturing BOM"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 28

    quantities = json.loads(QTY_PATH.read_text("utf-8")) if QTY_PATH.exists() else {}

    for idx, (cd, sf, desc, typ) in enumerate(ROWS, start=1):
        n = int(quantities.get(cd, 0) or 0)
        if n > 0:
            qty = f"{n} Each"
            note = ""
        else:
            qty = "1 Each"
            note = "Not in MFG-00044-OrderFiles.iam — defaulted to 1"
        ws.append([
            idx, sf, qty, "N/A",
            materials.get(cd, ""), cd, desc, note,
        ])
        row = ws.max_row
        for col_idx, cell in enumerate(ws[row], start=1):
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx != 7 else "left",
                vertical="center",
                wrap_text=True,
            )
            if row % 2 == 0:
                cell.fill = ALT_ROW_FILL
        ws.row_dimensions[row].height = 20

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Filter dropdowns on every header column
    last_col = chr(64 + len(HEADERS))
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    return wb


def main() -> int:
    print("Fetching Material property from Vault for each CD...")
    materials = fetch_materials()
    wb = build_workbook(materials)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\nWrote: {OUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
