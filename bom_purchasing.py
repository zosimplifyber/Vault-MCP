"""
BOM → Purchasing Sheet logic for the Vault MCP server.

Builds Simplifyber-branded Excel purchasing sheets from Vault BOM data.
Mirrors the logic in the standalone Simplifyber_BOM_Purchasing tool so a
single conversation can take a part number, pull its BOM from Vault, and
emit a formatted purchasing sheet — no manual file export required.
"""

from __future__ import annotations

import glob
import os
import sys
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import purchasing_reference  # Microsoft-List reference source (with Excel fallback)


# ---------------------------------------------------------------------------
# Simplifyber design constants
# ---------------------------------------------------------------------------
DARK_BLUE = "1F3864"
MID_BLUE = "2E75B6"
PALE_BLUE = "EAF3FB"
LIGHT_GRAY = "F2F2F2"
GRAY_BDR = "CCCCCC"
DARK_GRAY = "888888"
WHITE = "FFFFFF"
OLIVE_GREEN = "D8E4BC"
UNMATCHED_FILL = "FCE4D6"  # light orange — a Buy part with no price found


# ---------------------------------------------------------------------------
# SharePoint / OneDrive path to the purchased-items reference file
# ---------------------------------------------------------------------------
ONEDRIVE_SUBFOLDER = os.path.join(
    "OneDrive - simplifyber.com",
    "Shared Documents - Simplifyber",
    "Tech", "Hardware", "Design and Development", "Purchasing",
)
PURCHASED_ITEMS_FILENAME = "purchased items.xlsx"
PURCHASED_ITEMS_SHEET = "purchased parts"


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------
BOM_COLUMNS = [
    "Number", "Row Order", "Position Number", "Item Qty", "Units",
    "Category Name", "Revision", "State", "Title (Item,CO)",
    "Description (Item,CO)", "Source",
]

PURCHASE_COLUMNS = [
    "Material", "Material Finish", "Vendor", "Vendor Number", "Cost Per",
    "HS/HTS Code", "Shipping", "Tax/Tariff", "Sub Total",
    "Lead Time (Business Days)",
]

LOOKUP_COLUMNS = [
    "Material", "Vendor", "Vendor Number", "Cost Per",
    "HS/HTS Code", "Shipping", "Tax/Tariff", "Lead Time (Business Days)",
]

ALL_COLUMNS = BOM_COLUMNS + PURCHASE_COLUMNS

COLUMN_WIDTHS = {
    "Number": 14, "Row Order": 12, "Position Number": 16,
    "Item Qty": 10, "Units": 8, "Category Name": 22, "Revision": 10,
    "State": 14, "Title (Item,CO)": 22, "Description (Item,CO)": 44,
    "Source": 10, "Material": 22, "Material Finish": 18,
    "Vendor": 18, "Vendor Number": 18,
    "Cost Per": 12, "HS/HTS Code": 14, "Shipping": 12, "Tax/Tariff": 12,
    "Sub Total": 14, "Lead Time (Business Days)": 24,
}

VENDOR_COLUMNS = [
    "Vendor", "Number", "Description (Item,CO)", "Material",
    "Vendor Number", "Total Qty", "Unit Cost", "Line Total",
]
VENDOR_COL_WIDTHS = {
    "Vendor": 18, "Number": 14, "Description (Item,CO)": 44, "Material": 22,
    "Vendor Number": 20, "Total Qty": 12, "Unit Cost": 14, "Line Total": 14,
}


# ---------------------------------------------------------------------------
# Vault BOM field-name normalisation
# Vault returns BOM rows with property names that vary by configuration.
# This map collapses each known synonym onto our canonical column name.
# ---------------------------------------------------------------------------
VAULT_FIELD_MAP: dict[str, str] = {
    # Part number
    "Number": "Number",
    "Item Number": "Number",
    "Part Number": "Number",
    "PartNumber": "Number",
    # Row order / BOM structure
    "Row Order": "Row Order",
    "RowOrder": "Row Order",
    "Row": "Row Order",
    "BomRowOrder": "Row Order",
    # Position
    "Position Number": "Position Number",
    "PositionNumber": "Position Number",
    "Position": "Position Number",
    "Pos": "Position Number",
    # Quantity
    "Item Qty": "Item Qty",
    "Qty": "Item Qty",
    "Quantity": "Item Qty",
    "BomQuantity": "Item Qty",
    # Units
    "Units": "Units",
    "Unit": "Units",
    "UOM": "Units",
    # Category
    "Category Name": "Category Name",
    "Category": "Category Name",
    "ItemCategory": "Category Name",
    # Revision
    "Revision": "Revision",
    "Rev": "Revision",
    # State / lifecycle
    "State": "State",
    "LifecycleState": "State",
    "Lifecycle State": "State",
    "Status": "State",
    # Title
    "Title (Item,CO)": "Title (Item,CO)",
    "Title": "Title (Item,CO)",
    "Name": "Title (Item,CO)",
    # Description
    "Description (Item,CO)": "Description (Item,CO)",
    "Description": "Description (Item,CO)",
    "Desc": "Description (Item,CO)",
    # Source / BOM structure type
    "Source": "Source",
    "BOM Structure": "Source",
    "BomStructure": "Source",
    "ItemSource": "Source",
}


# Inventor BOM export headers → our canonical BOM columns (case-insensitive).
INVENTOR_FIELD_MAP: dict[str, str] = {
    "Item": "Row Order",
    "Part Number": "Number",
    "QTY": "Item Qty",
    "Unit QTY": "Units",
    "BOM Structure": "Source",
    "Description": "Description (Item,CO)",
    "REV": "Revision",
    "Material": "Material",
    "Material Finish": "Material Finish",
}

# Inventor "BOM Structure" values → our canonical "Source" values.
SOURCE_VALUE_MAP: dict[str, str] = {
    "Purchased": "Buy",
    "Normal": "Make",
    "Phantom": "Make",
    "Inseparable": "Make",
    "Reference": "Make",
}


def vault_bom_to_dataframe(vault_bom: list[dict[str, Any]]) -> pd.DataFrame:
    """Convert a Vault BOM rows list into a DataFrame matching BOM_COLUMNS.

    Each Vault row is flattened (some Vault responses nest user props inside
    a 'properties' sub-dict) and mapped through VAULT_FIELD_MAP onto our
    canonical column names.
    """
    rows: list[dict[str, Any]] = []
    for item in vault_bom:
        flat: dict[str, Any] = {}
        for k, v in item.items():
            if isinstance(v, dict):
                for kk, vv in v.items():
                    flat[kk] = vv
            else:
                flat[k] = v

        row: dict[str, Any] = {col: None for col in BOM_COLUMNS}
        for vault_key, value in flat.items():
            our_key = VAULT_FIELD_MAP.get(vault_key)
            if our_key and row[our_key] is None:
                row[our_key] = value
        rows.append(row)

    return pd.DataFrame(rows, columns=BOM_COLUMNS)


def extract_bom_list(vault_bom_response: Any) -> list[dict[str, Any]]:
    """Pull the list of BOM rows out of a vault_get_bom_by_part_number response.

    The MCP tool returns a JSON object whose 'bom' field is itself the standard
    Vault REST envelope ({"error": ..., "data": {...}}). The actual rows live
    a few keys deep, and the location varies by endpoint version. This walks
    the common shapes and returns the first list of dicts it finds.
    """
    if isinstance(vault_bom_response, list):
        return [r for r in vault_bom_response if isinstance(r, dict)]

    if not isinstance(vault_bom_response, dict):
        return []

    candidates: list[Any] = []
    bom = vault_bom_response.get("bom")
    if bom is not None:
        candidates.append(bom)
    candidates.extend([
        vault_bom_response.get("rows"),
        vault_bom_response.get("items"),
        vault_bom_response.get("data"),
    ])

    for candidate in candidates:
        if isinstance(candidate, list) and candidate:
            return [r for r in candidate if isinstance(r, dict)]
        if isinstance(candidate, dict):
            inner = candidate.get("data", candidate)
            if isinstance(inner, list) and inner:
                return [r for r in inner if isinstance(r, dict)]
            if isinstance(inner, dict):
                for key in ("results", "items", "itemVersions", "rows", "data", "value"):
                    rows = inner.get(key)
                    if isinstance(rows, list) and rows:
                        return [r for r in rows if isinstance(r, dict)]

    return []


# ---------------------------------------------------------------------------
# Reference file (purchased items.xlsx) discovery & loading
# ---------------------------------------------------------------------------

def find_purchased_items_file() -> str | None:
    """Locate `purchased items.xlsx` in the user's OneDrive Purchasing folder."""
    user_profile = os.environ.get("USERPROFILE", os.path.expanduser("~"))

    exact = os.path.join(user_profile, ONEDRIVE_SUBFOLDER, PURCHASED_ITEMS_FILENAME)
    if os.path.isfile(exact):
        return exact

    purchasing_dir = os.path.join(user_profile, ONEDRIVE_SUBFOLDER)
    if os.path.isdir(purchasing_dir):
        for f in glob.glob(os.path.join(purchasing_dir, "*.xlsx")):
            if "purchased" in os.path.basename(f).lower():
                return f

    if os.path.isdir(user_profile):
        for entry in os.listdir(user_profile):
            if entry.lower().startswith("onedrive") and "simplifyber" in entry.lower():
                candidate = os.path.join(
                    user_profile, entry,
                    "Shared Documents - Simplifyber",
                    "Tech", "Hardware", "Design and Development", "Purchasing",
                    PURCHASED_ITEMS_FILENAME,
                )
                if os.path.isfile(candidate):
                    return candidate

    return None


def load_reference_file(filepath: str) -> pd.DataFrame | None:
    """Load the 'purchased parts' sheet from the reference file."""
    try:
        try:
            df = pd.read_excel(filepath, sheet_name=PURCHASED_ITEMS_SHEET)
        except Exception:
            df = pd.read_excel(filepath, sheet_name=0)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception:
        return None


def lookup_purchased_data(
    bom_df: pd.DataFrame, ref_df: pd.DataFrame
) -> tuple[pd.DataFrame, int, int]:
    """Match BOM rows against the reference file and pre-fill purchasing columns.

    Returns (enriched_df, matched_count, total_buy_count) where matched_count
    is the number of Buy/Other rows that received a Vendor from the lookup.
    """
    result = bom_df.copy()
    ref_key = next(
        (c for c in ("Number", "Part Number", "Item Number", "PartNumber")
         if c in ref_df.columns),
        None,
    )
    if ref_key is None:
        return result, 0, 0

    available = [c for c in LOOKUP_COLUMNS if c in ref_df.columns]

    # Reference workbooks frequently have a part number listed on multiple rows
    # (e.g. one per vendor quote). The previous code did `ref_df.set_index(ref_key)`
    # without deduping, which made `ref_idx.at[num, c]` return a *Series* for any
    # duplicated number — and the surrounding `pd.notna(...)` then raised
    # "truth value of a Series is ambiguous". Drop duplicates keeping the first
    # occurrence (matches the historical "lookup first match" intent), then
    # build a {number: value} dict per column and use pandas' vectorized map.
    ref_unique = ref_df.drop_duplicates(subset=[ref_key], keep="first")
    for col in available:
        col_map = dict(zip(ref_unique[ref_key], ref_unique[col]))
        # `Series.map(dict)` returns NaN for missing keys and for cells whose
        # source value is NaN — same semantics as the prior None branch from
        # downstream `.isna()` / `.notna()` checks' perspective.
        result[col] = result["Number"].map(col_map)

    buy_mask = result["Source"].isin(["Buy", "Other"])
    matched = int(result[buy_mask]["Vendor"].notna().sum())
    total_buy = int(buy_mask.sum())
    return result, matched, total_buy


# ---------------------------------------------------------------------------
# Hierarchy parser — turns the dotted "Row Order" column into a parent→children map
# ---------------------------------------------------------------------------

def build_children_map(df: pd.DataFrame) -> dict[int, list[int]]:
    """Walk the Row Order column once and record each row's direct children.

    A row's depth is its number of dotted segments ("1.2.3" → depth 3); the
    root assembly's "-" (or empty) maps to depth 0. We use a stack of
    (depth, row_index) pairs: before adding row N, pop everything at the
    same depth or deeper, so the top of the stack is row N's parent.
    """
    def depth(row_order: Any) -> int:
        s = str(row_order).strip()
        return 0 if s in ("-", "nan", "") else s.count(".") + 1

    depths = [depth(df.iloc[i]["Row Order"]) for i in range(len(df))]
    children: dict[int, list[int]] = {i: [] for i in range(len(df))}
    stack: list[tuple[int, int]] = []

    for i, d in enumerate(depths):
        while stack and stack[-1][0] >= d:
            stack.pop()
        if stack:
            children[stack[-1][1]].append(i)
        stack.append((d, i))
    return children


# ---------------------------------------------------------------------------
# Excel sheet builders
# ---------------------------------------------------------------------------

def _border(color: str = GRAY_BDR) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _resource_path(filename: str) -> str:
    """Return absolute path to a bundled resource (logo image, etc.)."""
    if getattr(sys, "frozen", False):
        base = sys._MEIPASS  # type: ignore[attr-defined]
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, filename)


def _add_logo(ws, logo_filename: str = "Simplifyber_Logo_White.png") -> None:
    logo_path = _resource_path(logo_filename)
    if not os.path.isfile(logo_path):
        return
    try:
        xl_img = XLImage(logo_path)
        target_h = 38
        orig_w, orig_h = xl_img.width, xl_img.height
        xl_img.height = target_h
        xl_img.width = int(target_h * orig_w / orig_h)
        try:
            from openpyxl.drawing.spreadsheet_drawing import (
                AnchorMarker,
                OneCellAnchor,
            )
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            EMU = 9525  # EMUs per pixel at 96 dpi
            marker = AnchorMarker(col=0, colOff=10 * EMU, row=0, rowOff=6 * EMU)
            size = XDRPositiveSize2D(cx=xl_img.width * EMU, cy=xl_img.height * EMU)
            xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(xl_img)
        except Exception:
            ws.add_image(xl_img, "A1")
    except Exception:
        pass


def build_purchasing_sheet(
    df: pd.DataFrame, output_path: str, assembly_number: str
) -> str:
    """Write the formatted purchasing workbook (Purchasing + By Vendor tabs)."""
    df = df.reset_index(drop=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchasing"
    n_cols = len(ALL_COLUMNS)

    # Title bar
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = assembly_number
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50
    _add_logo(ws)

    # Date row
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    c.font = Font(name="Arial", size=9, color=DARK_GRAY, italic=True)
    c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 16

    # Column header row
    HDR_ROW = 3
    hdr_font = Font(name="Arial", bold=True, color=WHITE, size=10)
    hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr = _border()
    for ci, name in enumerate(ALL_COLUMNS, 1):
        c = ws.cell(row=HDR_ROW, column=ci, value=name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = bdr
    ws.row_dimensions[HDR_ROW].height = 36

    # Data rows
    body_font = Font(name="Arial", size=10)
    white_fill = PatternFill("solid", fgColor=WHITE)
    alt_fill = PatternFill("solid", fgColor=PALE_BLUE)
    olive_fill = PatternFill("solid", fgColor=OLIVE_GREEN)
    body_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    qty_col = get_column_letter(ALL_COLUMNS.index("Item Qty") + 1)
    cost_col = get_column_letter(ALL_COLUMNS.index("Cost Per") + 1)
    ship_col = get_column_letter(ALL_COLUMNS.index("Shipping") + 1)
    tax_col = get_column_letter(ALL_COLUMNS.index("Tax/Tariff") + 1)
    sub_col = get_column_letter(ALL_COLUMNS.index("Sub Total") + 1)

    children_map = build_children_map(df)
    FIRST_DATA_ROW = HDR_ROW + 1
    DOLLAR_FMT = '"$"#,##0.00'
    DOLLAR_COLS = {"Cost Per", "Shipping", "Tax/Tariff", "Sub Total"}
    unmatched_nums: list[str] = []

    for ri, (df_idx, row) in enumerate(df.iterrows(), start=FIRST_DATA_ROW):
        category = str(row.get("Category Name", "")).strip().lower()
        child_indices = children_map.get(df_idx, [])
        # Inventor exports carry no Category Name — a row with children IS an
        # assembly, so detect by structure as well as category.
        is_assembly = ("assembly" in category) or bool(child_indices)

        if is_assembly:
            fill = olive_fill
        else:
            fill = alt_fill if (ri - HDR_ROW) % 2 == 0 else white_fill

        vendor = row.get("Vendor")
        vendor_blank = pd.isna(vendor) or str(vendor).strip() == ""
        is_unmatched = (
            not is_assembly
            and str(row.get("Source", "")).strip() in {"Buy", "Other"}
            and vendor_blank
        )
        if is_unmatched:
            fill = PatternFill("solid", fgColor=UNMATCHED_FILL)
            num_val = row.get("Number")
            if pd.notna(num_val):
                num_str = str(num_val)
                if num_str not in unmatched_nums:
                    unmatched_nums.append(num_str)

        for ci, col_name in enumerate(ALL_COLUMNS, 1):
            c = ws.cell(row=ri, column=ci)

            if col_name in BOM_COLUMNS:
                val = row.get(col_name)
                c.value = None if pd.isna(val) else val

            elif col_name == "Material Finish":
                val = row.get("Material Finish")
                c.value = None if pd.isna(val) else val

            elif col_name in LOOKUP_COLUMNS:
                val = row.get(col_name)
                c.value = None if pd.isna(val) else val
                # Assembly Cost Per = SUM of children's Sub Totals
                if col_name == "Cost Per" and is_assembly and child_indices:
                    child_cells = ",".join(
                        f"{sub_col}{FIRST_DATA_ROW + ci_}" for ci_ in child_indices
                    )
                    c.value = f"=SUM({child_cells})"

            elif col_name == "Sub Total":
                if not is_assembly:
                    c.value = (
                        f"=({cost_col}{ri}*{qty_col}{ri})"
                        f"+{ship_col}{ri}+{tax_col}{ri}"
                    )
                elif child_indices:
                    # Root assembly Qty often = "-"; ISNUMBER guard treats it as 1
                    c.value = (
                        f"={cost_col}{ri}"
                        f"*IF(ISNUMBER({qty_col}{ri}),{qty_col}{ri},1)"
                    )

            if col_name in DOLLAR_COLS:
                c.number_format = DOLLAR_FMT

            c.font = body_font
            c.fill = fill
            c.border = bdr
            c.alignment = body_align

    # Unmatched note — list Buy parts with no reference match so a $0 line is
    # clearly "no price found", not "free".
    if unmatched_nums:
        note_row = FIRST_DATA_ROW + len(df) + 1
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=n_cols)
        nc = ws.cell(row=note_row, column=1)
        nc.value = (f"Unmatched ({len(unmatched_nums)}) — no price in reference: "
                    + ", ".join(unmatched_nums))
        nc.font = Font(name="Arial", size=9, italic=True, color=DARK_GRAY)
        nc.fill = PatternFill("solid", fgColor=UNMATCHED_FILL)
        nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column widths & freeze
    for ci, col_name in enumerate(ALL_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COLUMN_WIDTHS.get(col_name, 15)
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    # Mid-blue accent border on top of header row
    for ci in range(1, n_cols + 1):
        c = ws.cell(row=HDR_ROW, column=ci)
        c.border = Border(
            top=Side(style="medium", color=MID_BLUE),
            left=Side(style="thin", color=GRAY_BDR),
            right=Side(style="thin", color=GRAY_BDR),
            bottom=Side(style="thin", color=GRAY_BDR),
        )

    _build_vendor_tab(wb, df)
    _build_assembly_costs_tab(wb, df, children_map, FIRST_DATA_ROW)
    wb.save(output_path)
    return output_path


def _build_vendor_tab(wb: Workbook, df: pd.DataFrame) -> None:
    """Add the 'By Vendor' tab — a flat, filterable, vendor-grouped table."""
    ws = wb.create_sheet("By Vendor")
    last_col = get_column_letter(len(VENDOR_COLUMNS))
    bdr = _border()
    DOLLAR_FMT = '"$"#,##0.00'

    # Title
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "By Vendor"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 50
    _add_logo(ws)

    # Headers
    HDR = 2
    for ci, name in enumerate(VENDOR_COLUMNS, 1):
        c = ws.cell(row=HDR, column=ci, value=name)
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    ws.row_dimensions[HDR].height = 36

    # Aggregate Buy/Other parts by vendor
    parts = df[df["Source"].apply(lambda s: str(s).strip() in {"Buy", "Other"})].copy()
    for col in ("Vendor", "Vendor Number", "Material", "Cost Per"):
        if col not in parts.columns:
            parts[col] = None
    parts = parts[parts["Vendor"].notna() & (parts["Vendor"].astype(str).str.strip() != "")]

    def safe_qty(q: Any) -> float:
        try:
            return float(q)
        except (TypeError, ValueError):
            return 0.0

    parts["_qty_num"] = parts["Item Qty"].apply(safe_qty)
    parts["_cost_num"] = pd.to_numeric(parts["Cost Per"], errors="coerce").fillna(0)

    agg = (
        parts.groupby(
            ["Vendor", "Number", "Description (Item,CO)", "Material",
             "Vendor Number", "_cost_num"],
            dropna=False, sort=False,
        )
        .agg(total_qty=("_qty_num", "sum"))
        .reset_index()
        .sort_values(["Vendor", "Number"])
        .reset_index(drop=True)
    )

    qty_col = get_column_letter(VENDOR_COLUMNS.index("Total Qty") + 1)
    unit_cost_col = get_column_letter(VENDOR_COLUMNS.index("Unit Cost") + 1)
    white_fill = PatternFill("solid", fgColor=WHITE)
    alt_fill = PatternFill("solid", fgColor=PALE_BLUE)

    for i, (_, row) in enumerate(agg.iterrows()):
        ri = HDR + 1 + i
        fill = alt_fill if i % 2 == 0 else white_fill
        qty = row["total_qty"]
        vals = [
            row["Vendor"],
            row["Number"],
            row["Description (Item,CO)"],
            row["Material"] if pd.notna(row.get("Material")) else None,
            row["Vendor Number"] if pd.notna(row.get("Vendor Number")) else None,
            int(qty) if qty == int(qty) else qty,
            row["_cost_num"] if row["_cost_num"] != 0 else None,
            f"={unit_cost_col}{ri}*{qty_col}{ri}",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.border = bdr
            c.alignment = Alignment(horizontal="left", vertical="center")
            if VENDOR_COLUMNS[ci - 1] in ("Unit Cost", "Line Total"):
                c.number_format = DOLLAR_FMT

    last_data_row = HDR + len(agg)
    ws.auto_filter.ref = f"A{HDR}:{last_col}{last_data_row}"

    for ci, col_name in enumerate(VENDOR_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = VENDOR_COL_WIDTHS.get(col_name, 15)
    ws.freeze_panes = "A3"

    for ci in range(1, len(VENDOR_COLUMNS) + 1):
        c = ws.cell(row=HDR, column=ci)
        c.border = Border(
            top=Side(style="medium", color=MID_BLUE),
            left=Side(style="thin", color=GRAY_BDR),
            right=Side(style="thin", color=GRAY_BDR),
            bottom=Side(style="thin", color=GRAY_BDR),
        )


def _build_assembly_costs_tab(
    wb: Workbook, df: pd.DataFrame, children_map: dict, first_data_row: int,
) -> None:
    """Add the 'Assembly Costs' summary — each sub-assembly's cost-to-make-one
    (references the Purchasing sheet's roll-up formulas) plus a grand total for
    the whole build."""
    all_children = {ci for kids in children_map.values() for ci in kids}
    assemblies = [i for i in range(len(df)) if children_map.get(i)]
    top_level = [i for i in range(len(df)) if i not in all_children]

    def _subtree_incomplete(i: int) -> bool:
        kids = children_map.get(i, [])
        if not kids:  # leaf
            return pd.isna(df.iloc[i].get("Cost Per"))
        return any(_subtree_incomplete(c) for c in kids)

    has_incomplete = False

    cost_col = get_column_letter(ALL_COLUMNS.index("Cost Per") + 1)
    sub_col = get_column_letter(ALL_COLUMNS.index("Sub Total") + 1)

    ws = wb.create_sheet("Assembly Costs")
    bdr = _border()
    DOLLAR_FMT = '"$"#,##0.00'
    cols = ["Item", "Part #", "Description", "Cost to Make One"]
    last_col = get_column_letter(len(cols))

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "Assembly Costs"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    _add_logo(ws)

    HDR = 2
    for ci, name in enumerate(cols, 1):
        cc = ws.cell(row=HDR, column=ci, value=name)
        cc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cc.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = bdr
    ws.row_dimensions[HDR].height = 36

    alt_fill = PatternFill("solid", fgColor=PALE_BLUE)
    white_fill = PatternFill("solid", fgColor=WHITE)
    for n, i in enumerate(assemblies):
        ri = HDR + 1 + n
        row = df.iloc[i]
        purch_row = first_data_row + i
        incomplete = _subtree_incomplete(i)
        if incomplete:
            has_incomplete = True
        desc = row.get("Description (Item,CO)")
        desc = "" if pd.isna(desc) else str(desc)
        if incomplete:
            desc = (desc + " *").strip()
        vals = [
            None if pd.isna(row.get("Row Order")) else row.get("Row Order"),
            None if pd.isna(row.get("Number")) else row.get("Number"),
            desc if desc else None,
            f"=Purchasing!{cost_col}{purch_row}",   # cost to make one
        ]
        for ci, val in enumerate(vals, 1):
            cc = ws.cell(row=ri, column=ci, value=val)
            cc.font = Font(name="Arial", size=10)
            cc.fill = alt_fill if n % 2 == 0 else white_fill
            cc.border = bdr
            cc.alignment = Alignment(horizontal="left", vertical="center")
            if ci == 4:
                cc.number_format = DOLLAR_FMT

    total_row = HDR + 1 + len(assemblies)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    build_incomplete = any(_subtree_incomplete(i) for i in top_level)
    tc = ws.cell(row=total_row, column=1,
                 value="GRAND TOTAL — whole build" + (" *" if build_incomplete else ""))
    tc.font = Font(name="Arial", bold=True, color=DARK_BLUE, size=10)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    grand = ws.cell(row=total_row, column=4)
    if top_level:
        refs = "+".join(f"Purchasing!{sub_col}{first_data_row + i}" for i in top_level)
        grand.value = f"={refs}"
    else:
        grand.value = 0
    grand.number_format = DOLLAR_FMT
    grand.font = Font(name="Arial", bold=True, color=DARK_BLUE, size=10)

    if has_incomplete or build_incomplete:
        foot_row = total_row + 1
        ws.merge_cells(start_row=foot_row, start_column=1,
                       end_row=foot_row, end_column=len(cols))
        fc = ws.cell(row=foot_row, column=1,
                     value="* Includes unpriced parts — this total may be understated.")
        fc.font = Font(name="Arial", size=9, italic=True, color=DARK_GRAY)
        fc.alignment = Alignment(horizontal="left", vertical="center")

    for ci in range(1, len(cols) + 1):
        hc = ws.cell(row=HDR, column=ci)
        hc.border = Border(
            top=Side(style="medium", color=MID_BLUE),
            left=Side(style="thin", color=GRAY_BDR),
            right=Side(style="thin", color=GRAY_BDR),
            bottom=Side(style="thin", color=GRAY_BDR),
        )

    widths = {"Item": 12, "Part #": 16, "Description": 46, "Cost to Make One": 18}
    for ci, name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths[name]
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# High-level orchestrators used by the MCP tools
# ---------------------------------------------------------------------------

def _enrich_with_reference(df: pd.DataFrame, reference_path: str = "") -> tuple[pd.DataFrame, int, int, list[str], list[str]]:
    """Try to fill purchasing columns from the reference file.

    Returns (df, matched, total, unmatched_part_numbers, warnings).
    """
    warnings: list[str] = []
    matched = total = 0
    unmatched: list[str] = []

    ref_df = None
    # 1) An explicit Excel file always wins.
    if reference_path and os.path.isfile(reference_path):
        ref_df = load_reference_file(reference_path)
        if ref_df is None:
            warnings.append(f"Found reference file at {reference_path} but could not read it.")
    else:
        # Wrapped so reference resolution can NEVER break sheet generation — any
        # failure just leaves the purchasing columns blank with a warning.
        try:
            cfg = purchasing_reference.resolve_reference_config()
            # 2) Microsoft List, when configured (and someone has signed in).
            if cfg.get("source") in ("auto", "mslist") and purchasing_reference.mslist_is_configured(cfg):
                try:
                    ref_df = purchasing_reference.load_mslist_dataframe(
                        cfg["mslist"], cfg.get("column_map"))
                except Exception as exc:  # noqa: BLE001
                    warnings.append(
                        f"Microsoft List reference unavailable ({exc}); using the Excel file.")
                    ref_df = None
            # 3) Excel auto-discovery (unless the source is pinned to 'mslist').
            if ref_df is None and cfg.get("source") != "mslist":
                path = find_purchased_items_file()
                if path:
                    ref_df = load_reference_file(path)
                    if ref_df is None:
                        warnings.append(f"Found reference file at {path} but could not read it.")
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"Could not resolve a purchasing reference ({exc}); columns left blank.")
            ref_df = None

    if ref_df is None:
        warnings.append(
            "Purchased items reference not found. Ensure OneDrive is syncing the "
            "Purchasing folder (or sign in to the Microsoft List), or the columns "
            "Material/Vendor/Cost Per will be left blank."
        )
        return df, matched, total, unmatched, warnings

    df, matched, total = lookup_purchased_data(df, ref_df)
    buy_mask = df["Source"].isin(["Buy", "Other"])
    unmatched = df[buy_mask & df["Vendor"].isna()]["Number"].dropna().astype(str).tolist()
    return df, matched, total, unmatched, warnings


def generate_from_vault_bom(
    vault_bom_response: Any,
    assembly_number: str,
    output_dir: str = "",
) -> dict[str, Any]:
    """Build a purchasing sheet from a vault_get_bom_by_part_number response."""
    bom_list = extract_bom_list(vault_bom_response)
    if not bom_list:
        return {
            "error": True,
            "message": (
                "No BOM rows found in vault_bom_response. Pass the full JSON "
                "object returned by vault_get_bom_by_part_number, including "
                "the 'bom' key."
            ),
        }

    df = vault_bom_to_dataframe(bom_list)

    warnings: list[str] = []
    if df["Number"].isna().all():
        warnings.append(
            "Could not map the 'Number' column from Vault BOM data. "
            "Check VAULT_FIELD_MAP for missing field-name aliases."
        )
    if df["Item Qty"].isna().all():
        warnings.append(
            "Could not map the 'Item Qty' column from Vault BOM data. "
            "Sub-Total formulas will evaluate to zero."
        )

    df, matched, total, unmatched, ref_warnings = _enrich_with_reference(df)
    warnings.extend(ref_warnings)

    if not output_dir:
        output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"{assembly_number}-PurchasingExport.xlsx")

    build_purchasing_sheet(df, out_file, assembly_number)

    return {
        "output_path": out_file,
        "matched_parts": matched,
        "total_purchased_parts": total,
        "unmatched_parts": unmatched,
        "warnings": warnings,
    }


def coerce_bom_dataframe(
    df: pd.DataFrame,
) -> tuple[pd.DataFrame, str | None]:
    """Normalize a raw BOM DataFrame (Vault-canonical OR Inventor export).

    Returns (normalized_df, error_message). error_message is None on success.
    Guarantees the result carries every BOM_COLUMNS entry plus 'Material' and
    'Material Finish' (missing ones filled with None), a positional index, and
    translated Source values. Errors if the critical Number/Item Qty are absent.
    """
    df = df.rename(columns={c: str(c).strip() for c in df.columns})

    if not set(BOM_COLUMNS).issubset(df.columns):
        # Inventor export — map its headers case-insensitively.
        lower = {c.lower(): c for c in df.columns}
        rename: dict[str, str] = {}
        for inv_name, canon in INVENTOR_FIELD_MAP.items():
            src = lower.get(inv_name.lower())
            if (src is not None
                    and canon not in rename.values()
                    and canon not in df.columns):
                rename[src] = canon
        df = df.rename(columns=rename)
        if "Source" in df.columns:
            df["Source"] = df["Source"].map(
                lambda v: SOURCE_VALUE_MAP.get(str(v).strip(), v)
            )

    for col in BOM_COLUMNS + ["Material", "Material Finish"]:
        if col not in df.columns:
            df[col] = None

    df = df.reset_index(drop=True)

    # Row Order can arrive as float64: pandas infers a column of "1"/"2.1"
    # numeric strings as floats when the BOM has no multi-level (x.y.z) rows,
    # which corrupts the dotted-depth hierarchy parser ("14" -> "14.0", landing
    # a parent at the same depth as its "14.1" child). Normalize to clean strings.
    def _clean_row_order(v):
        if pd.isna(v):
            return v
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)
    df["Row Order"] = df["Row Order"].map(_clean_row_order)

    blank_number = df["Number"].isna() | (df["Number"].astype(str).str.strip() == "")
    if blank_number.all():
        return df, ("No part numbers found — the BOM needs a 'Part Number' "
                    "(Inventor) or 'Number' (Vault) column.")
    blank_qty = df["Item Qty"].isna() | (df["Item Qty"].astype(str).str.strip() == "")
    if blank_qty.all():
        return df, ("No quantities found — the BOM needs a 'QTY' (Inventor) "
                    "or 'Item Qty' (Vault) column.")
    return df, None


def read_bom_file(bom_file_path: str) -> pd.DataFrame:
    """Read a BOM export into a DataFrame by extension.

    .csv -> comma; .txt -> tab-delimited (Inventor's text export); .xls/.xlsx ->
    first sheet. Identity/hierarchy columns are forced to string so a
    numeric-looking value ("14", "2.10", an all-numeric part number) is never
    coerced to float — float coercion collapses "2.10" and "2.1" to the same
    value and breaks the dotted-depth BOM hierarchy parser. Raises ValueError on
    an unsupported extension.
    """
    str_cols = {c: str for c in
                ("Item", "Row Order", "Position Number", "Part Number", "Number")}
    ext = os.path.splitext(bom_file_path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(bom_file_path, dtype=str_cols)
    if ext == ".txt":
        return pd.read_csv(bom_file_path, sep="\t", dtype=str_cols)
    if ext in (".xls", ".xlsx"):
        return pd.read_excel(bom_file_path, sheet_name=0, dtype=str_cols)
    raise ValueError(f"Unsupported file type: {ext}. Use .xlsx, .xls, .csv, or .txt.")


def generate_from_file(
    bom_file_path: str,
    assembly_number: str,
    output_dir: str = "",
    reference_path: str = "",
) -> dict[str, Any]:
    """Build a purchasing sheet from an exported BOM file.

    Accepts a Vault BOM export (canonical columns) or an Inventor BOM export
    (auto-detected + header-mapped). Supports .xlsx/.xls/.csv/.txt(tab).
    """
    if not os.path.isfile(bom_file_path):
        return {"error": True, "message": f"BOM file not found: {bom_file_path}"}

    try:
        raw = read_bom_file(bom_file_path)
    except ValueError as exc:
        return {"error": True, "message": str(exc)}
    except Exception as exc:  # noqa: BLE001
        return {"error": True, "message": f"Could not read BOM file: {exc}"}

    df, err = coerce_bom_dataframe(raw)
    if err:
        return {"error": True, "message": err}

    # Material precedence: the export's Material wins where non-blank; the
    # reference file only fills blanks.
    export_material = df["Material"].copy()
    df, matched, total, unmatched, warnings = _enrich_with_reference(df, reference_path=reference_path)
    keep = export_material.notna() & (export_material.astype(str).str.strip() != "")
    df.loc[keep, "Material"] = export_material[keep]

    if not output_dir:
        output_dir = os.path.dirname(bom_file_path)
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"{assembly_number}-PurchasingExport.xlsx")

    build_purchasing_sheet(df, out_file, assembly_number)

    return {
        "output_path": out_file,
        "matched_parts": matched,
        "total_purchased_parts": total,
        "unmatched_parts": unmatched,
        "warnings": warnings,
    }


def lookup_part(part_number: str) -> dict[str, Any]:
    """Look up purchasing data for a single part number in the reference file."""
    ref_path = find_purchased_items_file()
    if not ref_path:
        return {
            "found": False,
            "note": (
                "Purchased items reference file not found. "
                "Check that OneDrive is syncing the Purchasing folder."
            ),
        }

    ref_df = load_reference_file(ref_path)
    if ref_df is None:
        return {"found": False, "note": f"Could not read reference file at {ref_path}."}

    ref_key = next(
        (c for c in ("Number", "Part Number", "Item Number", "PartNumber")
         if c in ref_df.columns),
        None,
    )
    if ref_key is None:
        return {
            "found": False,
            "note": "Reference file has no recognisable part-number column.",
        }

    match = ref_df[ref_df[ref_key].astype(str).str.strip() == str(part_number).strip()]
    if match.empty:
        return {
            "found": False,
            "note": f"Part number '{part_number}' not found in reference file.",
        }

    row = match.iloc[0]
    result: dict[str, Any] = {
        col: (None if pd.isna(row[col]) else row[col])
        for col in LOOKUP_COLUMNS
        if col in row.index
    }
    result["found"] = True
    result["part_number"] = str(part_number).strip()
    return result


def reference_file_status() -> dict[str, Any]:
    """Diagnose whether the reference file is locatable and readable."""
    ref_path = find_purchased_items_file()
    if not ref_path:
        user_profile = os.environ.get("USERPROFILE", os.path.expanduser("~"))
        expected = os.path.join(user_profile, ONEDRIVE_SUBFOLDER, PURCHASED_ITEMS_FILENAME)
        return {
            "found": False,
            "expected": expected,
            "note": "File not found. Make sure OneDrive is syncing the Purchasing folder.",
        }

    ref_df = load_reference_file(ref_path)
    return {
        "found": True,
        "path": ref_path,
        "part_count": len(ref_df) if ref_df is not None else 0,
        "columns": list(ref_df.columns) if ref_df is not None else [],
    }
